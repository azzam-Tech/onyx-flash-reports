import os
import io
import json
import csv
from urllib.parse import urlencode
from datetime import datetime
from flask import Flask, request, render_template, Response, session, redirect, jsonify
from database import get_conn
from reports_config import TABS, TABMAP, find_report, run_report, lookups, jv_options, compute_dash, save_hidden, load_hidden, load_hide_profit, load_hidden_raw

app = Flask(__name__)
app.secret_key = os.environ.get('SREEN_SECRET', 'sreen-reports-2026-secret-key')
SETTINGS_PIN = os.environ.get("SETTINGS_PIN", "00900")
APP_PIN = os.environ.get("APP_PIN", "00900")

@app.before_request
def require_login():
    if request.endpoint not in ('login', 'static') and not session.get('logged_in'):
        return redirect('/login')

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if request.form.get("pin") == APP_PIN:
            session['logged_in'] = True
            return redirect('/')
        else:
            error = "الرمز غير صحيح، حاول مرة أخرى."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.pop('logged_in', None)
    return redirect('/login')

@app.route("/")
def index():
    hidden_tabs, hidden_reports = load_hidden()
    _vis = [t for t in TABS if t["id"] not in hidden_tabs] or TABS
    cur_tab = request.args.get("tab", _vis[0]["id"])
    rid = request.args.get("report", "")
    tab, rpt = find_report(cur_tab, rid)
    cur_tab = tab["id"]
    for _p in rpt["params"]:
        if _p.get("dynamic") == "jv": _p["options"] = jv_options()
        if _p["name"] in ("rep_code","c_code","v_code","i_code","a_code"): _p["_list"] = lookups(_p["name"])
    display = {p["name"]: request.args.get(p["name"], p.get("default","")) for p in rpt["params"]}
    qsp = {"tab": cur_tab, "report": rpt["id"]}
    for p in rpt["params"]: qsp[p["name"]] = request.args.get(p["name"], p.get("default",""))
    qs = urlencode(qsp)
    error = None; cols=[]; rows=[]; dash=None
    if tab.get("dash"):
        try:
            dash = compute_dash(request.args.get("date_from","2026-01-01"), request.args.get("date_to","2026-12-31"))
        except Exception as e:
            error = str(e)
    else:
        try:
            cols, rows = run_report(rpt, request.args)
        except Exception as e:
            error = str(e)
    return render_template("page.html", tabs=TABS, tab=tab, cur_tab=cur_tab, rpt=rpt,
                                  binds=display, cols=cols, rows=rows, error=error, qs=qs, dash=dash, hidden_tabs=hidden_tabs, hidden_reports=hidden_reports, hide_profit=load_hide_profit())

@app.route("/export")
def export():
    tab, rpt = find_report(request.args.get("tab", TABS[0]["id"]), request.args.get("report",""))
    try:
        cols, rows = run_report(rpt, request.args)
    except Exception as e:
        return "خطأ: " + str(e), 500
    try:
        hidden_cols = ["إشعار خصم مستقل (-)", "الخصم في الفاتورة", "إيداعات وتسويات (بدون عميل)"]
        valid_indices = [i for i, col in enumerate(cols) if col not in hidden_cols]
        filtered_cols = [cols[i] for i in valid_indices]

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws = wb.active; ws.title = "تقرير"
        ws.sheet_view.rightToLeft = True
        ws.append(filtered_cols)
        
        # تنسيق الرأس
        header_fill = PatternFill("solid", fgColor="4F46E5")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin', color='E2E8F0'), 
                        right=Side(style='thin', color='E2E8F0'), 
                        top=Side(style='thin', color='E2E8F0'), 
                        bottom=Side(style='thin', color='E2E8F0'))
                        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="right")
            cell.border = border
            
        # تنسيق صف الإجمالي (الصف الأول من البيانات وهو الصف الثاني في الإكسل)
        total_fill = PatternFill("solid", fgColor="EEF2FF")
        total_font = Font(bold=True, color="3730A3")
        
        for row_idx, r in enumerate(rows, start=2):
            filtered_r = [str(r[i]) if r[i] is not None else '' for i in valid_indices]
            ws.append(filtered_r)
            for cell_idx, cell in enumerate(ws[row_idx], start=1):
                cell.border = border
                cell.alignment = Alignment(horizontal="right")
                if row_idx == 2:  # صف الإجمالي
                    cell.fill = total_fill
                    cell.font = total_font

        for i in range(1, len(filtered_cols)+1):
            ws.column_dimensions[get_column_letter(i)].width = 22
            
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=%s.xlsx" % rpt["id"]})
    except ImportError:
        import csv
        buf = io.StringIO()
        buf.write('\ufeff')
        writer = csv.writer(buf)
        writer.writerow(filtered_cols)
        for r in rows:
            filtered_r = [str(r[i]) if r[i] is not None else '' for i in valid_indices]
            writer.writerow(filtered_r)
        return Response(buf.getvalue().encode('utf-8'),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=%s.csv" % rpt["id"]})

@app.route("/print")
def printview():
    tab, rpt = find_report(request.args.get("tab", TABS[0]["id"]), request.args.get("report",""))
    try:
        cols, rows = run_report(rpt, request.args)
    except Exception as e:
        return "خطأ: " + str(e), 500
    filt = []
    for p in rpt["params"]:
        v = request.args.get(p["name"], p.get("default",""))
        if v not in ("", None): filt.append((p["label"], v))
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    return render_template("print.html", title=rpt["title"], cols=cols, rows=rows, filt=filt, now=now)

@app.route("/settings/logout")
def settings_logout():
    session.pop("set_ok", None)
    return render_template("pin.html", error=False)

@app.route("/settings", methods=["GET", "POST"])
def settings():
    # بوابة رمز الدخول
    if not session.get("set_ok"):
        if request.method == "POST" and request.form.get("pin") is not None:
            if request.form.get("pin") == SETTINGS_PIN:
                session["set_ok"] = True
            else:
                return render_template("pin.html", error=True)
        else:
            return render_template("pin.html", error=False)
    saved = False
    if request.method == "POST" and request.form.get("action") == "save":
        htabs = [t["id"] for t in TABS if request.form.get("tab_"+t["id"])]
        hreps = []
        for t in TABS:
            for r in t["reports"]:
                key = t["id"]+"/"+r["id"]
                if request.form.get("rep_"+key):
                    hreps.append(key)
        save_hidden(htabs, hreps, bool(request.form.get("hide_profit")))
        saved = True
    ht, hr = load_hidden_raw()
    return render_template("settings.html", tabs=TABS, hidden_tabs=ht, hidden_reports=hr,
                                  saved=saved, hide_profit=load_hide_profit())

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
