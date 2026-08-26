import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
import re

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_pooled_conn

def normalize_name(name):
    if not name:
        return ""
    name = str(name).strip()
    name = re.sub(r'^منطقة\s+', '', name)
    name = re.sub(r'^حي\s+', '', name)
    name = name.replace(" ", "")
    return name

try:
    with get_pooled_conn() as conn:
        with conn.cursor() as cur:
            # Fetch all regions ordered by City then Region Name
            query = """
                SELECT 
                    c.CITY_NO, c.CITY_A_NAME,
                    r.R_CODE, r.R_A_NAME
                FROM REGIONS r
                JOIN CITIES c ON r.CITY_NO = c.CITY_NO
                ORDER BY c.CITY_NO, r.R_A_NAME
            """
            cur.execute(query)
            regions = cur.fetchall()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "دمج المناطق المتكررة"
            ws.sheet_view.rightToLeft = True
            
            headers = [
                'رقم المدينة', 'اسم المدينة',
                'رقم المنطقة (الحالي)', 'اسم المنطقة',
                'لدمج هذه المنطقة، ضع رقم المنطقة الصحيحة هنا'
            ]
            ws.append(headers)
            
            header_font = Font(bold=True, color="FFFFFF")
            fill_info = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            fill_action = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
            
            for i in range(1, 5):
                ws.cell(row=1, column=i).font = header_font
                ws.cell(row=1, column=i).fill = fill_info
            
            ws.cell(row=1, column=5).font = header_font
            ws.cell(row=1, column=5).fill = fill_action
            
            # Group by (City_No, Normalized_Region_Name)
            groups = defaultdict(list)
            for row_idx, r in enumerate(regions, start=2):
                city_no, city_name, r_code, r_name = r
                norm_name = normalize_name(r_name)
                groups[(city_no, norm_name)].append({
                    'row_idx': row_idx,
                    'r_code': r_code,
                    'r_name': r_name,
                    'full_row': [city_no, city_name, r_code, r_name, '']
                })
                
            merges_done = 0
            
            for (city_no, norm_name), regs in groups.items():
                if len(regs) > 1:
                    # Pick the primary region (smallest R_CODE)
                    primary = min(regs, key=lambda x: int(x['r_code']))
                    primary_id = primary['r_code']
                    
                    for reg in regs:
                        if reg['r_code'] != primary_id:
                            reg['full_row'][4] = primary_id
                            merges_done += 1
                            
            # Append rows to sheet
            for (city_no, norm_name), regs in groups.items():
                for reg in regs:
                    ws.append(reg['full_row'])
                
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 45

            os.makedirs('Results', exist_ok=True)
            output_path = 'Results/Region_Merge_Template_Prefilled.xlsx'
            wb.save(output_path)
            
            print(f"Generated Region Merge Template with {len(regions)} regions.")
            print(f"Proposed {merges_done} automatic merges.")
            print(f"File saved to {output_path}")

except Exception as e:
    print(f"Error: {e}")
