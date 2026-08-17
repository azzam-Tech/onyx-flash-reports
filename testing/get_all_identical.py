import oracledb
import os
import openpyxl

os.environ["PATH"] = r"C:\oracle\instantclient\instantclient_23_0;" + os.environ.get("PATH", "")
oracledb.init_oracle_client(lib_dir=r"C:\oracle\instantclient\instantclient_23_0")

def normalize(code):
    cleaned = code.replace('-', '').replace('.', '').replace('_', '').replace(' ', '').upper()
    return ''.join(sorted(cleaned))

def generate_identical_excel():
    try:
        # Manual overrides for items that can't be fuzzy-matched
        manual_overrides = {
            'TWS-8SRWM': 'SRWM-8TW.',
            'TWS-12SRWM': 'SRWM-12TW.',
            'HIKT-50S4KW3': 'HIKT-50S3'
        }
        
        # 1. Connect to DB and fetch all items for fuzzy matching
        conn = oracledb.connect(user='RPT_USER', password='ULT2016', dsn='100.100.1.100:1521/ORCL')
        cur = conn.cursor()
        
        cur.execute("SELECT I_CODE, I_NAME FROM IAS_ITM_MST")
        all_items = cur.fetchall()
        
        db_map = {}
        for row in all_items:
            i_code = row[0]
            if i_code:
                norm_code = normalize(i_code)
                if norm_code not in db_map:
                    db_map[norm_code] = []
                db_map[norm_code].append((i_code, row[1]))
                
        # Fetch current retail prices for all items to speed up
        cur.execute("SELECT I_CODE, I_PRICE FROM IAS_ITEM_PRICE WHERE LEV_NO = 2")
        price_rows = cur.fetchall()
        db_prices = {r[0]: r[1] for r in price_rows if r[0]}

        # 2. Read Prices from Excel
        file_path = r"C:\Users\amarn\OneDrive\Desktop\dbOnyxOnAntigravity\pricing\تحديث الاسعار.xlsx"
        wb_in = openpyxl.load_workbook(file_path, data_only=True)
        sheet_in = wb_in.active
        
        identical_items = []
        
        for row in sheet_in.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[2]:
                continue
            
            excel_code = str(row[0]).strip()
            excel_name = str(row[1]).strip() if row[1] else ""
            try:
                excel_price = float(row[2])
            except ValueError:
                continue
                
            # Determine correct code
            correct_code = None
            correct_name = ""
            
            # Direct mapping
            if excel_code in manual_overrides:
                correct_code = manual_overrides[excel_code]
            else:
                # First try exact match, handling dot issue
                temp_code = excel_code
                if temp_code.startswith('.'):
                    temp_code = temp_code[1:] + '.'
                    
                if temp_code in db_prices:
                    correct_code = temp_code
                else:
                    # Try fuzzy match
                    norm = normalize(excel_code)
                    if norm in db_map:
                        correct_code = db_map[norm][0][0]
                        correct_name = db_map[norm][0][1]
                        
            if not correct_code:
                continue
                
            if not correct_name:
                # Find name in db_map if possible, else just keep empty
                norm = normalize(correct_code)
                if norm in db_map:
                    correct_name = db_map[norm][0][1]

            # Check price
            if correct_code in db_prices:
                db_price = db_prices[correct_code]
                if db_price == excel_price:
                    identical_items.append((excel_code, correct_code, correct_name, db_price))
                    
        cur.close()
        conn.close()
        
        # 3. Write to Excel
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "الأصناف الغير محدثة (لتطابقها)"
        
        ws_out.append(["الكود في الإكسيل (الملخبط)", "الكود في قاعدة البيانات (الصحيح)", "اسم الصنف", "سعر التجزئة (المتطابق)"])
        
        for item in identical_items:
            ws_out.append([item[0], item[1], item[2], item[3]])
            
        out_path = r"C:\Users\amarn\OneDrive\Desktop\dbOnyxOnAntigravity\pricing\الأصناف_الغير_محدثة_لتطابق_السعر.xlsx"
        wb_out.save(out_path)
        print(f"Successfully generated Excel with {len(identical_items)} items at: {out_path}")
        
    except Exception as e:
        print(f"Failed: {e}")

if __name__ == '__main__':
    generate_identical_excel()
