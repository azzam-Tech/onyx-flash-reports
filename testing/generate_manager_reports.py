import sys
import os
import openpyxl

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_conn

def generate_manager_reports():
    file_path = r"C:\Users\amarn\Downloads\الثلاجات_تصنيف_اونكس_برو_مرتب.xlsx"
    missed_file = r"C:\Users\amarn\Downloads\الثلاجات_النشطة_المنسية.xlsx"
    stagnant_file = r"C:\Users\amarn\Downloads\الثلاجات_الراكدة_المدرجة_بالخطأ.xlsx"
    
    try:
        # 1. Read Manager's Excel
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        manager_icodes = set()
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0: continue
            icode = str(row[8]).strip() if row[8] else None
            if icode:
                manager_icodes.add(icode)
                
        with get_conn() as con:
            with con.cursor() as cur:
                # 2. Get Active Items
                cur.execute("""
                    SELECT DISTINCT I_CODE 
                    FROM ITEM_MOVEMENT 
                    WHERE I_CODE IN (SELECT I_CODE FROM IAS_ITM_MST WHERE G_CODE = '003')
                """)
                active_icodes = {r[0] for r in cur.fetchall()}
                
                # 3. Get Stagnant Items
                cur.execute("""
                    SELECT I_CODE 
                    FROM IAS_ITM_MST 
                    WHERE G_CODE = '003' 
                    AND I_CODE NOT IN (
                        SELECT DISTINCT I_CODE FROM ITEM_MOVEMENT WHERE I_CODE IS NOT NULL
                    )
                """)
                stagnant_icodes = {r[0] for r in cur.fetchall()}
                
                # 4. Calculate the sets
                missed_active = active_icodes - manager_icodes
                manager_stagnant = manager_icodes.intersection(stagnant_icodes)
                
                # Helper function to get item details and write to Excel
                def write_to_excel(icodes, out_path, title):
                    if not icodes:
                        return
                    format_strings = ','.join([f"'{c}'" for c in icodes])
                    cur.execute(f"SELECT I_CODE, I_NAME, I_E_NAME, AD_DATE FROM IAS_ITM_MST WHERE I_CODE IN ({format_strings}) ORDER BY I_CODE")
                    rows = cur.fetchall()
                    
                    out_wb = openpyxl.Workbook()
                    out_ws = out_wb.active
                    out_ws.title = title
                    out_ws.append(["رقم الصنف", "اسم الصنف عربي", "اسم الصنف انجليزي", "تاريخ الإضافة"])
                    for r in rows:
                        out_ws.append(r)
                    out_wb.save(out_path)
                    print(f"Generated: {out_path} ({len(rows)} items)")

                # 5. Generate files
                write_to_excel(missed_active, missed_file, "المنسية")
                write_to_excel(manager_stagnant, stagnant_file, "الراكدة")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    generate_manager_reports()
