import openpyxl
from collections import defaultdict

def safe_str(val):
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val)).strip()
    return str(val).strip()

try:
    old_file = r"C:\Users\amarn\Downloads\الثلاجات_تصنيف_اونكس_برو.xlsx"
    new_file = r"C:\Users\amarn\Downloads\الثلاجات_تصنيف_اونكس_برو_مرتب (1).xlsx"
    
    # 1. Read Old File
    wb_old = openpyxl.load_workbook(old_file, data_only=True)
    ws_old = wb_old.active
    
    old_data = defaultdict(list)
    for row_idx, row in enumerate(ws_old.iter_rows(min_row=2, values_only=True), start=2):
        i_code = safe_str(row[5])
        mng_code = safe_str(row[2])
        subg_code = safe_str(row[3])
        if i_code:
            old_data[i_code].append((row_idx, mng_code, subg_code))
            
    # Find some duplicates (items with >1 entries)
    duplicates = {k: v for k, v in old_data.items() if len(v) > 1}
    sample_dups = list(duplicates.keys())[:10]
    
    # 2. Read New File
    wb_new = openpyxl.load_workbook(new_file, data_only=True)
    ws_new = wb_new.active
    
    new_data = defaultdict(list)
    # The new file might have different columns depending on whether it has the "distracting" columns or not.
    # We will assume it has the original format:
    # 0: Main Group
    # 1: Main Group Name
    # 2: Sub Group Code (MNG_CODE)
    # 3: Sub Group Name
    # 4: Source
    # 5: Sub-Sub Group Code (SUBG_CODE)
    # 6: Sub-Sub Group Name
    # 7: Full Code
    # 8: Item Code
    
    # Let's first check headers of the new file to be sure
    headers = [cell.value for cell in ws_new[1]]
    
    # Let's find index for I_CODE, MNG_CODE, SUBG_CODE based on headers or just iterate all and guess based on positions
    # If the user says it's "before duplication", it probably has the full 11 columns like we saw first.
    # 0: G_CODE, 2: MNG_CODE, 5: SUBG_CODE, 8: I_CODE
    for row_idx, row in enumerate(ws_new.iter_rows(min_row=2, values_only=True), start=2):
        # We assume column 8 is I_CODE if len > 8
        if len(row) > 8:
            i_code = safe_str(row[8])
            mng_code = safe_str(row[2])
            subg_code = safe_str(row[5])
        else:
            # Maybe it's the 7 column format
            i_code = safe_str(row[5])
            mng_code = safe_str(row[2])
            subg_code = safe_str(row[3])
            
        if i_code:
            new_data[i_code].append((row_idx, mng_code, subg_code))
            
    # 3. Compare
    print(f"--- Comparison for 10 duplicated items ---")
    print(f"Columns found in NEW file: {headers}\n")
    
    for item in sample_dups:
        print(f"Item: {item}")
        print("  In OLD file (with duplicates):")
        for r in old_data[item]:
            print(f"    - Row {r[0]}: MNG_CODE={r[1]}, SUBG_CODE={r[2]}")
            
        print("  In NEW file (مرتب):")
        if item in new_data:
            for r in new_data[item]:
                print(f"    - Row {r[0]}: MNG_CODE={r[1]}, SUBG_CODE={r[2]}")
        else:
            print("    - NOT FOUND in NEW file!")
        print("-" * 50)
        
except Exception as e:
    print(f"Error: {e}")
