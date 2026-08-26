import openpyxl
from collections import defaultdict
import re

def normalize_name(name):
    if not name:
        return ""
    # Remove leading/trailing spaces
    name = str(name).strip()
    # Remove specific prefixes
    name = re.sub(r'^مدينة\s+', '', name)
    name = re.sub(r'^محافظة\s+', '', name)
    # Remove all spaces for exact matching
    name = name.replace(" ", "")
    return name

try:
    wb = openpyxl.load_workbook('Results/City_Merge_Template.xlsx')
    ws = wb['دمج المدن المتكررة']
    
    # Group rows by (Province_No, Normalized_City_Name)
    # Store tuples of (row_index, city_no)
    groups = defaultdict(list)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        prov_no, prov_name, city_no, city_name, merge_target = row
        norm_name = normalize_name(city_name)
        groups[(prov_no, norm_name)].append({'row_idx': row_idx, 'city_no': int(city_no)})
        
    merges_done = 0
    
    for (prov_no, norm_name), cities in groups.items():
        if len(cities) > 1:
            # We have duplicates! Find the primary city.
            # Rule: Prefer cities >= 101. If multiple, pick smallest.
            valid_cities = [c for c in cities if c['city_no'] >= 101]
            if valid_cities:
                primary_city = min(valid_cities, key=lambda x: x['city_no'])
            else:
                # If none are >= 101, pick the smallest overall
                primary_city = min(cities, key=lambda x: x['city_no'])
                
            primary_id = primary_city['city_no']
            
            # Update all other cities in this group to merge into primary_id
            for city in cities:
                if city['city_no'] != primary_id:
                    ws.cell(row=city['row_idx'], column=5).value = primary_id
                    merges_done += 1
                    
    wb.save('Results/City_Merge_Template_Prefilled.xlsx')
    print(f"Successfully pre-filled the template. Proposed {merges_done} automatic merges.")

except Exception as e:
    print(f"Error: {e}")
