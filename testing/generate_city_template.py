import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_pooled_conn

try:
    with get_pooled_conn() as conn:
        with conn.cursor() as cur:
            # Fetch all cities ordered by Province then City Name
            query = """
                SELECT 
                    p.PROV_NO, p.PROV_A_NAME,
                    c.CITY_NO, c.CITY_A_NAME
                FROM CITIES c
                JOIN IAS_PROVINCES p ON c.PROV_NO = p.PROV_NO
                WHERE p.PROV_NO BETWEEN 101 AND 114
                ORDER BY p.PROV_NO, c.CITY_A_NAME
            """
            cur.execute(query)
            cities = cur.fetchall()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "دمج المدن المتكررة"
            ws.sheet_view.rightToLeft = True
            
            headers = [
                'رقم المحافظة', 'اسم المحافظة',
                'رقم المدينة (الحالي)', 'اسم المدينة',
                'لدمج هذه المدينة مع أخرى، ضع رقم المدينة الصحيحة هنا'
            ]
            ws.append(headers)
            
            header_font = Font(bold=True, color="FFFFFF")
            fill_info = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            fill_action = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
            
            for i in range(1, 5):
                ws.cell(row=1, column=i).font = header_font
                ws.cell(row=1, column=i).fill = fill_info
            
            ws.cell(row=1, column=5).font = header_font
            ws.cell(row=1, column=5).fill = fill_action
            
            for row in cities:
                ws.append([row[0], row[1], row[2], row[3], ''])
                
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 45

            os.makedirs('Results', exist_ok=True)
            output_path = 'Results/City_Merge_Template.xlsx'
            wb.save(output_path)
            
            print(f"Generated City Merge Template with {len(cities)} cities.")
            print(f"File saved to {output_path}")

except Exception as e:
    print(f"Error: {e}")
