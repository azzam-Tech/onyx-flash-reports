# -*- coding: utf-8 -*-
"""لوحة تقارير SREEN — Flask + Oracle (RTL). تبويبات رئيسية وفرعية موصولة بالاستعلامات."""

import os
import json
import oracledb
import io
from urllib.parse import urlencode
from datetime import datetime
from flask import Flask, request, render_template_string, Response, session

app = Flask(__name__)
app.secret_key = os.environ.get("SREEN_SECRET", "sreen-reports-2026-secret-key")
SETTINGS_PIN = os.environ.get("SETTINGS_PIN", "00900")
APP_PIN = os.environ.get("APP_PIN", "12345")

_lib = os.environ.get("ORA_LIB_DIR")
try:
    oracledb.init_oracle_client(lib_dir=_lib) if _lib else oracledb.init_oracle_client()
    print("Thick mode ON")
except Exception as e:
    print("thick warn:", e)

DB_USER     = os.environ.get("ORA_USER",     "RPT_USER")
DB_PASSWORD = os.environ.get("ORA_PASSWORD", "ULT2016")
DB_DSN      = os.environ.get("ORA_DSN",      "192.168.1.10:1521/ORCL")


LOGIN_PAGE = """<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="UTF-8">
  <title>تسجيل الدخول - نظام التقارير</title>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;800&display=swap');
    body { margin: 0; padding: 0; background-color: #f4f5f8; font-family: 'Cairo', sans-serif; display: flex; align-items: center; justify-content: center; height: 100vh; }
    .card { background: #fff; padding: 40px; border-radius: 16px; box-shadow: 0 10px 30px rgba(0,0,0,0.05); text-align: center; max-width: 400px; width: 100%; border-top: 6px solid #4f46e5; }
    h2 { color: #1e293b; font-weight: 800; margin-bottom: 5px; }
    p { color: #64748b; margin-bottom: 25px; }
    input[type=password] { width: 100%; padding: 12px; margin-bottom: 20px; border: 2px solid #e2e8f0; border-radius: 8px; font-size: 16px; text-align: center; font-family: inherit; font-weight: 600; box-sizing: border-box; }
    input[type=password]:focus { outline: none; border-color: #4f46e5; }
    button { background: #4f46e5; color: #fff; border: none; border-radius: 8px; padding: 12px 20px; font-size: 16px; cursor: pointer; width: 100%; font-weight: 600; transition: background 0.2s; }
    button:hover { background: #4338ca; }
    .err { color: #ef4444; background: #fee2e2; padding: 10px; border-radius: 8px; margin-bottom: 15px; font-size: 14px; }
  </style>
</head>
<body>
  <div class="card">
    <svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="#4f46e5" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="margin-bottom:10px;"><rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect><path d="M7 11V7a5 5 0 0 1 10 0v4"></path></svg>
    <h2>نظام التقارير</h2>
    <p>يرجى إدخال رمز المرور للمتابعة</p>
    {% if error %}<div class="err">{{ error }}</div>{% endif %}
    <form method="POST">
      <input type="password" name="pin" placeholder="الرمز السري (PIN)" autofocus required>
      <button type="submit">دخول آمن</button>
    </form>
  </div>
</body>
</html>"""

from flask import redirect

@app.before_request
def require_login():
    if request.endpoint not in ('login', 'static') and not session.get('logged_in'):
        return redirect('/login')

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if request.form.get("pin") == APP_PIN:
            session['logged_in'] = True
            return redirect('/')
        else:
            error = "الرمز غير صحيح، حاول مرة أخرى."
    return render_template_string(LOGIN_PAGE, error=error)

@app.route("/logout")
def logout():
    session.pop('logged_in', None)
    return redirect('/login')



def get_conn():
    return oracledb.connect(user=DB_USER, password=DB_PASSWORD, dsn=DB_DSN)

SETTINGS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")
# كل ما يخص الربح: يُخفى عند تفعيل "إخفاء الربح"
PROFIT_TABS = {"prof"}
PROFIT_REPORTS = {"fin/income_statement", "fin/cost_centers"}
def _load_raw():
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}
def load_hidden_raw():
    d = _load_raw()
    return set(d.get("tabs", [])), set(d.get("reports", []))
def load_hide_profit():
    return bool(_load_raw().get("hide_profit"))
def load_hidden():
    """المجموعات الفعّالة المطبَّقة على الواجهة (تشمل إخفاء الربح إن كان مفعّلاً)."""
    tabs, reps = load_hidden_raw()
    if load_hide_profit():
        tabs = tabs | PROFIT_TABS
        reps = reps | PROFIT_REPORTS
    return tabs, reps
def save_hidden(tabs, reports, hide_profit=False):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump({"tabs": list(tabs), "reports": list(reports),
                       "hide_profit": bool(hide_profit)}, f, ensure_ascii=False)
    except Exception as e:
        print("settings save error:", e)

DFROM = {"name":"date_from","label":"من تاريخ","type":"date","default":"2026-01-01"}
DTO   = {"name":"date_to","label":"إلى تاريخ","type":"date","default":"2026-07-10"}
REP   = {"name":"rep_code","label":"المندوب (اختياري)","type":"text","default":""}
INCR  = {"name":"inc_rcpt","label":"سندات القبض","type":"select","default":"1","options":[["1","تضمين"],["0","استبعاد"]]}
INCN  = {"name":"inc_net","label":"قيود الشبكة المنفصلة","type":"select","default":"1","options":[["1","تضمين"],["0","استبعاد"]]}
INCC  = {"name":"inc_cash","label":"المبيعات النقدية","type":"select","default":"1","options":[["1","تضمين"],["0","استبعاد"]]}
GRP   = {"name":"grp_by","label":"تجميع حسب","type":"select","default":"rep","options":[["rep","المندوب"],["cc","مركز التكلفة"],["cst","العميل"]]}
CST   = {"name":"c_code","label":"العميل (اختياري)","type":"text","default":""}
BTYPE = {"name":"bill_type","label":"نوع المستند","type":"select","default":"",
         "options":[["","الكل"],["1","مبيعات نقدية"],["4","مبيعات آجلة"],["2","مرتجع نقدي"],["5","مرتجع آجل"]]}

TABS = [
 {"id":"dash","title":"لوحة القيادة","icon":"M3 13h8V3H3zM13 21h8V3h-8zM3 21h8v-6H3z","dash":True,"reports":[{"id":"overview","title":"نظرة عامة","params":[{"name":"date_from","label":"من تاريخ","type":"date","default":"2026-01-01"},{"name":"date_to","label":"إلى تاريخ","type":"date","default":"2026-12-31"}]}]},
 {"id":"sales","title":"المبيعات","icon":"M4 20V10M10 20V4M16 20v-7M22 20H2","reports":[
   {"id":"bills","title":"فواتير المبيعات","params":[DFROM,DTO,BTYPE,REP,CST],"sql":"""
     SELECT CASE b.BILL_DOC_TYPE WHEN 1 THEN 'مبيعات نقدية' WHEN 4 THEN 'مبيعات آجلة' 
                 WHEN 2 THEN 'مرتجع نقدي' WHEN 5 THEN 'مرتجع آجل' ELSE 'أخرى' END AS "نوع المستند",
            b.BILL_NO AS "رقم الفاتورة", TO_CHAR(b.BILL_DATE,'YYYY-MM-DD') AS "التاريخ",
            b.C_CODE AS "كود العميل", MAX(c.C_A_NAME) AS "اسم العميل", b.REP_CODE AS "المندوب",
            TO_CHAR(NVL(b.BILL_AMT,0),'FM999,999,990.00') AS "المبلغ",
            TO_CHAR(NVL(d.itm_disc,0),'FM999,999,990.00') AS "خصم الأصناف",
            TO_CHAR(NVL(b.DISC_AMT,0),'FM999,999,990.00') AS "الخصم الإجمالي",
            TO_CHAR(NVL(b.VAT_AMT,0),'FM999,999,990.00') AS "الضريبة",
            TO_CHAR((NVL(b.BILL_AMT,0)-NVL(b.DISC_AMT,0)-NVL(d.itm_disc,0)+NVL(b.VAT_AMT,0)+NVL(b.OTHR_AMT,0)) * CASE WHEN b.BILL_DOC_TYPE IN (2,5) THEN -1 ELSE 1 END,'FM999,999,990.00') AS "الصافي",
            CASE NVL(b.BILL_POST,0) WHEN 1 THEN 'مرحّلة' ELSE 'غير مرحّلة' END AS "الحالة"
     FROM IAS20261.IAS_BILL_MST b
     LEFT JOIN IAS20261.CUSTOMER c ON c.C_CODE = b.C_CODE
     LEFT JOIN (
         SELECT BILL_DOC_TYPE, BILL_NO, BILL_SER, SUM(NVL(DIS_AMT,0)) as itm_disc
         FROM IAS20261.IAS_BILL_DTL
         GROUP BY BILL_DOC_TYPE, BILL_NO, BILL_SER
     ) d ON b.BILL_DOC_TYPE = d.BILL_DOC_TYPE AND b.BILL_NO = d.BILL_NO AND b.BILL_SER = d.BILL_SER
     WHERE b.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND b.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       AND b.BILL_DOC_TYPE IN (1,4,2,5)
       AND (:bill_type IS NULL OR b.BILL_DOC_TYPE = :bill_type)
       AND (:rep_code IS NULL OR b.REP_CODE = :rep_code)
       AND (:c_code IS NULL OR b.C_CODE = :c_code)
     GROUP BY b.BILL_DOC_TYPE, b.BILL_NO, b.BILL_DATE, b.C_CODE, b.REP_CODE, b.BILL_AMT, b.DISC_AMT, b.VAT_AMT, b.OTHR_AMT, b.BILL_POST, d.itm_disc
     ORDER BY b.BILL_DATE DESC, b.BILL_NO DESC FETCH FIRST 300 ROWS ONLY"""},
   {"id":"sales_vs_collection","title":"المبيعات مقابل التحصيل","params":[DFROM,DTO],"sql":"""
     WITH sales_data AS (
         SELECT CC_CODE,
                SUM(NVL(BILL_AMT,0)) - SUM(NVL(DISC_AMT_MST,0)) as sales
         FROM IAS20261.IAS_BILL_MST
         WHERE BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') 
           AND BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
           AND BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8)
         GROUP BY CC_CODE
     ),
     returns_data AS (
         SELECT CC_CODE,
                SUM(NVL(BILL_AMT,0)) - SUM(NVL(DISC_AMT_MST,0)) as returns
         FROM IAS20261.IAS_RT_BILL_MST
         WHERE RT_BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') 
           AND RT_BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
           AND RT_BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8)
         GROUP BY CC_CODE
     ),
     discount_notice AS (
         SELECT CC_CODE, ROUND(SUM(NVL(CR_AMT,0)) / 1.15, 2) as ext_disc
         FROM IAS20261.IAS_POST_DTL
         WHERE DOC_TYPE = 15 AND NVL(CR_AMT,0) > 0 AND NVL(DOC_POST,0) = 1
           AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') 
           AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         GROUP BY CC_CODE
     ),
     net_sales AS (
         SELECT NVL(NVL(s.CC_CODE, r.CC_CODE), d.CC_CODE) AS CC_CODE,
                SUM(NVL(s.sales, 0)) - SUM(NVL(r.returns, 0)) - SUM(NVL(d.ext_disc, 0)) AS net_sales_amt
         FROM sales_data s
         FULL OUTER JOIN returns_data r ON s.CC_CODE = r.CC_CODE
         FULL OUTER JOIN discount_notice d ON NVL(s.CC_CODE, r.CC_CODE) = d.CC_CODE
         GROUP BY NVL(NVL(s.CC_CODE, r.CC_CODE), d.CC_CODE)
     ),
     all_trans AS (
         SELECT TO_CHAR(CC_CODE) as grp_code,
                CR_AMT as rcpt, 0 as net_jrn, 0 as cash_sales, 0 as inv_disc, 0 as cash_ret, 0 as ext_notice, 0 as rcpt_unknown, 0 as unposted_rcpt, 0 as unposted_unknown
         FROM IAS20261.IAS_POST_DTL
         WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=2 AND NVL(CR_AMT,0)>0 AND C_CODE IS NOT NULL
           AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         UNION ALL
         SELECT TO_CHAR(CC_CODE), 0, 0, 0, 0, 0, 0, 0, CR_AMT, 0
         FROM IAS20261.IAS_POST_DTL
         WHERE NVL(DOC_POST,0)=0 AND DOC_TYPE=2 AND NVL(CR_AMT,0)>0 AND C_CODE IS NOT NULL
           AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         UNION ALL
         SELECT TO_CHAR(CC_CODE), 0, 0, 0, 0, 0, 0, 0, 0, CR_AMT
         FROM IAS20261.IAS_POST_DTL
         WHERE NVL(DOC_POST,0)=0 AND DOC_TYPE=2 AND NVL(CR_AMT,0)>0 AND C_CODE IS NULL
           AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         UNION ALL
         SELECT TO_CHAR(CC_CODE), 0, CR_AMT, 0, 0, 0, 0, 0, 0, 0
         FROM IAS20261.IAS_POST_DTL
         WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=1 AND JV_TYPE=2 AND NVL(CR_AMT,0)>0 AND C_CODE IS NOT NULL
           AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         UNION ALL
         SELECT TO_CHAR(b.CC_CODE), 0, 0, NVL(p.DR_AMT,0), NVL(b.DISC_AMT,0), 0, 0, 0, 0, 0
         FROM IAS20261.IAS_BILL_MST b
         JOIN IAS20261.IAS_POST_DTL p ON p.DOC_NO = b.BILL_NO AND p.DOC_SER = b.BILL_SER AND p.DOC_TYPE = 4 AND TO_CHAR(p.A_CODE) LIKE '111%'
         WHERE b.BILL_DOC_TYPE=1 AND NVL(p.DOC_POST,0)=1 AND p.DR_AMT > 0
           AND b.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND b.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         UNION ALL
         SELECT TO_CHAR(CC_CODE), 0, 0, 0, 0, CR_AMT, 0, 0, 0, 0
         FROM IAS20261.IAS_POST_DTL
         WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=5 AND A_CODE LIKE '111%' AND NVL(CR_AMT,0)>0
           AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         UNION ALL
         SELECT TO_CHAR(CC_CODE), 0, 0, 0, 0, 0, CR_AMT, 0, 0, 0
         FROM IAS20261.IAS_POST_DTL
         WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=15 AND NVL(CR_AMT,0)>0
           AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         UNION ALL
         SELECT TO_CHAR(CC_CODE), 0, 0, 0, 0, 0, 0, CR_AMT, 0, 0
         FROM IAS20261.IAS_POST_DTL
         WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=2 AND NVL(CR_AMT,0)>0 AND C_CODE IS NULL
           AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
     ),
     base_collection AS (
         SELECT grp_code,
                (SUM(rcpt) + SUM(unposted_rcpt) + SUM(unposted_unknown) + SUM(net_jrn) + SUM(cash_sales) - SUM(cash_ret)) as total_inc
         FROM all_trans
         WHERE grp_code IS NOT NULL
         GROUP BY grp_code
     )
     SELECT NVL(ns.CC_CODE, bc.grp_code) AS "المركز",
            MAX(cc.CC_A_NAME) AS "اسم المركز",
            TO_CHAR(SUM(NVL(ns.net_sales_amt, 0)), 'FM999,999,999,990.00') AS "صافي المبيعات",
            TO_CHAR(SUM(NVL(ns.net_sales_amt, 0)) * 1.15, 'FM999,999,999,990.00') AS "صافي المبيعات بالضريبة",
            TO_CHAR(SUM(NVL(bc.total_inc, 0)), 'FM999,999,999,990.00') AS "إجمالي التحصيل"
     FROM net_sales ns
     FULL OUTER JOIN base_collection bc ON ns.CC_CODE = bc.grp_code
     LEFT JOIN IAS20261.COST_CENTERS cc ON cc.CC_CODE = NVL(ns.CC_CODE, bc.grp_code)
     GROUP BY NVL(ns.CC_CODE, bc.grp_code)
     HAVING (SUM(NVL(ns.net_sales_amt, 0)) <> 0 OR SUM(NVL(bc.total_inc, 0)) <> 0)
     ORDER BY SUM(NVL(ns.net_sales_amt, 0)) DESC
     FETCH FIRST 300 ROWS ONLY"""},

   {"id":"by_item","title":"حسب الصنف","params":[DFROM,DTO],"sql":"""
     WITH dt AS (
       SELECT dt.I_CODE,
              CASE WHEN b.BILL_DOC_TYPE IN (1,4) THEN 1 WHEN b.BILL_DOC_TYPE IN (2,5) THEN -1 ELSE 0 END as sign,
              NVL(dt.I_QTY,0) as qty,
              (NVL(dt.I_QTY,0) * NVL(dt.I_PRICE,0) - NVL(dt.DIS_AMT,0)) as item_net,
              CASE WHEN NVL(b.BILL_AMT,0) = 0 THEN 0 ELSE 
                ((NVL(dt.I_QTY,0) * NVL(dt.I_PRICE,0) - NVL(dt.DIS_AMT,0)) / b.BILL_AMT) * NVL(b.DISC_AMT,0) 
              END as prorated_disc
       FROM IAS20261.IAS_BILL_DTL dt
       JOIN IAS20261.IAS_BILL_MST b ON b.BILL_DOC_TYPE=dt.BILL_DOC_TYPE AND b.BILL_NO=dt.BILL_NO AND b.BILL_SER=dt.BILL_SER
       WHERE b.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND b.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND b.BILL_DOC_TYPE IN (1,4,2,5)
     )
     SELECT dt.I_CODE AS "كود الصنف", MAX(m.I_NAME) AS "اسم الصنف",
            ROUND(SUM(CASE WHEN dt.sign=1 THEN dt.qty ELSE 0 END),2) AS "كمية المبيعات",
            ROUND(SUM(CASE WHEN dt.sign=-1 THEN dt.qty ELSE 0 END),2) AS "كمية المردودات (-)",
            TO_CHAR(SUM(dt.item_net * dt.sign),'FM999,999,999,990.00') AS "قيمة المبيعات",
            TO_CHAR(SUM(dt.prorated_disc * dt.sign),'FM999,999,999,990.00') AS "نصيب الصنف من الخصم (-)",
            TO_CHAR(SUM((dt.item_net - dt.prorated_disc) * dt.sign),'FM999,999,999,990.00') AS "الصافي"
     FROM dt
     LEFT JOIN IAS20261.IAS_ITM_MST m ON m.I_CODE=dt.I_CODE
     GROUP BY dt.I_CODE
     ORDER BY SUM((dt.item_net - dt.prorated_disc) * dt.sign) DESC 
     FETCH FIRST 300 ROWS ONLY"""},
   {"id":"by_customer","title":"حسب العميل","params":[DFROM,DTO],"sql":"""
     WITH s AS (
       SELECT b.C_CODE, b.REP_CODE,
              CASE WHEN b.BILL_DOC_TYPE IN (1,4) THEN 1 ELSE 0 END as is_sale,
              CASE WHEN b.BILL_DOC_TYPE IN (2,5) THEN 1 ELSE 0 END as is_ret,
              CASE WHEN b.BILL_DOC_TYPE IN (1,4) THEN 1 WHEN b.BILL_DOC_TYPE IN (2,5) THEN -1 ELSE 0 END as sign,
              NVL(b.BILL_AMT,0) amt, NVL(b.DISC_AMT,0) + NVL(d.itm_disc,0) disc, 0 as ext_disc, NVL(b.VAT_AMT,0) vat, NVL(b.OTHR_AMT,0) othr
       FROM IAS20261.IAS_BILL_MST b
       LEFT JOIN (
           SELECT BILL_DOC_TYPE, BILL_NO, BILL_SER, SUM(NVL(DIS_AMT,0)) as itm_disc
           FROM IAS20261.IAS_BILL_DTL
           GROUP BY BILL_DOC_TYPE, BILL_NO, BILL_SER
       ) d ON b.BILL_DOC_TYPE = d.BILL_DOC_TYPE AND b.BILL_NO = d.BILL_NO AND b.BILL_SER = d.BILL_SER
       WHERE b.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND b.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND b.BILL_DOC_TYPE IN (1,4,2,5)
       UNION ALL
       SELECT C_CODE, REP_CODE,
              0 as is_sale, 0 as is_ret, 0 as sign,
              0 as amt, 0 as disc, NVL(CR_AMT,0) as ext_disc, 0 as vat, 0 as othr
       FROM IAS20261.IAS_POST_DTL
       WHERE DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND DOC_TYPE = 15 AND NVL(CR_AMT,0) > 0 AND NVL(DOC_POST,0) = 1
     )
     SELECT s.C_CODE AS "كود العميل", MAX(c.C_A_NAME) AS "اسم العميل", MAX(s.REP_CODE) AS "المندوب",
            SUM(s.is_sale) AS "فواتير مبيعات",
            SUM(s.is_ret) AS "فواتير مرتجعات",
            TO_CHAR(SUM(s.amt * s.is_sale),'FM999,999,999,990.00') AS "المبيعات",
            TO_CHAR(SUM(s.amt * s.is_ret),'FM999,999,999,990.00') AS "المردودات (-)",
            TO_CHAR(SUM(s.disc * s.sign),'FM999,999,999,990.00') AS "خصم الفواتير والأصناف (-)",
            TO_CHAR(SUM(s.ext_disc),'FM999,999,999,990.00') AS "إشعار خصم مستقل (-)",
            TO_CHAR(SUM((s.amt - s.disc) * s.sign) - SUM(s.ext_disc),'FM999,999,999,990.00') AS "الصافي قبل الضريبة",
            TO_CHAR(SUM((s.amt - s.disc + s.vat + s.othr) * s.sign) - SUM(s.ext_disc),'FM999,999,999,990.00') AS "الإجمالي بالضريبة"
     FROM s LEFT JOIN IAS20261.CUSTOMER c ON c.C_CODE=s.C_CODE
     WHERE s.C_CODE IS NOT NULL
     GROUP BY s.C_CODE 
     ORDER BY SUM((s.amt - s.disc) * s.sign) - SUM(s.ext_disc) DESC 
     FETCH FIRST 300 ROWS ONLY"""},
   {"id":"by_salesman","title":"حسب المندوب","params":[DFROM,DTO],"sql":"""
     WITH s AS (
       SELECT b.REP_CODE, b.C_CODE,
              CASE WHEN b.BILL_DOC_TYPE IN (1,4) THEN 1 ELSE 0 END as is_sale,
              CASE WHEN b.BILL_DOC_TYPE IN (2,5) THEN 1 ELSE 0 END as is_ret,
              CASE WHEN b.BILL_DOC_TYPE IN (1,4) THEN 1 WHEN b.BILL_DOC_TYPE IN (2,5) THEN -1 ELSE 0 END as sign,
              NVL(b.BILL_AMT,0) amt, NVL(b.DISC_AMT,0) + NVL(d.itm_disc,0) disc, 0 as ext_disc, NVL(b.VAT_AMT,0) vat, NVL(b.OTHR_AMT,0) othr
       FROM IAS20261.IAS_BILL_MST b
       LEFT JOIN (
           SELECT BILL_DOC_TYPE, BILL_NO, BILL_SER, SUM(NVL(DIS_AMT,0)) as itm_disc
           FROM IAS20261.IAS_BILL_DTL
           GROUP BY BILL_DOC_TYPE, BILL_NO, BILL_SER
       ) d ON b.BILL_DOC_TYPE = d.BILL_DOC_TYPE AND b.BILL_NO = d.BILL_NO AND b.BILL_SER = d.BILL_SER
       WHERE b.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND b.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND b.BILL_DOC_TYPE IN (1,4,2,5)
       UNION ALL
       SELECT REP_CODE, C_CODE,
              0 as is_sale, 0 as is_ret, 0 as sign,
              0 as amt, 0 as disc, NVL(CR_AMT,0) as ext_disc, 0 as vat, 0 as othr
       FROM IAS20261.IAS_POST_DTL
       WHERE DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND DOC_TYPE = 15 AND NVL(CR_AMT,0) > 0 AND NVL(DOC_POST,0) = 1
     )
     SELECT s.REP_CODE AS "كود المندوب", MAX(sm.REPRS_A_NAME) AS "اسم المندوب",
            COUNT(DISTINCT s.C_CODE) AS "عدد العملاء",
            SUM(s.is_sale) AS "فواتير مبيعات",
            SUM(s.is_ret) AS "فواتير مرتجعات",
            TO_CHAR(SUM(s.amt * s.is_sale),'FM999,999,999,990.00') AS "المبيعات",
            TO_CHAR(SUM(s.amt * s.is_ret),'FM999,999,999,990.00') AS "المردودات (-)",
            TO_CHAR(SUM(s.disc * s.sign),'FM999,999,999,990.00') AS "خصم الفواتير والأصناف (-)",
            TO_CHAR(SUM(s.ext_disc),'FM999,999,999,990.00') AS "إشعار خصم مستقل (-)",
            TO_CHAR(SUM((s.amt - s.disc) * s.sign) - SUM(s.ext_disc),'FM999,999,999,990.00') AS "الصافي قبل الضريبة",
            TO_CHAR(SUM((s.amt - s.disc + s.vat + s.othr) * s.sign) - SUM(s.ext_disc),'FM999,999,999,990.00') AS "الإجمالي بالضريبة"
     FROM s LEFT JOIN IAS20261.SALES_MAN sm ON sm.REPRS_CODE=s.REP_CODE
     WHERE s.REP_CODE IS NOT NULL
     GROUP BY s.REP_CODE 
     ORDER BY SUM((s.amt - s.disc) * s.sign) - SUM(s.ext_disc) DESC 
     FETCH FIRST 300 ROWS ONLY"""},
 ]},
 {"id":"ar","title":"العملاء والمدينون","icon":"M9 8a3 3 0 100-6 3 3 0 000 6zM3 20c0-3 3-5 6-5s6 2 6 5","reports":[
   {"id":"balances","title":"أرصدة العملاء","params":[],"sql":"""
     SELECT p.C_CODE AS "كود العميل", MAX(c.C_A_NAME) AS "اسم العميل", MAX(c.REP_CODE) AS "المندوب",
            TO_CHAR(SUM(NVL(p.DR_AMT,0)-NVL(p.CR_AMT,0)),'FM999,999,999,990.00') AS "الرصيد (مدين)"
     FROM IAS20261.IAS_POST_DTL p LEFT JOIN IAS20261.CUSTOMER c ON c.C_CODE=p.C_CODE
     WHERE p.C_CODE IS NOT NULL AND NVL(p.DOC_POST,0)=1
     GROUP BY p.C_CODE HAVING SUM(NVL(p.DR_AMT,0)-NVL(p.CR_AMT,0))<>0
     ORDER BY SUM(NVL(p.DR_AMT,0)-NVL(p.CR_AMT,0)) DESC FETCH FIRST 300 ROWS ONLY"""},
   {"id":"statement","title":"كشف حساب عميل","params":[{"name":"c_code","label":"كود العميل","type":"text","default":"1381"},DFROM,DTO],"sql":"""
     SELECT "التاريخ","نوع المستند","رقم المستند","البيان","مدين","دائن","الرصيد" FROM (
       SELECT TO_CHAR(p.DOC_DATE,'YYYY-MM-DD') AS "التاريخ", d.JV_NAME AS "نوع المستند",
              p.DOC_NO AS "رقم المستند", p.DOC_DESC AS "البيان",
              TO_CHAR(NVL(p.DR_AMT,0),'FM999,999,990.00') AS "مدين",
              TO_CHAR(NVL(p.CR_AMT,0),'FM999,999,990.00') AS "دائن",
              TO_CHAR(SUM(NVL(p.DR_AMT,0)-NVL(p.CR_AMT,0)) OVER (ORDER BY p.DOC_DATE,p.DOC_NO,p.DOC_SER),'FM999,999,990.00') AS "الرصيد",
              p.DOC_DATE s1, p.DOC_NO s2, p.DOC_SER s3
       FROM IAS20261.IAS_POST_DTL p
       LEFT JOIN IAS_SYS.IAS_DOCJV_TYPE_SYSTEMS d ON d.DOC_TYPE=p.DOC_TYPE AND d.JV_TYPE=p.JV_TYPE AND d.LANG_NO=1
       WHERE p.C_CODE = :c_code AND NVL(p.DOC_POST,0)=1
         AND p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       ORDER BY p.DOC_DATE, p.DOC_NO, p.DOC_SER
     ) FETCH FIRST 1000 ROWS ONLY"""},
   {"id":"aging","title":"أعمار الديون","params":[
     {"name":"date_from","label":"من تاريخ (اختياري)","type":"date","default":""},
     {"name":"date_to","label":"إلى تاريخ","type":"date","default":"2026-07-10"},
     {"name":"rep_code","label":"المندوب (اختياري)","type":"text","default":""},
     {"name":"jv_type","label":"نوع القيد","type":"select","default":"","dynamic":"jv","options":[["","الكل"]]}
   ],"sql":"""
     WITH pay AS (SELECT C_CODE, SUM(NVL(CR_AMT,0)) paid FROM IAS20261.IAS_POST_DTL
                  WHERE NVL(DOC_POST,0)=1 AND C_CODE IS NOT NULL
                    AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
                    AND (:date_from IS NULL OR DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD'))
                    AND (:jv_type IS NULL OR JV_TYPE = :jv_type)
                  GROUP BY C_CODE),
     charges AS (SELECT p.C_CODE, p.DOC_DATE, NVL(p.DR_AMT,0) amt,
                   SUM(NVL(p.DR_AMT,0)) OVER (PARTITION BY p.C_CODE ORDER BY p.DOC_DATE,p.DOC_NO,p.DOC_SER
                        ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) cum
                 FROM IAS20261.IAS_POST_DTL p
                 WHERE NVL(p.DOC_POST,0)=1 AND p.C_CODE IS NOT NULL AND NVL(p.DR_AMT,0)>0
                   AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
                   AND (:date_from IS NULL OR p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD'))
                   AND (:jv_type IS NULL OR p.JV_TYPE = :jv_type)),
     openit AS (SELECT ch.C_CODE, GREATEST(0,LEAST(ch.amt,ch.cum-NVL(pay.paid,0))) unpaid,
                   TRUNC(TO_DATE(:date_to,'YYYY-MM-DD'))-TRUNC(ch.DOC_DATE) age
                FROM charges ch JOIN pay ON pay.C_CODE=ch.C_CODE)
     SELECT o.C_CODE AS "كود العميل", MAX(c.C_A_NAME) AS "اسم العميل", MAX(c.REP_CODE) AS "المندوب",
            TO_CHAR(SUM(CASE WHEN o.age<=30 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "0-30",
            TO_CHAR(SUM(CASE WHEN o.age BETWEEN 31 AND 60 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "31-60",
            TO_CHAR(SUM(CASE WHEN o.age BETWEEN 61 AND 90 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "61-90",
            TO_CHAR(SUM(CASE WHEN o.age BETWEEN 91 AND 120 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "91-120",
            TO_CHAR(SUM(CASE WHEN o.age>120 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "أكثر من 120",
            TO_CHAR(SUM(o.unpaid),'FM999,999,990.00') AS "الإجمالي"
     FROM openit o LEFT JOIN IAS20261.CUSTOMER c ON c.C_CODE=o.C_CODE
     WHERE o.unpaid>0 AND (:rep_code IS NULL OR c.REP_CODE = :rep_code)
     GROUP BY o.C_CODE ORDER BY SUM(o.unpaid) DESC FETCH FIRST 300 ROWS ONLY"""},
   {"id":"dormant","title":"العملاء الخاملون","params":[{"name":"as_of","label":"حتى تاريخ","type":"date","default":"2026-07-10"},{"name":"days","label":"أيام الخمول","type":"number","default":"90"}],"sql":"""
     SELECT * FROM (
       SELECT c.C_CODE AS "كود العميل", c.C_A_NAME AS "اسم العميل", c.REP_CODE AS "المندوب",
              TO_CHAR(lb.last_bill,'YYYY-MM-DD') AS "آخر فاتورة",
              (TRUNC(TO_DATE(:as_of,'YYYY-MM-DD'))-TRUNC(lb.last_bill)) AS "أيام منذ آخر تعامل"
       FROM IAS20261.CUSTOMER c
       LEFT JOIN (SELECT C_CODE, MAX(BILL_DATE) last_bill FROM IAS20261.IAS_BILL_MST WHERE BILL_DOC_TYPE IN (1,4) GROUP BY C_CODE) lb ON lb.C_CODE=c.C_CODE
       WHERE NVL(c.INACTIVE,0)=0 AND (lb.last_bill IS NULL OR lb.last_bill < TO_DATE(:as_of,'YYYY-MM-DD') - :days)
       ORDER BY lb.last_bill NULLS FIRST
     ) WHERE ROWNUM <= 300"""},

    {"id":"true_income_statement","title":"قائمة الدخل (الحقيقية)","params":[DFROM,DTO,REP],"sql":"""
        WITH gl_base AS (
          SELECT 
              A_CODE as acc_code,
              SUM(CASE WHEN DOC_DATE < TO_DATE(:date_from, 'YYYY-MM-DD') THEN NVL(DR_AMT,0) ELSE 0 END) as op_dr,
              SUM(CASE WHEN DOC_DATE < TO_DATE(:date_from, 'YYYY-MM-DD') THEN NVL(CR_AMT,0) ELSE 0 END) as op_cr,
              SUM(CASE WHEN DOC_DATE >= TO_DATE(:date_from, 'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to, 'YYYY-MM-DD')+1 THEN NVL(DR_AMT,0) ELSE 0 END) as mv_dr,
              SUM(CASE WHEN DOC_DATE >= TO_DATE(:date_from, 'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to, 'YYYY-MM-DD')+1 THEN NVL(CR_AMT,0) ELSE 0 END) as mv_cr
          FROM IAS20261.IAS_POST_DTL
          WHERE (:rep_code IS NULL OR REP_CODE = :rep_code OR CC_CODE = :rep_code)
            AND NVL(DOC_POST,0)=1
            AND (
                A_CODE LIKE '31102%' OR A_CODE LIKE '31104%' OR A_CODE LIKE '31105%' OR A_CODE LIKE '31109%' OR A_CODE LIKE '31110%' OR
                A_CODE LIKE '32101%' OR A_CODE LIKE '32201%' OR A_CODE LIKE '32401%' OR A_CODE LIKE '32801%' OR
                A_CODE LIKE '41101%' OR A_CODE LIKE '41202%'
            )
          GROUP BY A_CODE
        ),
        inv_cogs AS (
          SELECT 
              '311010001' as acc_code,
              SUM(CASE WHEN m.BILL_DATE < TO_DATE(:date_from, 'YYYY-MM-DD') THEN NVL(im.I_QTY,0) * NVL(it.PRIMARY_COST,0) ELSE 0 END) as op_dr,
              0 as op_cr,
              SUM(CASE WHEN m.BILL_DATE >= TO_DATE(:date_from, 'YYYY-MM-DD') AND m.BILL_DATE < TO_DATE(:date_to, 'YYYY-MM-DD')+1 THEN NVL(im.I_QTY,0) * NVL(it.PRIMARY_COST,0) ELSE 0 END) as mv_dr,
              0 as mv_cr
          FROM IAS20261.ITEM_MOVEMENT im
          JOIN IAS20261.IAS_ITM_MST it ON it.I_CODE = im.I_CODE
          JOIN IAS20261.IAS_BILL_MST m 
            ON m.BILL_DOC_TYPE = im.BILL_DOC_TYPE 
           AND m.BILL_NO = im.DOC_NO 
           AND m.BILL_SER = im.DOC_SER
          WHERE (:rep_code IS NULL OR m.REP_CODE = :rep_code)
            AND im.DOC_TYPE = 1 
            AND NVL(im.I_QTY,0) > 0
        ),
        inv_cogs_ret AS (
          SELECT 
              '311030001' as acc_code,
              0 as op_dr,
              SUM(CASE WHEN r.RT_BILL_DATE < TO_DATE(:date_from, 'YYYY-MM-DD') THEN NVL(im.I_QTY,0) * NVL(it.PRIMARY_COST,0) ELSE 0 END) as op_cr,
              0 as mv_dr,
              SUM(CASE WHEN r.RT_BILL_DATE >= TO_DATE(:date_from, 'YYYY-MM-DD') AND r.RT_BILL_DATE < TO_DATE(:date_to, 'YYYY-MM-DD')+1 THEN NVL(im.I_QTY,0) * NVL(it.PRIMARY_COST,0) ELSE 0 END) as mv_cr
          FROM IAS20261.ITEM_MOVEMENT im
          JOIN IAS20261.IAS_ITM_MST it ON it.I_CODE = im.I_CODE
          JOIN IAS20261.IAS_RT_BILL_MST r
            ON r.RT_BILL_DOC_TYPE = im.BILL_DOC_TYPE 
           AND r.RT_BILL_NO = im.DOC_NO 
           AND r.RT_BILL_SER = im.DOC_SER
          WHERE (:rep_code IS NULL OR r.REP_CODE = :rep_code)
            AND im.DOC_TYPE = 3
            AND r.PREV_YEAR IS NULL
            AND NVL(im.I_QTY,0) > 0
        ),
        inv_cogs_ret_prev AS (
          SELECT 
              '311060001' as acc_code,
              0 as op_dr,
              SUM(CASE WHEN r.RT_BILL_DATE < TO_DATE(:date_from, 'YYYY-MM-DD') THEN NVL(im.I_QTY,0) * NVL(it.PRIMARY_COST,0) ELSE 0 END) as op_cr,
              0 as mv_dr,
              SUM(CASE WHEN r.RT_BILL_DATE >= TO_DATE(:date_from, 'YYYY-MM-DD') AND r.RT_BILL_DATE < TO_DATE(:date_to, 'YYYY-MM-DD')+1 THEN NVL(im.I_QTY,0) * NVL(it.PRIMARY_COST,0) ELSE 0 END) as mv_cr
          FROM IAS20261.ITEM_MOVEMENT im
          JOIN IAS20261.IAS_ITM_MST it ON it.I_CODE = im.I_CODE
          JOIN IAS20261.IAS_RT_BILL_MST r
            ON r.RT_BILL_DOC_TYPE = im.BILL_DOC_TYPE 
           AND r.RT_BILL_NO = im.DOC_NO 
           AND r.RT_BILL_SER = im.DOC_SER
          WHERE (:rep_code IS NULL OR r.REP_CODE = :rep_code)
            AND im.DOC_TYPE = 3
            AND r.PREV_YEAR IS NOT NULL
            AND NVL(im.I_QTY,0) > 0
        ),
        all_data AS (
          SELECT * FROM gl_base
          UNION ALL
          SELECT * FROM inv_cogs
          UNION ALL
          SELECT * FROM inv_cogs_ret
          UNION ALL
          SELECT * FROM inv_cogs_ret_prev
        )
        SELECT 
            d.acc_code AS "الرقم", 
            MAX(a.A_NAME) AS "الاسم",
            TO_CHAR(NULLIF(ROUND(NVL(SUM(d.op_dr),0),2), 0),'FM999,999,990.00') AS "الرصيد الافتتاحي مدين",
            TO_CHAR(NULLIF(ROUND(NVL(SUM(d.op_cr),0),2), 0),'FM999,999,990.00') AS "الرصيد الافتتاحي دائن",
            TO_CHAR(NULLIF(ROUND(NVL(SUM(d.mv_dr),0),2), 0),'FM999,999,990.00') AS "رصيد الحركة مدين",
            TO_CHAR(NULLIF(ROUND(NVL(SUM(d.mv_cr),0),2), 0),'FM999,999,990.00') AS "رصيد الحركة دائن",
            TO_CHAR(NULLIF(ROUND(
              CASE WHEN (NVL(SUM(d.op_dr),0) + NVL(SUM(d.mv_dr),0)) - (NVL(SUM(d.op_cr),0) + NVL(SUM(d.mv_cr),0)) > 0 
                   THEN (NVL(SUM(d.op_dr),0) + NVL(SUM(d.mv_dr),0)) - (NVL(SUM(d.op_cr),0) + NVL(SUM(d.mv_cr),0))
                   ELSE 0 END, 2), 0), 'FM999,999,990.00'
            ) AS "الأرصدة مدين",
            TO_CHAR(NULLIF(ROUND(
              CASE WHEN (NVL(SUM(d.op_cr),0) + NVL(SUM(d.mv_cr),0)) - (NVL(SUM(d.op_dr),0) + NVL(SUM(d.mv_dr),0)) > 0 
                   THEN (NVL(SUM(d.op_cr),0) + NVL(SUM(d.mv_cr),0)) - (NVL(SUM(d.op_dr),0) + NVL(SUM(d.mv_dr),0))
                   ELSE 0 END, 2), 0), 'FM999,999,990.00'
            ) AS "الأرصدة دائن"
        FROM all_data d
        LEFT JOIN IAS20261.ACCOUNT a ON a.A_CODE = d.acc_code
        GROUP BY d.acc_code
        ORDER BY d.acc_code
      """},

 ]},
 {"id":"dts","title":"التوزيع والمناديب","icon":"M3 13l3-7h7l3 4h4v5M3 13h17M6 18a2 2 0 100-4 2 2 0 000 4zm11 0a2 2 0 100-4 2 2 0 000 4z","reports":[
   {"id":"collections","title":"تحصيل المناديب","params":[DFROM,DTO,REP],"sql":"""
     SELECT p.REP_CODE AS "كود المندوب", MAX(s.REPRS_A_NAME) AS "اسم المندوب",
            COUNT(DISTINCT p.C_CODE) AS "عدد العملاء",
            TO_CHAR(SUM(NVL(p.CR_AMT,0)),'FM999,999,999,990.00') AS "إجمالي التحصيل"
     FROM IAS20261.IAS_POST_DTL p LEFT JOIN IAS20261.SALES_MAN s ON s.REPRS_CODE = p.REP_CODE
     WHERE NVL(p.DOC_POST,0)=1 AND p.C_CODE IS NOT NULL AND NVL(p.CR_AMT,0)>0 AND p.REP_CODE IS NOT NULL AND p.DOC_TYPE = 2
       AND p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       AND (:rep_code IS NULL OR p.REP_CODE = :rep_code)
     GROUP BY p.REP_CODE ORDER BY SUM(NVL(p.CR_AMT,0)) DESC FETCH FIRST 300 ROWS ONLY"""},
   {"id":"performance","title":"أداء المناديب","params":[DFROM,DTO,REP],"sql":"""
     SELECT * FROM (
       SELECT s.REPRS_CODE AS "كود المندوب", s.REPRS_A_NAME AS "اسم المندوب",
              NVL(sl.inv,0) AS "عدد الفواتير", NVL(sl.custs,0) AS "عدد العملاء",
              TO_CHAR(NVL(sl.sales,0),'FM999,999,999,990.00') AS "المبيعات بالضريبة",
              TO_CHAR(NVL(col.coll,0),'FM999,999,999,990.00') AS "التحصيل",
              TO_CHAR(NVL(sl.sales,0)-NVL(col.coll,0),'FM999,999,999,990.00') AS "صافي (مبيعات - تحصيل)"
       FROM IAS20261.SALES_MAN s
       LEFT JOIN (SELECT REP_CODE, SUM(NVL(BILL_AMT,0)-NVL(DISC_AMT,0)+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)) sales, COUNT(*) inv, COUNT(DISTINCT C_CODE) custs
                  FROM IAS20261.IAS_BILL_MST
                  WHERE BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1 AND BILL_DOC_TYPE IN (1,4)
                  GROUP BY REP_CODE) sl ON sl.REP_CODE = s.REPRS_CODE
       LEFT JOIN (SELECT REP_CODE, SUM(NVL(CR_AMT,0)) coll
                  FROM IAS20261.IAS_POST_DTL
                  WHERE NVL(DOC_POST,0)=1 AND C_CODE IS NOT NULL AND NVL(CR_AMT,0)>0 AND DOC_TYPE = 2
                    AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
                  GROUP BY REP_CODE) col ON col.REP_CODE = s.REPRS_CODE
       WHERE (sl.sales IS NOT NULL OR col.coll IS NOT NULL) AND (:rep_code IS NULL OR s.REPRS_CODE = :rep_code)
       ORDER BY NVL(sl.sales,0) DESC
     ) WHERE ROWNUM <= 300"""},
   {"id":"perf_aging","title":"أداء المناديب الموسّع (أعمار التحصيل)","params":[DFROM,DTO,REP],"sql":"""
     WITH custs AS (
       SELECT DISTINCT p.C_CODE FROM IAS20261.IAS_POST_DTL p
       WHERE NVL(p.DOC_POST,0)=1 AND NVL(p.CR_AMT,0)>0 AND p.DOC_TYPE IN (1,2)
         AND p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1),
     inv AS (
       SELECT p.C_CODE, p.DOC_DATE inv_date,
              SUM(NVL(p.DR_AMT,0)) OVER (PARTITION BY p.C_CODE ORDER BY p.DOC_DATE,p.DOC_NO,p.DOC_SER ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) inv_hi,
              SUM(NVL(p.DR_AMT,0)) OVER (PARTITION BY p.C_CODE ORDER BY p.DOC_DATE,p.DOC_NO,p.DOC_SER ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) - NVL(p.DR_AMT,0) inv_lo
       FROM IAS20261.IAS_POST_DTL p
       WHERE NVL(p.DOC_POST,0)=1 AND p.C_CODE IS NOT NULL AND NVL(p.DR_AMT,0)>0
         AND p.C_CODE IN (SELECT C_CODE FROM custs)),
     col AS (
       SELECT p.C_CODE, p.DOC_DATE coll_date, p.DOC_TYPE dtype,
              SUM(NVL(p.CR_AMT,0)) OVER (PARTITION BY p.C_CODE ORDER BY p.DOC_DATE,p.DOC_NO,p.DOC_SER ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) coll_hi,
              SUM(NVL(p.CR_AMT,0)) OVER (PARTITION BY p.C_CODE ORDER BY p.DOC_DATE,p.DOC_NO,p.DOC_SER ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) - NVL(p.CR_AMT,0) coll_lo
       FROM IAS20261.IAS_POST_DTL p
       WHERE NVL(p.DOC_POST,0)=1 AND p.C_CODE IS NOT NULL AND NVL(p.CR_AMT,0)>0
         AND p.C_CODE IN (SELECT C_CODE FROM custs)),
     mt AS (
       SELECT co.C_CODE, co.coll_date, iv.inv_date,
              GREATEST(0, LEAST(co.coll_hi, iv.inv_hi) - GREATEST(co.coll_lo, iv.inv_lo)) amt
       FROM col co JOIN inv iv ON iv.C_CODE=co.C_CODE AND iv.inv_lo < co.coll_hi AND iv.inv_hi > co.coll_lo
       WHERE co.coll_date >= TO_DATE(:date_from,'YYYY-MM-DD') AND co.coll_date < TO_DATE(:date_to,'YYYY-MM-DD')+1 AND co.dtype IN (1,2)),
     s AS (
       SELECT C_CODE, amt,
              CASE WHEN TRUNC(mt.inv_date) <= TRUNC(mt.coll_date)
                   THEN TRUNC(mt.coll_date)-TRUNC(mt.inv_date) ELSE 0 END age
       FROM mt
       WHERE amt > 0
         AND ( TRUNC(mt.inv_date) <= TRUNC(mt.coll_date)
               OR mt.inv_date >= TO_DATE(:date_to,'YYYY-MM-DD')+1 ))
     SELECT * FROM (
       SELECT c.REP_CODE AS "كود المندوب", MAX(sm.REPRS_A_NAME) AS "اسم المندوب",
              COUNT(DISTINCT s.C_CODE) AS "عدد العملاء",
              TO_CHAR(SUM(CASE WHEN s.age<=30 THEN s.amt ELSE 0 END),'FM999,999,990.00') AS "0-30",
              TO_CHAR(SUM(CASE WHEN s.age BETWEEN 31 AND 60 THEN s.amt ELSE 0 END),'FM999,999,990.00') AS "31-60",
              TO_CHAR(SUM(CASE WHEN s.age BETWEEN 61 AND 90 THEN s.amt ELSE 0 END),'FM999,999,990.00') AS "61-90",
              TO_CHAR(SUM(CASE WHEN s.age BETWEEN 91 AND 120 THEN s.amt ELSE 0 END),'FM999,999,990.00') AS "91-120",
              TO_CHAR(SUM(CASE WHEN s.age>120 THEN s.amt ELSE 0 END),'FM999,999,990.00') AS "أكثر من 120",
              TO_CHAR(SUM(CASE WHEN s.age IS NOT NULL THEN s.amt ELSE 0 END),'FM999,999,990.00') AS "المبلغ المحصل"
       FROM s JOIN IAS20261.CUSTOMER c ON c.C_CODE=s.C_CODE
       LEFT JOIN IAS20261.SALES_MAN sm ON sm.REPRS_CODE=c.REP_CODE
       WHERE (:rep_code IS NULL OR c.REP_CODE = :rep_code)
       GROUP BY c.REP_CODE ORDER BY SUM(s.amt) DESC
     ) WHERE ROWNUM <= 300"""},
   {"id":"perf_aging_exact","title":"أعمار التحصيل (مطابق أونكس 100%)","params":[DFROM,DTO,REP],"sql":"""
     WITH cr AS (
       SELECT C_CODE,
         SUM(CASE WHEN PER_NO BETWEEN 0 AND 30   THEN DR_AMT ELSE 0 END) pos,
         SUM(CASE WHEN PER_NO BETWEEN -30 AND -1 THEN DR_AMT ELSE 0 END) neg,
         SUM(CASE WHEN PER_NO BETWEEN 31 AND 60  THEN DR_AMT ELSE 0 END) b2,
         SUM(CASE WHEN PER_NO BETWEEN 61 AND 90  THEN DR_AMT ELSE 0 END) b3,
         SUM(CASE WHEN PER_NO BETWEEN 91 AND 120 THEN DR_AMT ELSE 0 END) b4,
         SUM(CASE WHEN PER_NO > 120              THEN DR_AMT ELSE 0 END) b5,
         SUM(DR_AMT) crlim
       FROM IAS20261.IAS_CRLIMIT_TMP
       WHERE DOC_TYPE<>1 AND DOC_TYPE_REF IN (1,2)
         AND PAID_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND PAID_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       GROUP BY C_CODE),
     co AS (
       SELECT C_CODE, SUM(CR_AMT) col FROM IAS20261.IAS_COL_TMP
       WHERE DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       GROUP BY C_CODE),
     perc AS (
       SELECT cr.C_CODE,
         CASE WHEN cr.pos > 0 THEN cr.pos + cr.neg + GREATEST(0, NVL(co.col,0) - cr.crlim) ELSE 0 END b1,
         cr.b2, cr.b3, cr.b4, cr.b5
       FROM cr LEFT JOIN co ON co.C_CODE = cr.C_CODE)
     SELECT * FROM (
       SELECT c.REP_CODE AS "كود المندوب", MAX(sm.REPRS_A_NAME) AS "اسم المندوب",
              COUNT(DISTINCT perc.C_CODE) AS "عدد العملاء",
              TO_CHAR(SUM(perc.b1),'FM999,999,990.00') AS "0-30",
              TO_CHAR(SUM(perc.b2),'FM999,999,990.00') AS "31-60",
              TO_CHAR(SUM(perc.b3),'FM999,999,990.00') AS "61-90",
              TO_CHAR(SUM(perc.b4),'FM999,999,990.00') AS "91-120",
              TO_CHAR(SUM(perc.b5),'FM999,999,990.00') AS "أكثر من 120",
              TO_CHAR(SUM(perc.b1+perc.b2+perc.b3+perc.b4+perc.b5),'FM999,999,990.00') AS "إجمالي التحصيل"
       FROM perc JOIN IAS20261.CUSTOMER c ON c.C_CODE=perc.C_CODE
       LEFT JOIN IAS20261.SALES_MAN sm ON sm.REPRS_CODE=c.REP_CODE
       WHERE (:rep_code IS NULL OR c.REP_CODE = :rep_code)
       GROUP BY c.REP_CODE ORDER BY SUM(perc.b1+perc.b2+perc.b3+perc.b4+perc.b5) DESC
     ) WHERE ROWNUM <= 300"""},
       {"id":"collection_adopted","title":"التحصيل المعتمد (ديناميكي)","params":[DFROM,DTO,GRP,REP,INCR,INCN,INCC,INCRT,INCEX],"sql":"""
     WITH 
     grp AS (
       SELECT 'rep' as typ, TO_CHAR(REPRS_CODE) as cd, MAX(REPRS_A_NAME) as nm FROM IAS20261.SALES_MAN GROUP BY TO_CHAR(REPRS_CODE)
       UNION ALL 
       SELECT 'cc' as typ, TO_CHAR(CC_CODE) as cd, MAX(CC_A_NAME) as nm FROM IAS20261.COST_CENTERS GROUP BY TO_CHAR(CC_CODE)
       UNION ALL
       SELECT 'cst' as typ, TO_CHAR(C_CODE) as cd, MAX(C_A_NAME) as nm FROM IAS20261.CUSTOMER GROUP BY TO_CHAR(C_CODE)
       UNION ALL
       SELECT 'cst' as typ, 'UNKNOWN' as cd, 'عميل نقدي عام' as nm FROM DUAL
     ),
     all_trans AS (
       SELECT CASE WHEN :grp_by='cc' THEN TO_CHAR(CC_CODE) WHEN :grp_by='cst' THEN NVL(TO_CHAR(C_CODE),'UNKNOWN') ELSE TO_CHAR(REP_CODE) END as grp_code,
              CR_AMT as rcpt, 0 as net_jrn, 0 as cash_sales, 0 as inv_disc, 0 as cash_ret, 0 as ext_notice
       FROM IAS20261.IAS_POST_DTL
       WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=2 AND NVL(CR_AMT,0)>0
         AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       UNION ALL
       SELECT CASE WHEN :grp_by='cc' THEN TO_CHAR(CC_CODE) WHEN :grp_by='cst' THEN NVL(TO_CHAR(C_CODE),'UNKNOWN') ELSE TO_CHAR(REP_CODE) END,
              0, CR_AMT, 0, 0, 0, 0
       FROM IAS20261.IAS_POST_DTL
       WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=1 AND JV_TYPE=2 AND NVL(CR_AMT,0)>0
         AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       UNION ALL
       SELECT CASE WHEN :grp_by='cc' THEN TO_CHAR(CC_CODE) WHEN :grp_by='cst' THEN NVL(TO_CHAR(C_CODE),'UNKNOWN') ELSE TO_CHAR(REP_CODE) END,
              0, 0, NVL(BILL_AMT,0)-NVL(DISC_AMT,0)+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0), NVL(DISC_AMT,0), 0, 0
       FROM IAS20261.IAS_BILL_MST
       WHERE BILL_DOC_TYPE=1
         AND BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       UNION ALL
       SELECT CASE WHEN :grp_by='cc' THEN TO_CHAR(CC_CODE) WHEN :grp_by='cst' THEN NVL(TO_CHAR(C_CODE),'UNKNOWN') ELSE TO_CHAR(REP_CODE) END,
              0, 0, 0, 0, CR_AMT, 0
       FROM IAS20261.IAS_POST_DTL
       WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=5 AND A_CODE LIKE '111%' AND NVL(CR_AMT,0)>0
         AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       UNION ALL
       SELECT CASE WHEN :grp_by='cc' THEN TO_CHAR(CC_CODE) WHEN :grp_by='cst' THEN NVL(TO_CHAR(C_CODE),'UNKNOWN') ELSE TO_CHAR(REP_CODE) END,
              0, 0, 0, 0, 0, CR_AMT
       FROM IAS20261.IAS_POST_DTL
       WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=15 AND NVL(CR_AMT,0)>0
         AND DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
     ),
     base AS (
       SELECT grp_code,
              SUM(rcpt) rcpt, SUM(net_jrn) net_jrn, SUM(cash_sales) cash_sales, SUM(inv_disc) inv_disc, SUM(cash_ret) cash_ret, SUM(ext_notice) ext_notice,
              (CASE WHEN :inc_rcpt='1' THEN SUM(rcpt) ELSE 0 END
             + CASE WHEN :inc_net='1'  THEN SUM(net_jrn) ELSE 0 END
             + CASE WHEN :inc_cash='1' THEN SUM(cash_sales) ELSE 0 END
             - CASE WHEN :inc_ret='1'  THEN SUM(cash_ret) ELSE 0 END
             - CASE WHEN :inc_ext='1'  THEN SUM(ext_notice) ELSE 0 END) total_inc
       FROM all_trans
       WHERE grp_code IS NOT NULL
         AND (:rep_code IS NULL OR (:grp_by = 'rep' AND grp_code = :rep_code))
       GROUP BY grp_code
     )
     SELECT * FROM (
       SELECT b.grp_code AS "الكود", NVL(MAX(g.nm), b.grp_code) AS "الجهة / الاسم",
              TO_CHAR(MAX(b.rcpt),'FM999,999,990.00')      AS "سندات القبض",
              TO_CHAR(MAX(b.net_jrn),'FM999,999,990.00')   AS "قيود الشبكة المنفصلة",
              TO_CHAR(MAX(b.cash_sales),'FM999,999,990.00') AS "المبيعات النقدية",
              TO_CHAR(MAX(b.inv_disc),'FM999,999,990.00')   AS "الخصم في الفاتورة",
              TO_CHAR(MAX(b.ext_notice),'FM999,999,990.00') AS "إشعار خصم مستقل (-)",
              TO_CHAR(MAX(b.cash_ret),'FM999,999,990.00')   AS "المرتجع النقدي (-)",
              TO_CHAR(MAX(b.total_inc),'FM999,999,990.00') AS "إجمالي التحصيل"
       FROM base b
       LEFT JOIN grp g ON g.cd = b.grp_code AND g.typ = :grp_by
       WHERE (b.rcpt > 0 OR b.net_jrn > 0 OR b.cash_sales > 0 OR b.cash_ret > 0 OR b.inv_disc > 0 OR b.ext_notice > 0)
       GROUP BY b.grp_code
       ORDER BY MAX(b.total_inc) DESC
     ) WHERE ROWNUM <= 300"""},
 ]},
 {"id":"pur","title":"المشتريات والموردون","icon":"M6 6h15l-1.5 9h-12zM6 6L5 3H2M9 20a1 1 0 100-2 1 1 0 000 2zm9 0a1 1 0 100-2 1 1 0 000 2z","reports":[
   {"id":"pi_bills","title":"فواتير المشتريات","params":[DFROM,DTO,{"name":"v_code","label":"المورد (اختياري)","type":"text","default":""}],"sql":"""
     SELECT BILL_NO AS "رقم الفاتورة", TO_CHAR(BILL_DATE,'YYYY-MM-DD') AS "التاريخ",
            V_CODE AS "كود المورد", V_NAME AS "اسم المورد",
            TO_CHAR(NVL(BILL_AMT,0),'FM999,999,990.00') AS "المبلغ",
            TO_CHAR(NVL(DISC_AMT,0),'FM999,999,990.00') AS "الخصم",
            TO_CHAR(NVL(VAT_AMT,0),'FM999,999,990.00') AS "الضريبة",
            TO_CHAR(NVL(BILL_AMT,0)-NVL(DISC_AMT,0)+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0),'FM999,999,990.00') AS "الصافي",
            CASE NVL(BILL_POST,0) WHEN 1 THEN 'مرحّلة' ELSE 'غير مرحّلة' END AS "الحالة"
     FROM IAS20261.IAS_PI_BILL_MST
     WHERE BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       AND (:v_code IS NULL OR V_CODE = :v_code)
     ORDER BY BILL_DATE DESC, BILL_NO DESC FETCH FIRST 300 ROWS ONLY"""},
   {"id":"pi_by_vendor","title":"حسب المورد","params":[DFROM,DTO],"sql":"""
     SELECT V_CODE AS "كود المورد", MAX(V_NAME) AS "اسم المورد", COUNT(*) AS "عدد الفواتير",
            TO_CHAR(SUM(NVL(BILL_AMT,0)-NVL(DISC_AMT,0)),'FM999,999,999,990.00') AS "صافي قبل الضريبة",
            TO_CHAR(SUM(NVL(BILL_AMT,0)-NVL(DISC_AMT,0)+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)),'FM999,999,999,990.00') AS "الإجمالي بالضريبة"
     FROM IAS20261.IAS_PI_BILL_MST
     WHERE BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
     GROUP BY V_CODE ORDER BY SUM(NVL(BILL_AMT,0)-NVL(DISC_AMT,0)) DESC FETCH FIRST 300 ROWS ONLY"""},
   {"id":"pi_by_item","title":"حسب الصنف","params":[DFROM,DTO,{"name":"i_code","label":"الصنف (اختياري)","type":"text","default":""}],"sql":"""
     SELECT dt.I_CODE AS "كود الصنف", MAX(m.I_NAME) AS "اسم الصنف",
            ROUND(SUM(NVL(dt.I_QTY,0)),2) AS "إجمالي الكمية",
            TO_CHAR(ROUND(SUM(NVL(dt.I_QTY,0)*NVL(dt.I_PRICE,0)),2),'FM999,999,999,990.00') AS "قيمة المشتريات",
            COUNT(DISTINCT b.BILL_NO) AS "عدد الفواتير"
     FROM IAS20261.IAS_PI_BILL_DTL dt
     JOIN IAS20261.IAS_PI_BILL_MST b ON b.BILL_DOC_TYPE=dt.BILL_DOC_TYPE AND b.BILL_NO=dt.BILL_NO AND b.BILL_SER=dt.BILL_SER
     LEFT JOIN IAS20261.IAS_ITM_MST m ON m.I_CODE=dt.I_CODE
     WHERE b.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND b.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       AND (:i_code IS NULL OR dt.I_CODE = :i_code)
     GROUP BY dt.I_CODE ORDER BY SUM(NVL(dt.I_QTY,0)*NVL(dt.I_PRICE,0)) DESC FETCH FIRST 300 ROWS ONLY"""},
   {"id":"vendor_statement","title":"كشف حساب مورد","params":[{"name":"v_code","label":"كود المورد","type":"text","default":"222"},DFROM,DTO],"sql":"""
     SELECT "التاريخ","نوع المستند","رقم المستند","البيان","مدين","دائن","الرصيد" FROM (
       SELECT TO_CHAR(p.DOC_DATE,'YYYY-MM-DD') AS "التاريخ", d.JV_NAME AS "نوع المستند",
              p.DOC_NO AS "رقم المستند", p.DOC_DESC AS "البيان",
              TO_CHAR(NVL(p.DR_AMT,0),'FM999,999,990.00') AS "مدين",
              TO_CHAR(NVL(p.CR_AMT,0),'FM999,999,990.00') AS "دائن",
              TO_CHAR(SUM(NVL(p.CR_AMT,0)-NVL(p.DR_AMT,0)) OVER (ORDER BY p.DOC_DATE,p.DOC_NO,p.DOC_SER),'FM999,999,990.00') AS "الرصيد",
              p.DOC_DATE s1, p.DOC_NO s2, p.DOC_SER s3
       FROM IAS20261.IAS_POST_DTL p
       LEFT JOIN IAS_SYS.IAS_DOCJV_TYPE_SYSTEMS d ON d.DOC_TYPE=p.DOC_TYPE AND d.JV_TYPE=p.JV_TYPE AND d.LANG_NO=1
       WHERE p.V_CODE = :v_code AND NVL(p.DOC_POST,0)=1
         AND p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       ORDER BY p.DOC_DATE, p.DOC_NO, p.DOC_SER
     ) FETCH FIRST 1000 ROWS ONLY"""},
   {"id":"vendor_aging","title":"أعمار الدائنين","params":[{"name":"as_of","label":"حتى تاريخ","type":"date","default":"2026-07-10"}],"sql":"""
     WITH pay AS (SELECT V_CODE, SUM(NVL(DR_AMT,0)) paid FROM IAS20261.IAS_POST_DTL
                  WHERE NVL(DOC_POST,0)=1 AND V_CODE IS NOT NULL GROUP BY V_CODE),
     charges AS (SELECT p.V_CODE, p.DOC_DATE, NVL(p.CR_AMT,0) amt,
                   SUM(NVL(p.CR_AMT,0)) OVER (PARTITION BY p.V_CODE ORDER BY p.DOC_DATE,p.DOC_NO,p.DOC_SER
                        ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) cum
                 FROM IAS20261.IAS_POST_DTL p WHERE NVL(p.DOC_POST,0)=1 AND p.V_CODE IS NOT NULL AND NVL(p.CR_AMT,0)>0),
     openit AS (SELECT ch.V_CODE, GREATEST(0,LEAST(ch.amt,ch.cum-NVL(pay.paid,0))) unpaid,
                   TRUNC(TO_DATE(:as_of,'YYYY-MM-DD'))-TRUNC(ch.DOC_DATE) age
                FROM charges ch JOIN pay ON pay.V_CODE=ch.V_CODE)
     SELECT o.V_CODE AS "كود المورد", MAX(v.V_NAME) AS "اسم المورد",
            TO_CHAR(SUM(CASE WHEN o.age<=30 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "0-30",
            TO_CHAR(SUM(CASE WHEN o.age BETWEEN 31 AND 60 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "31-60",
            TO_CHAR(SUM(CASE WHEN o.age BETWEEN 61 AND 90 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "61-90",
            TO_CHAR(SUM(CASE WHEN o.age BETWEEN 91 AND 120 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "91-120",
            TO_CHAR(SUM(CASE WHEN o.age>120 THEN o.unpaid ELSE 0 END),'FM999,999,990.00') AS "أكثر من 120",
            TO_CHAR(SUM(o.unpaid),'FM999,999,990.00') AS "الإجمالي"
     FROM openit o LEFT JOIN (SELECT V_CODE, MAX(V_NAME) V_NAME FROM IAS20261.IAS_PI_BILL_MST GROUP BY V_CODE) v ON v.V_CODE=o.V_CODE
     WHERE o.unpaid>0 GROUP BY o.V_CODE ORDER BY SUM(o.unpaid) DESC FETCH FIRST 300 ROWS ONLY"""},
 ]},
 {"id":"fin","title":"المالية والمحاسبة","icon":"M4 20V4h16v16zM8 16v-4M12 16V8M16 16v-6","reports":[
   {"id":"trial_balance","title":"ميزان المراجعة","params":[DFROM,DTO],"sql":"""
     SELECT * FROM (
       SELECT p.A_CODE AS "رقم الحساب", MAX(a.A_NAME) AS "اسم الحساب",
              TO_CHAR(SUM(NVL(p.DR_AMT,0)),'FM999,999,999,990.00') AS "إجمالي مدين",
              TO_CHAR(SUM(NVL(p.CR_AMT,0)),'FM999,999,999,990.00') AS "إجمالي دائن",
              TO_CHAR(SUM(NVL(p.DR_AMT,0)-NVL(p.CR_AMT,0)),'FM999,999,999,990.00') AS "الرصيد"
       FROM IAS20261.IAS_POST_DTL p LEFT JOIN IAS20261.ACCOUNT a ON a.A_CODE=p.A_CODE
       WHERE NVL(p.DOC_POST,0)=1 AND p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       GROUP BY p.A_CODE HAVING SUM(NVL(p.DR_AMT,0))<>0 OR SUM(NVL(p.CR_AMT,0))<>0
       ORDER BY p.A_CODE
     ) WHERE ROWNUM <= 500"""},
   {"id":"income_statement","title":"قائمة الدخل","params":[DFROM,DTO],"sql":"""
     SELECT NVL(pa.A_NAME, a.A_PARENT) AS "البند",
            TO_CHAR(SUM(NVL(p.CR_AMT,0)-NVL(p.DR_AMT,0)),'FM999,999,999,990.00') AS "الصافي"
     FROM IAS20261.IAS_POST_DTL p
     JOIN IAS20261.ACCOUNT a ON a.A_CODE=p.A_CODE
     LEFT JOIN IAS20261.ACCOUNT pa ON pa.A_CODE=a.A_PARENT
     WHERE NVL(p.DOC_POST,0)=1 AND a.A_REPORT=2
       AND p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
     GROUP BY a.A_PARENT, pa.A_NAME
     ORDER BY SUM(NVL(p.CR_AMT,0)-NVL(p.DR_AMT,0)) DESC"""},
   {"id":"cost_centers","title":"مراكز التكلفة","params":[DFROM,DTO],"sql":"""
     SELECT * FROM (
       SELECT p.CC_CODE AS "مركز التكلفة", MAX(cc.CC_A_NAME) AS "الاسم",
              TO_CHAR(SUM(NVL(p.CR_AMT,0)-NVL(p.DR_AMT,0)),'FM999,999,999,990.00') AS "صافي الربح/الخسارة"
       FROM IAS20261.IAS_POST_DTL p JOIN IAS20261.ACCOUNT a ON a.A_CODE=p.A_CODE
       LEFT JOIN IAS20261.COST_CENTERS cc ON cc.CC_CODE=p.CC_CODE
       WHERE NVL(p.DOC_POST,0)=1 AND a.A_REPORT=2
         AND p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       GROUP BY p.CC_CODE HAVING SUM(NVL(p.CR_AMT,0)-NVL(p.DR_AMT,0))<>0
       ORDER BY SUM(NVL(p.CR_AMT,0)-NVL(p.DR_AMT,0)) DESC
     ) WHERE ROWNUM <= 200"""},
   {"id":"journal","title":"قيود اليومية","params":[{"name":"date_from","label":"من تاريخ","type":"date","default":"2026-07-01"},{"name":"date_to","label":"إلى تاريخ","type":"date","default":"2026-07-10"},{"name":"a_code","label":"الحساب (اختياري)","type":"text","default":""}],"sql":"""
     SELECT * FROM (
       SELECT TO_CHAR(p.DOC_DATE,'YYYY-MM-DD') AS "التاريخ", d.JV_NAME AS "نوع القيد",
              p.DOC_NO AS "رقم المستند", p.A_CODE AS "رقم الحساب", a.A_NAME AS "اسم الحساب",
              p.DOC_DESC AS "البيان",
              TO_CHAR(NVL(p.DR_AMT,0),'FM999,999,990.00') AS "مدين",
              TO_CHAR(NVL(p.CR_AMT,0),'FM999,999,990.00') AS "دائن"
       FROM IAS20261.IAS_POST_DTL p
       LEFT JOIN IAS20261.ACCOUNT a ON a.A_CODE=p.A_CODE
       LEFT JOIN IAS20261.JV_TYPES d ON d.JV_TYPE=p.JV_TYPE
       WHERE NVL(p.DOC_POST,0)=1 AND p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND (:a_code IS NULL OR p.A_CODE = :a_code)
       ORDER BY p.DOC_DATE DESC, p.DOC_NO DESC, p.DOC_SER
     ) WHERE ROWNUM <= 300"""},
 ]},
 {"id":"tax","title":"الضريبة","icon":"M4 4h16v4H4zM6 8v12M18 8v12M4 20h16M9 12h6M9 16h6","reports":[
   {"id":"vat_decl","title":"الإقرار الضريبي (شهري)","params":[],"sql":"""
     SELECT PRD_NM AS "الفترة",
       TO_CHAR(-(SUM(CASE WHEN DOC_TYPE IN (4,5,15) THEN DOC_AMT_VAT ELSE 0 END)),'FM999,999,999,990.00') AS "المبيعات الخاضعة",
       TO_CHAR(-(SUM(CASE WHEN DOC_TYPE IN (4,5,15) THEN VAT_AMT ELSE 0 END)),'FM999,999,999,990.00') AS "ضريبة المخرجات",
       TO_CHAR(SUM(CASE WHEN DOC_TYPE IN (6,7,16) THEN DOC_AMT_VAT ELSE 0 END),'FM999,999,999,990.00') AS "المشتريات الخاضعة",
       TO_CHAR(SUM(CASE WHEN DOC_TYPE IN (6,7,16) THEN VAT_AMT ELSE 0 END),'FM999,999,999,990.00') AS "ضريبة المدخلات",
       TO_CHAR(SUM(CASE WHEN DOC_TYPE IN (1,3) THEN VAT_AMT ELSE 0 END),'FM999,999,999,990.00') AS "تعديلات",
       TO_CHAR(-(SUM(VAT_AMT)),'FM999,999,999,990.00') AS "صافي الضريبة المستحقة"
     FROM IAS20261.GNR_TAX_SUM_VW
     GROUP BY PRD_NO, PRD_NM ORDER BY PRD_NO"""},
   {"id":"vat_out","title":"تفصيل ضريبة المخرجات","params":[],"sql":"""
     SELECT PRD_NM AS "الفترة", DOC_TYP_NAME AS "نوع المستند",
       TO_CHAR(-(SUM(DOC_AMT_VAT)),'FM999,999,999,990.00') AS "الوعاء الخاضع",
       TO_CHAR(-(SUM(VAT_AMT)),'FM999,999,999,990.00') AS "الضريبة"
     FROM IAS20261.GNR_TAX_SUM_VW WHERE DOC_TYPE IN (4,5,15)
     GROUP BY PRD_NO, PRD_NM, DOC_TYPE, DOC_TYP_NAME ORDER BY PRD_NO, DOC_TYPE"""},
   {"id":"vat_in","title":"تفصيل ضريبة المدخلات","params":[],"sql":"""
     SELECT PRD_NM AS "الفترة", DOC_TYP_NAME AS "نوع المستند",
       TO_CHAR(SUM(DOC_AMT_VAT),'FM999,999,999,990.00') AS "الوعاء الخاضع",
       TO_CHAR(SUM(VAT_AMT),'FM999,999,999,990.00') AS "الضريبة"
     FROM IAS20261.GNR_TAX_SUM_VW WHERE DOC_TYPE IN (6,7,16,1,3)
     GROUP BY PRD_NO, PRD_NM, DOC_TYPE, DOC_TYP_NAME ORDER BY PRD_NO, DOC_TYPE"""},
 ]},
 {"id":"prof","title":"الربحية","icon":"M3 3v18h18M7 14l3-4 3 3 5-6","reports":[
   {"id":"prof_summary","title":"ملخّص مجمل الربح للفترة","params":[DFROM,DTO,REP],"sql":"""
     WITH s AS (
       SELECT CASE WHEN m.BILL_DOC_TYPE IN (1,4) THEN 1 WHEN m.BILL_DOC_TYPE IN (2,5) THEN -1 ELSE 0 END as sign,
              NVL(d.I_QTY,0) as qty, NVL(d.I_PRICE,0) as price, NVL(d.DIS_AMT,0) as line_disc,
              CASE WHEN NVL(m.BILL_AMT,0)=0 THEN 0 ELSE ((NVL(d.I_QTY,0)*NVL(d.I_PRICE,0))/m.BILL_AMT)*NVL(m.DISC_AMT,0) END as hdr_disc,
              NVL(d.OTHR_AMT,0) as othr, NVL(d.STK_COST,0) as unit_cost
       FROM IAS20261.IAS_BILL_DTL d JOIN IAS20261.IAS_BILL_MST m ON m.BILL_SER=d.BILL_SER
       WHERE m.BILL_DOC_TYPE IN (1,4,2,5)
         AND m.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND m.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND (:rep_code IS NULL OR m.REP_CODE = :rep_code)
     ),
     ext AS (
       SELECT NVL(CR_AMT,0) as ext_disc
       FROM IAS20261.IAS_POST_DTL
       WHERE DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND DOC_TYPE = 15 AND NVL(CR_AMT,0) > 0 AND NVL(DOC_POST,0) = 1
         AND (:rep_code IS NULL OR REP_CODE = :rep_code)
     )
     SELECT
       TO_CHAR(SUM(rev),'FM999,999,999,990.00') AS "المبيعات (بلا ضريبة)",
       TO_CHAR(SUM(cst),'FM999,999,999,990.00') AS "تكلفة المبيعات",
       TO_CHAR(SUM(rev)-SUM(cst),'FM999,999,999,990.00') AS "مجمل الربح",
       TO_CHAR(ROUND(100*(SUM(rev)-SUM(cst))/NULLIF(SUM(rev),0),1),'FM990.0')||' %' AS "الهامش"
     FROM (
       SELECT SUM(((qty*price) - line_disc - hdr_disc + othr) * sign) - (SELECT NVL(SUM(ext_disc),0) FROM ext) as rev,
              SUM(qty * unit_cost * sign) as cst
       FROM s
     )"""},
   {"id":"net_profit","title":"صافي الربح للفترة (بعد كل المصاريف)","params":[DFROM,DTO],"sql":"""
     SELECT
       TO_CHAR(SUM(CASE WHEN nt>0 THEN nt ELSE 0 END),'FM999,999,999,990.00') AS "الإيرادات",
       TO_CHAR(SUM(CASE WHEN nt<0 THEN -nt ELSE 0 END),'FM999,999,999,990.00') AS "المصاريف",
       TO_CHAR(SUM(nt),'FM999,999,999,990.00') AS "صافي الربح"
     FROM (SELECT p.A_CODE, SUM(NVL(p.CR_AMT,0)-NVL(p.DR_AMT,0)) nt
           FROM IAS20261.IAS_POST_DTL p JOIN IAS20261.ACCOUNT a ON a.A_CODE=p.A_CODE
           WHERE NVL(p.DOC_POST,0)=1 AND a.A_REPORT=2
             AND p.DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND p.DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
           GROUP BY p.A_CODE)"""},
   {"id":"prof_item","title":"ربحية الصنف","params":[DFROM,DTO,REP],"sql":"""
     WITH s AS (
       SELECT d.I_CODE,
              CASE WHEN m.BILL_DOC_TYPE IN (1,4) THEN 1 WHEN m.BILL_DOC_TYPE IN (2,5) THEN -1 ELSE 0 END as sign,
              NVL(d.I_QTY,0) as qty, NVL(d.I_PRICE,0) as price, NVL(d.DIS_AMT,0) as line_disc,
              CASE WHEN NVL(m.BILL_AMT,0)=0 THEN 0 ELSE ((NVL(d.I_QTY,0)*NVL(d.I_PRICE,0))/m.BILL_AMT)*NVL(m.DISC_AMT,0) END as hdr_disc,
              NVL(d.OTHR_AMT,0) as othr, NVL(d.STK_COST,0) as unit_cost
       FROM IAS20261.IAS_BILL_DTL d JOIN IAS20261.IAS_BILL_MST m ON m.BILL_SER=d.BILL_SER
       WHERE m.BILL_DOC_TYPE IN (1,4,2,5)
         AND m.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND m.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND (:rep_code IS NULL OR m.REP_CODE = :rep_code)
     )
     SELECT * FROM (
       SELECT s.I_CODE AS "كود الصنف", MAX(i.I_NAME) AS "اسم الصنف",
              TO_CHAR(SUM(qty * sign),'FM999,999,990.00') AS "الكمية المباعة",
              TO_CHAR(SUM(((qty*price) - line_disc - hdr_disc + othr) * sign),'FM999,999,999,990.00') AS "المبيعات",
              TO_CHAR(SUM(qty * unit_cost * sign),'FM999,999,999,990.00') AS "التكلفة",
              TO_CHAR(SUM(((qty*price) - line_disc - hdr_disc + othr) * sign) - SUM(qty * unit_cost * sign),'FM999,999,999,990.00') AS "الربح",
              TO_CHAR(ROUND(100*(SUM(((qty*price) - line_disc - hdr_disc + othr) * sign) - SUM(qty * unit_cost * sign))/NULLIF(SUM(((qty*price) - line_disc - hdr_disc + othr) * sign),0),1),'FM990.0')||' %' AS "هامش"
       FROM s LEFT JOIN IAS20261.IAS_ITM_MST i ON i.I_CODE=s.I_CODE
       GROUP BY s.I_CODE ORDER BY SUM(((qty*price) - line_disc - hdr_disc + othr) * sign) - SUM(qty * unit_cost * sign) DESC
     ) WHERE ROWNUM<=300"""},
   {"id":"prof_cust","title":"ربحية العميل","params":[DFROM,DTO,REP],"sql":"""
     WITH s AS (
       SELECT m.C_CODE,
              CASE WHEN m.BILL_DOC_TYPE IN (1,4) THEN 1 WHEN m.BILL_DOC_TYPE IN (2,5) THEN -1 ELSE 0 END as sign,
              ((NVL(d.I_QTY,0)*NVL(d.I_PRICE,0)) - NVL(d.DIS_AMT,0) - (CASE WHEN NVL(m.BILL_AMT,0)=0 THEN 0 ELSE ((NVL(d.I_QTY,0)*NVL(d.I_PRICE,0))/m.BILL_AMT)*NVL(m.DISC_AMT,0) END) + NVL(d.OTHR_AMT,0)) as rev,
              (NVL(d.I_QTY,0)*NVL(d.STK_COST,0)) as cst,
              0 as ext_disc
       FROM IAS20261.IAS_BILL_DTL d JOIN IAS20261.IAS_BILL_MST m ON m.BILL_SER=d.BILL_SER
       WHERE m.BILL_DOC_TYPE IN (1,4,2,5)
         AND m.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND m.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND (:rep_code IS NULL OR m.REP_CODE = :rep_code)
       UNION ALL
       SELECT C_CODE, 1 as sign, 0 as rev, 0 as cst, NVL(CR_AMT,0) as ext_disc
       FROM IAS20261.IAS_POST_DTL
       WHERE DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND DOC_TYPE = 15 AND NVL(CR_AMT,0) > 0 AND NVL(DOC_POST,0) = 1
         AND (:rep_code IS NULL OR REP_CODE = :rep_code)
     )
     SELECT * FROM (
       SELECT s.C_CODE AS "كود العميل", MAX(c.C_A_NAME) AS "اسم العميل",
              TO_CHAR(SUM(rev * sign) - SUM(ext_disc),'FM999,999,999,990.00') AS "المبيعات",
              TO_CHAR(SUM(cst * sign),'FM999,999,999,990.00') AS "التكلفة",
              TO_CHAR((SUM(rev * sign) - SUM(ext_disc)) - SUM(cst * sign),'FM999,999,999,990.00') AS "الربح",
              TO_CHAR(ROUND(100*((SUM(rev * sign) - SUM(ext_disc)) - SUM(cst * sign))/NULLIF(SUM(rev * sign) - SUM(ext_disc),0),1),'FM990.0')||' %' AS "هامش"
       FROM s LEFT JOIN IAS20261.CUSTOMER c ON c.C_CODE=s.C_CODE
       WHERE s.C_CODE IS NOT NULL
       GROUP BY s.C_CODE ORDER BY (SUM(rev * sign) - SUM(ext_disc)) - SUM(cst * sign) DESC
     ) WHERE ROWNUM<=300"""},
   {"id":"prof_rep","title":"ربحية المندوب","params":[DFROM,DTO],"sql":"""
     WITH s AS (
       SELECT m.REP_CODE,
              CASE WHEN m.BILL_DOC_TYPE IN (1,4) THEN 1 WHEN m.BILL_DOC_TYPE IN (2,5) THEN -1 ELSE 0 END as sign,
              ((NVL(d.I_QTY,0)*NVL(d.I_PRICE,0)) - NVL(d.DIS_AMT,0) - (CASE WHEN NVL(m.BILL_AMT,0)=0 THEN 0 ELSE ((NVL(d.I_QTY,0)*NVL(d.I_PRICE,0))/m.BILL_AMT)*NVL(m.DISC_AMT,0) END) + NVL(d.OTHR_AMT,0)) as rev,
              (NVL(d.I_QTY,0)*NVL(d.STK_COST,0)) as cst,
              0 as ext_disc
       FROM IAS20261.IAS_BILL_DTL d JOIN IAS20261.IAS_BILL_MST m ON m.BILL_SER=d.BILL_SER
       WHERE m.BILL_DOC_TYPE IN (1,4,2,5)
         AND m.BILL_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND m.BILL_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       UNION ALL
       SELECT REP_CODE, 1 as sign, 0 as rev, 0 as cst, NVL(CR_AMT,0) as ext_disc
       FROM IAS20261.IAS_POST_DTL
       WHERE DOC_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND DOC_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
         AND DOC_TYPE = 15 AND NVL(CR_AMT,0) > 0 AND NVL(DOC_POST,0) = 1
     )
     SELECT s.REP_CODE AS "كود المندوب", MAX(sm.REPRS_A_NAME) AS "اسم المندوب",
            TO_CHAR(SUM(rev * sign) - SUM(ext_disc),'FM999,999,999,990.00') AS "المبيعات",
            TO_CHAR(SUM(cst * sign),'FM999,999,999,990.00') AS "التكلفة",
            TO_CHAR((SUM(rev * sign) - SUM(ext_disc)) - SUM(cst * sign),'FM999,999,999,990.00') AS "الربح",
            TO_CHAR(ROUND(100*((SUM(rev * sign) - SUM(ext_disc)) - SUM(cst * sign))/NULLIF(SUM(rev * sign) - SUM(ext_disc),0),1),'FM990.0')||' %' AS "هامش"
     FROM s LEFT JOIN IAS20261.SALES_MAN sm ON sm.REPRS_CODE=s.REP_CODE
     WHERE s.REP_CODE IS NOT NULL
     GROUP BY s.REP_CODE ORDER BY (SUM(rev * sign) - SUM(ext_disc)) - SUM(cst * sign) DESC"""},
 ]},
 {"id":"stock","title":"المخزون","icon":"M3 7l9-4 9 4-9 4zM3 7v10l9 4 9-4V7M12 11v10","reports":[
   {"id":"stock_bal","title":"أرصدة الأصناف","params":[{"name":"as_of","label":"حتى تاريخ","type":"date","default":"2026-07-10"},{"name":"w_code","label":"المستودع (اختياري)","type":"text","default":""}],"sql":"""
     SELECT * FROM (
       SELECT mv.I_CODE AS "كود الصنف", MAX(i.I_NAME) AS "اسم الصنف",
              TO_CHAR(SUM(DECODE(NVL(mv.IN_OUT,0),1,NVL(mv.I_QTY,0),-NVL(mv.I_QTY,0))),'FM999,999,990.00') AS "الرصيد",
              TO_CHAR(SUM(DECODE(NVL(mv.IN_OUT,0),1,NVL(mv.I_QTY,0),-NVL(mv.I_QTY,0))*NVL(mv.STK_COST,0)),'FM999,999,999,990.00') AS "قيمة الرصيد (تقريبية)"
       FROM IAS20261.ITEM_MOVEMENT mv LEFT JOIN IAS20261.IAS_ITM_MST i ON i.I_CODE=mv.I_CODE
       WHERE mv.I_DATE < TO_DATE(:as_of,'YYYY-MM-DD')+1
         AND (:w_code IS NULL OR mv.W_CODE = :w_code)
       GROUP BY mv.I_CODE HAVING SUM(DECODE(NVL(mv.IN_OUT,0),1,NVL(mv.I_QTY,0),-NVL(mv.I_QTY,0))) <> 0
       ORDER BY SUM(DECODE(NVL(mv.IN_OUT,0),1,NVL(mv.I_QTY,0),-NVL(mv.I_QTY,0))*NVL(mv.STK_COST,0)) DESC
     ) WHERE ROWNUM<=300"""},
   {"id":"stock_move","title":"حركة صنف","params":[{"name":"i_code","label":"كود الصنف","type":"text","default":""},DFROM,DTO],"sql":"""
     SELECT * FROM (
       SELECT TO_CHAR(mv.I_DATE,'DD/MM/YYYY') AS "التاريخ", mv.DOC_NO AS "المستند",
              CASE NVL(mv.IN_OUT,0) WHEN 1 THEN 'وارد' ELSE 'صادر' END AS "الاتجاه",
              TO_CHAR(NVL(mv.I_QTY,0),'FM999,999,990.00') AS "الكمية",
              TO_CHAR(NVL(mv.STK_COST,0),'FM999,999,990.00') AS "التكلفة",
              mv.W_CODE AS "المستودع"
       FROM IAS20261.ITEM_MOVEMENT mv
       WHERE mv.I_CODE = :i_code
         AND mv.I_DATE >= TO_DATE(:date_from,'YYYY-MM-DD') AND mv.I_DATE < TO_DATE(:date_to,'YYYY-MM-DD')+1
       ORDER BY mv.I_DATE, mv.DOC_NO
     ) WHERE ROWNUM<=500"""},
   {"id":"stock_dormant","title":"الأصناف الراكدة (لم تُبَع)","params":[{"name":"as_of","label":"حتى تاريخ","type":"date","default":"2026-07-10"},{"name":"days","label":"أيام الركود","type":"number","default":"90"}],"sql":"""
     SELECT * FROM (
       SELECT mv.I_CODE AS "كود الصنف", MAX(i.I_NAME) AS "اسم الصنف",
              TO_CHAR(SUM(DECODE(NVL(mv.IN_OUT,0),1,NVL(mv.I_QTY,0),-NVL(mv.I_QTY,0))),'FM999,999,990.00') AS "الرصيد",
              TO_CHAR(MAX(CASE WHEN NVL(mv.IN_OUT,0)<>1 THEN mv.I_DATE END),'DD/MM/YYYY') AS "آخر صرف",
              TRUNC(TO_DATE(:as_of,'YYYY-MM-DD')) - TRUNC(MAX(CASE WHEN NVL(mv.IN_OUT,0)<>1 THEN mv.I_DATE END)) AS "أيام منذ آخر صرف"
       FROM IAS20261.ITEM_MOVEMENT mv LEFT JOIN IAS20261.IAS_ITM_MST i ON i.I_CODE=mv.I_CODE
       WHERE mv.I_DATE < TO_DATE(:as_of,'YYYY-MM-DD')+1
       GROUP BY mv.I_CODE
       HAVING SUM(DECODE(NVL(mv.IN_OUT,0),1,NVL(mv.I_QTY,0),-NVL(mv.I_QTY,0))) > 0
          AND ( MAX(CASE WHEN NVL(mv.IN_OUT,0)<>1 THEN mv.I_DATE END) IS NULL
                OR TRUNC(TO_DATE(:as_of,'YYYY-MM-DD')) - TRUNC(MAX(CASE WHEN NVL(mv.IN_OUT,0)<>1 THEN mv.I_DATE END)) >= :days )
       ORDER BY SUM(DECODE(NVL(mv.IN_OUT,0),1,NVL(mv.I_QTY,0),-NVL(mv.I_QTY,0))) DESC
     ) WHERE ROWNUM<=300"""},
 ]},
    ,{
        "id": "global_vars",
        "title": "المتغيرات العامة",
        "icon": '<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M10.325 4.317c.426-1.756 2.924-1.756 3.35 0a1.724 1.724 0 0 0 2.573 1.066c1.543-.94 3.31.826 2.37 2.37a1.724 1.724 0 0 0 1.065 2.572c1.756.426 1.756 2.924 0 3.35a1.724 1.724 0 0 0 -1.066 2.573c.94 1.543-.826 3.31-2.37 2.37a1.724 1.724 0 0 0 -2.572 1.065c-.426 1.756-2.924 1.756-3.35 0a1.724 1.724 0 0 0 -2.573 -1.066c-1.543.94-3.31-.826-2.37-2.37a1.724 1.724 0 0 0 -1.065 -2.572c-1.756-.426-1.756-2.924 0-3.35a1.724 1.724 0 0 0 1.066 -2.573c-.94-1.543.826-3.31 2.37-2.37.996.15 2.015-.267 2.572-1.065z" /><path d="M9 12a3 3 0 1 0 6 0a3 3 0 0 0 -6 0" />',
        "reports": [
            {
                "id": "targets_ui",
                "title": "تارقت المناديب",
                "sql": "SELECT 1 FROM DUAL",
                "params": [],
                "cols": ["Dummy"]
            }
        ]
    }
]

TABMAP = {t["id"]: t for t in TABS}
def find_report(tab, rid):
    t = TABMAP.get(tab) or TABS[0]
    for r in t["reports"]:
        if r["id"] == rid:
            return t, r
    return t, t["reports"][0]
    }
]
def run_report(rpt, args):
    binds = {}
    for p in rpt["params"]:
        v = args.get(p["name"], p.get("default",""))
        if p["type"] == "number":
            try: v = int(v)
            except: v = 0
            binds[p["name"]] = v
        else:
            if p["name"] in ("rep_code","c_code","v_code","i_code","a_code") and v:
                v = v.split(" - ")[0].strip()
            binds[p["name"]] = v if v != "" else None
    with get_conn() as con:
        with con.cursor() as cur:
            cur.execute(rpt["sql"], binds)
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
    return cols, rows

_JV_CACHE = None
def jv_options():
    global _JV_CACHE
    if _JV_CACHE is not None:
        return _JV_CACHE
    opts = [["","الكل"]]
    try:
        with get_conn() as con:
            with con.cursor() as cur:
                cur.execute("""SELECT p.JV_TYPE, MAX(j.JV_NAME) nm
                    FROM IAS20261.IAS_POST_DTL p
                    LEFT JOIN IAS20261.JV_TYPES j ON j.JV_TYPE=p.JV_TYPE
                    WHERE p.C_CODE IS NOT NULL AND NVL(p.DOC_POST,0)=1
                    GROUP BY p.JV_TYPE ORDER BY COUNT(*) DESC""")
                for t, nm in cur.fetchall():
                    if t is None: continue
                    code = str(int(t))
                    label = nm or ("بدون نوع" if int(t)==0 else "نوع "+code)
                    opts.append([code, label])
        _JV_CACHE = opts
        return _JV_CACHE
    except Exception:
        return [["","الكل"],["1","قيد يومية"],["2","قيود الشبكة"],["4","قيود دائنون"],["9","قيد ضريبي"],["11","رصيد افتتاحي للعملاء"],["0","بدون نوع"]]

_LK_CACHE = {}
def lookups(name):
    if name in _LK_CACHE:
        return _LK_CACHE[name]
    q = {
      "rep_code": "SELECT REPRS_CODE, REPRS_A_NAME FROM IAS20261.SALES_MAN WHERE REPRS_CODE IS NOT NULL ORDER BY REPRS_A_NAME",
      "c_code":   "SELECT C_CODE, C_A_NAME FROM IAS20261.CUSTOMER WHERE NVL(INACTIVE,0)=0 AND C_CODE IS NOT NULL ORDER BY C_A_NAME",
      "v_code":   "SELECT V_CODE, MAX(V_NAME) FROM IAS20261.IAS_PI_BILL_MST WHERE V_CODE IS NOT NULL GROUP BY V_CODE ORDER BY MAX(V_NAME)",
      "i_code":   "SELECT I_CODE, I_NAME FROM IAS20261.IAS_ITM_MST WHERE I_CODE IS NOT NULL ORDER BY I_NAME",
      "a_code":   "SELECT A_CODE, A_NAME FROM IAS20261.ACCOUNT WHERE A_CODE IS NOT NULL ORDER BY A_CODE",
    }.get(name)
    if not q:
        return []
    out = []
    try:
        with get_conn() as con:
            with con.cursor() as cur:
                cur.execute(q)
                for code, nm in cur.fetchall():
                    if code is None:
                        continue
                    out.append(("%s - %s" % (str(code).strip(), (nm or "").strip())).strip(" -"))
        _LK_CACHE[name] = out
        return out
    except Exception:
        return []

STYLE = """<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;500;600;700;800&family=Inter:wght@400;500;600;700;800&display=swap');
:root {
  --bg: #f4f5f8;
  --sb-bg: #ffffff;
  --card-bg: #ffffff;
  --primary: #4f46e5;
  --primary-hover: #4338ca;
  --ink: #64748b;
  --ink-dark: #1e293b;
  --line: #f1f5f9;
  --sh: 0 10px 40px rgba(0,0,0,0.04);
}
body { background: var(--bg); color: var(--ink); font-family: 'Cairo', 'Inter', sans-serif; direction: rtl; margin:0; padding:0; box-sizing:border-box; }
* { box-sizing: border-box; margin:0; padding:0; }
a { text-decoration: none; }
.card { background: var(--card-bg); border-radius: 20px; padding: 24px; box-shadow: var(--sh); }
.app { display: flex; min-height: 100vh; padding: 20px; gap: 24px; }
.sb { width: 260px; background: var(--sb-bg); border-radius: 24px; display: flex; flex-direction: column; padding: 30px 20px; flex-shrink: 0; box-shadow: var(--sh); }
.brand { display: flex; align-items: center; gap: 12px; font-size: 24px; font-weight: 800; color: var(--ink-dark); margin-bottom: 40px; }
.brand svg { width: 32px; height: 32px; fill: var(--primary); }
.menu-lbl { font-size: 11px; font-weight: 700; color: #94a3b8; margin: 20px 10px 10px; letter-spacing: 1px; }
.sb a { display: flex; align-items: center; gap: 14px; padding: 14px 20px; border-radius: 16px; color: var(--ink); font-weight: 600; font-size: 15px; margin-bottom: 8px; transition: all 0.3s; }
.sb a:hover { background: #f8fafc; color: var(--ink-dark); }
.sb a.on { background: var(--primary); color: #fff; box-shadow: 0 10px 20px rgba(79, 70, 229, 0.25); }
.sb svg { width: 22px; height: 22px; stroke: currentColor; fill: none; stroke-width: 2; }
.sb a.on svg { stroke: #fff; }

.main { flex: 1; min-width: 0; display: flex; flex-direction: column; }
.top { display: flex; align-items: center; padding: 10px 0 30px; gap: 15px; }
.logo { display:none; }
.ttl { font-size: 26px; font-weight: 800; color: var(--ink-dark); }
.wrap { display: flex; flex-direction: column; gap: 24px; }

.pills { display: flex; gap: 12px; flex-wrap: wrap; }
.pill { background: var(--card-bg); border-radius: 12px; padding: 12px 24px; font-size: 14px; font-weight: 600; color: var(--ink); box-shadow: var(--sh); transition: 0.3s; }
.pill:hover { transform: translateY(-2px); color: var(--primary); }
.pill.on { background: var(--primary); color: #fff; box-shadow: 0 10px 20px rgba(79, 70, 229, 0.25); }

.filters { background: var(--card-bg); border-radius: 20px; padding: 24px; display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 20px; align-items: end; box-shadow: var(--sh); margin-bottom: 24px; }
.filters label { display: block; font-size: 13px; font-weight: 600; color: var(--ink); margin-bottom: 8px; }
.filters input, .filters select { width: 100%; padding: 12px 16px; border: 1px solid var(--line); border-radius: 12px; font-family: inherit; font-size: 14px; font-weight: 500; color: var(--ink-dark); background: #f8fafc; outline: none; transition: 0.3s; }
.filters input:focus, .filters select:focus { border-color: var(--primary); background: #fff; box-shadow: 0 0 0 4px rgba(79, 70, 229, 0.1); }
.filters .btn { background: var(--primary); color: #fff; border: 0; padding: 14px 24px; border-radius: 12px; font-weight: 600; font-size: 14px; cursor: pointer; transition: 0.3s; height: 46px; }
.filters .btn:hover { background: var(--primary-hover); transform: translateY(-2px); box-shadow: 0 10px 20px rgba(79, 70, 229, 0.2); }

.tw { overflow-x: auto; background: var(--card-bg); border-radius: 20px; box-shadow: var(--sh); padding: 10px; }
table { border-collapse: collapse; width: 100%;  }
thead th { white-space: nowrap; color: var(--ink); padding: 8px 12px; text-align: right; font-size: 13px; font-weight: 600; border-bottom: 2px solid var(--line);  }
tbody td { white-space: nowrap; padding: 6px 12px; border-bottom: 1px solid var(--line); font-size: 13px; font-weight: 500; color: var(--ink-dark);  transition: 0.2s; }
tbody tr:hover td { background: #f8fafc; }

.rhead { display: flex; align-items: center; gap: 16px; margin-bottom: 10px; }
.rhead h1 { margin: 0; flex: 1; font-size: 20px; color: var(--ink-dark); font-weight: 800; border:0; padding:0; }
.rhead h1::before { display: none; }
.cnt { color: var(--ink); font-size: 13px; font-weight: 600; margin-bottom: 10px; }
.exps { display: flex; gap: 10px; }
.exp { border: 0; border-radius: 10px; padding: 10px 20px; font-weight: 600; font-size: 13px; color: #fff; cursor: pointer; transition: 0.3s; }
.exp:hover { transform: translateY(-2px); box-shadow: 0 5px 15px rgba(0,0,0,0.1); }
.exp.xl { background: #10b981; } .exp.pf { background: #ef4444; }
.err { background: #fef2f2; color: #b91c1c; padding: 16px; border-radius: 12px; font-weight: 600; }

.gdwrap { display: flex; flex-direction: column; gap: 24px; }
.gkpis { display: grid; grid-template-columns: repeat(4, 1fr); gap: 24px; }
.gk { background: var(--card-bg); border-radius: 24px; padding: 24px; display: flex; flex-direction: column; gap: 16px; box-shadow: var(--sh); position: relative; overflow: hidden; }
.gk:nth-child(1) { background: var(--primary); color: #fff; }
.gk:nth-child(1) .gl { color: rgba(255,255,255,0.8); }
.gk:nth-child(1) .gv { color: #fff; }
.gk .gic { width: 48px; height: 48px; border-radius: 16px; display: flex; align-items: center; justify-content: center; font-size: 20px; }
.gk:nth-child(1) .gic { background: rgba(255,255,255,0.2); }
.gk:nth-child(2) .gic { background: #dcfce7; color: #16a34a; }
.gk:nth-child(3) .gic { background: #ffedd5; color: #f97316; }
.gk:nth-child(4) .gic { background: #e0e7ff; color: #4f46e5; }
.gk:nth-child(5) .gic { background: #d1fae5; color: #059669; }
.gk:nth-child(6) .gic { background: #fee2e2; color: #dc2626; }
.gk:nth-child(7) .gic { background: #e0f2fe; color: #0284c7; }
.gk:nth-child(8) .gic { background: #fef3c7; color: #d97706; }
.gk .gl { font-size: 13px; font-weight: 600; color: var(--ink); margin-bottom: 4px; }
.gk .gv { font-size: 26px; font-weight: 800; color: var(--ink-dark); }
.gcharts { display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }
.gc { background: var(--card-bg); border-radius: 24px; padding: 24px; box-shadow: var(--sh); }
.gc h3 { font-size: 16px; font-weight: 700; margin: 0 0 20px; color: var(--ink-dark); }

@media(max-width:900px){.app{flex-direction:column;padding:10px;} .sb{width:100%;flex-direction:row;padding:15px;overflow-x:auto;border-radius:16px;gap:10px; align-items:center;} .brand{margin:0;padding-right:15px;} .brand span{display:none;} .menu-lbl{display:none;} .sb a{margin:0;padding:10px;} .sb a span{display:none;} .gkpis{grid-template-columns:repeat(2,1fr)}.gcharts{grid-template-columns:1fr}}
</style>"""

LOGO = '<div style="color:#4f46e5;font-weight:900;font-size:26px;letter-spacing:-1px">تقارير الأونكس الحديثة</div>'

PAGE = """<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>تقارير SREEN</title>""" + STYLE + """</head><body>
<div class="app">
 <aside class="sb">
   <div class="brand"><svg viewBox="0 0 24 24"><rect x="3" y="3" width="7" height="7" rx="1"/><rect x="14" y="3" width="7" height="7" rx="1"/><rect x="3" y="14" width="7" height="7" rx="1"/><rect x="14" y="14" width="7" height="7" rx="1"/></svg><span>Onyx Deck</span></div>
   <div class="menu-lbl">القائمة الرئيسية</div>
   
   {% for t in tabs %}{% if t.id not in hidden_tabs %}
     <a class="{{ 'on' if t.id==cur_tab else '' }}" href="/?tab={{t.id}}">
       <svg viewBox="0 0 24 24"><path d="{{t.icon}}"/></svg><span>{{ t.title }}</span></a>
   {% endif %}{% endfor %}
   <div class="menu-lbl" style="margin-top:auto">أدوات</div>
   <a href="/settings"><svg viewBox="0 0 24 24"><path d="M4 6h9M4 12h5M4 18h7"/><circle cx="17" cy="6" r="2.3"/><circle cx="13" cy="12" r="2.3"/><circle cx="15" cy="18" r="2.3"/></svg><span>الإعدادات</span></a>
 </aside>
 <div class="main">
   <div class="top">""" + LOGO + """<div class="ttl">لوحة <b>التقارير</b></div></div>
   <div class="wrap">
     {% if dash %}
     <div class="rhead"><h1>لوحة القيادة</h1></div>
     <form class="filters" method="get" action="/">
       <input type="hidden" name="tab" value="{{cur_tab}}"><input type="hidden" name="report" value="overview">
       <div><label>من تاريخ</label><input type="date" name="date_from" value="{{ binds.get('date_from') or '2026-01-01' }}"></div>
       <div><label>إلى تاريخ</label><input type="date" name="date_to" value="{{ binds.get('date_to') or '2026-12-31' }}"></div>
       <div><button class="btn" type="submit">تحديث</button></div>
     </form>
     {% if error %}<div class="err">خطأ: {{error}}</div>{% else %}
     <div class="gdwrap">
       <div class="gkpis">
         <div class="gk"><div class="gic" style="background:#dbeafe">💵</div><div><div class="gl">إجمالي المبيعات</div><div class="gv">{{ "{:,.0f}".format(dash.sales) }}</div></div></div>
         <div class="gk"><div class="gic" style="background:#dcfce7">💰</div><div><div class="gl">إجمالي التحصيل</div><div class="gv">{{ "{:,.0f}".format(dash.collect) }}</div></div></div>
         <div class="gk"><div class="gic" style="background:#ffedd5">🛒</div><div><div class="gl">إجمالي المشتريات</div><div class="gv">{{ "{:,.0f}".format(dash.purch) }}</div></div></div>
         {% if not hide_profit %}<div class="gk"><div class="gic" style="background:#ede9fe">📈</div><div><div class="gl">مجمل الربح</div><div class="gv">{{ "{:,.0f}".format(dash.gross) }}</div></div></div>
         <div class="gk"><div class="gic" style="background:#d1fae5">✅</div><div><div class="gl">صافي الربح</div><div class="gv">{{ "{:,.0f}".format(dash.netprofit) }}</div></div></div>{% endif %}
         <div class="gk"><div class="gic" style="background:#fee2e2">🧾</div><div><div class="gl">الذمم المدينة</div><div class="gv">{{ "{:,.0f}".format(dash.recv) }}</div></div></div>
         <div class="gk"><div class="gic" style="background:#e0f2fe">📦</div><div><div class="gl">قيمة المخزون</div><div class="gv">{{ "{:,.0f}".format(dash.invval) }}</div></div></div>
         <div class="gk"><div class="gic" style="background:#fef3c7">🏛️</div><div><div class="gl">صافي الضريبة</div><div class="gv">{{ "{:,.0f}".format(dash.vat) }}</div></div></div>
       </div>
       <div class="gcharts" style="grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));">
          <div class="gc" style="grid-column: 1 / -1;"><h3>المبيعات والتحصيل شهرياً</h3><div style="position:relative;height:280px;width:100%"><canvas id="c1"></canvas></div></div>
          <div class="gc"><h3>أفضل 5 مناديب</h3><div style="position:relative;height:250px;width:100%"><canvas id="c2"></canvas></div></div>
          <div class="gc"><h3>أفضل 5 أصناف</h3><div style="position:relative;height:250px;width:100%"><canvas id="c3"></canvas></div></div>
          <div class="gc" style="grid-column: 1 / -1;"><h3>المشتريات شهرياً</h3><div style="position:relative;height:280px;width:100%"><canvas id="c4"></canvas></div></div>
        </div>
     </div>
          <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
     <script>
     var D={{ dash|tojson }};
     window.addEventListener("load",function(){ 
       if(!window.Chart) return; 
       Chart.defaults.font.family = "'Cairo', 'Inter', sans-serif";
       Chart.defaults.color = "#64748b";
       
       const commonOptions = {
         responsive: true,
         maintainAspectRatio: false,
         plugins: {
           legend: { display: false },
           tooltip: { backgroundColor: '#1e293b', padding: 14, titleFont: { size: 14, family: "'Cairo', sans-serif", weight: 'bold' }, bodyFont: { size: 14, family: "'Cairo', sans-serif" }, cornerRadius: 10, displayColors: true, boxPadding: 6 }
         }
       };
       
       // C1: Bar Chart (Sales & Collection)
       new Chart(document.getElementById("c1"),{
         type:"bar",
         data:{
           labels:D.months,
           datasets:[
             {label:"مبيعات", data:D.msales, backgroundColor:"#4f46e5", borderRadius:8, maxBarThickness: 32},
             {label:"تحصيل", data:D.mcollect, backgroundColor:"#38bdf8", borderRadius:8, maxBarThickness: 32}
           ]
         },
         options: {
           ...commonOptions,
           plugins: { ...commonOptions.plugins, legend: { display: true, position: 'top', align: 'end', labels: { usePointStyle: true, boxWidth: 10, font: { family: "'Cairo'", size: 13, weight: 'bold' } } } },
           scales: {
             x: { grid: { display: false }, border: { display: false } },
             y: { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false } }
           }
         }
       });

       // C2: Doughnut (Salesmen)
       new Chart(document.getElementById("c2"),{
         type:"doughnut",
         data:{
           labels:D.rep_labels.slice(0,5),
           datasets:[{data:D.rep_vals.slice(0,5), backgroundColor:["#4f46e5", "#38bdf8", "#10b981", "#f59e0b", "#8b5cf6"], borderWidth: 0, hoverOffset: 4}]
         },
         options: {
           responsive: true, maintainAspectRatio: false, cutout: '75%',
           plugins: { legend: { display: true, position: 'bottom', labels: { usePointStyle: true, boxWidth: 8, font: { size: 11 } } } }
         }
       });

       // C3: Doughnut (Items)
       new Chart(document.getElementById("c3"),{
         type:"doughnut",
         data:{
           labels:D.itm_labels.slice(0,5),
           datasets:[{data:D.itm_vals.slice(0,5), backgroundColor:["#f43f5e", "#d946ef", "#0ea5e9", "#14b8a6", "#eab308"], borderWidth: 0, hoverOffset: 4}]
         },
         options: {
           responsive: true, maintainAspectRatio: false, cutout: '75%',
           plugins: { legend: { display: true, position: 'bottom', labels: { usePointStyle: true, boxWidth: 8, font: { size: 11 } } } }
         }
       });

       // C4: Line Chart (Purchases)
       const ctx4 = document.getElementById("c4").getContext('2d');
       const grad4 = ctx4.createLinearGradient(0, 0, 0, 300);
       grad4.addColorStop(0, 'rgba(16, 185, 129, 0.4)');
       grad4.addColorStop(1, 'rgba(16, 185, 129, 0.0)');
       
       new Chart(ctx4,{
         type:"line",
         data:{
           labels:D.months,
           datasets:[{
             label: "مشتريات", data:D.mpurch, borderColor:"#10b981", borderWidth: 3, backgroundColor: grad4, fill:true, tension:0.4, pointRadius: 0, pointHoverRadius: 6, pointBackgroundColor: "#fff", pointBorderColor: "#10b981", pointBorderWidth: 2
           }]
         },
         options: {
           ...commonOptions,
           interaction: { mode: 'index', intersect: false },
           scales: {
             x: { grid: { display: false }, border: { display: false } },
             y: { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false } }
           }
         }
       });
     });
     </script>
     <script>
     var D={{ dash|tojson }};
     window.addEventListener("load",function(){ 
       if(!window.Chart) return; 
       Chart.defaults.font.family = "'Cairo', 'Inter', sans-serif";
       Chart.defaults.color = "#64748b";
       
       const commonOptions = {
         responsive: true,
         maintainAspectRatio: false,
         plugins: {
           legend: { display: false },
           tooltip: {
             backgroundColor: '#1e293b',
             padding: 14,
             titleFont: { size: 14, family: "'Cairo', sans-serif", weight: 'bold' },
             bodyFont: { size: 14, family: "'Cairo', sans-serif" },
             cornerRadius: 10,
             displayColors: true,
             boxPadding: 6
           }
         },
         scales: {
           x: { grid: { display: false }, border: { display: false }, ticks: { font: { weight: '600' } } },
           y: { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false }, ticks: { font: { weight: '600' }, padding: 10 } }
         }
       };

       const horizontalOptions = JSON.parse(JSON.stringify(commonOptions));
       horizontalOptions.indexAxis = "y";
       horizontalOptions.scales.x = { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false }, ticks: { font: { weight: '600' } } };
       horizontalOptions.scales.y = { grid: { display: false }, border: { display: false }, ticks: { font: { weight: '600' }, padding: 10 } };
       
       // C1: Bar Chart (Sales & Collection)
       new Chart(document.getElementById("c1"),{
         type:"bar",
         data:{
           labels:D.months,
           datasets:[
             {label:"مبيعات", data:D.msales, backgroundColor:"#4f46e5", borderRadius:8, maxBarThickness: 32, borderSkipped: false},
             {label:"تحصيل", data:D.mcollect, backgroundColor:"#38bdf8", borderRadius:8, maxBarThickness: 32, borderSkipped: false}
           ]
         },
         options: {
           ...commonOptions,
           plugins: {
             ...commonOptions.plugins,
             legend: { display: true, position: 'top', align: 'end', labels: { usePointStyle: true, boxWidth: 10, padding: 20, font: { family: "'Cairo'", size: 13, weight: 'bold' } } }
           }
         }
       });

       // C2: Horizontal Bar (Salesmen)
       new Chart(document.getElementById("c2"),{
         type:"bar",
         data:{
           labels:D.rep_labels,
           datasets:[{label: "مبيعات", data:D.rep_vals, backgroundColor:"#8b5cf6", borderRadius:8, maxBarThickness: 24, borderSkipped: false}]
         },
         options: horizontalOptions
       });

       // C3: Horizontal Bar (Items)
       new Chart(document.getElementById("c3"),{
         type:"bar",
         data:{
           labels:D.itm_labels,
           datasets:[{label: "مبيعات", data:D.itm_vals, backgroundColor:"#10b981", borderRadius:8, maxBarThickness: 24, borderSkipped: false}]
         },
         options: horizontalOptions
       });

       // C4: Line Chart (Purchases)
       const ctx4 = document.getElementById("c4").getContext('2d');
       const grad4 = ctx4.createLinearGradient(0, 0, 0, 300);
       grad4.addColorStop(0, 'rgba(249, 115, 22, 0.4)');
       grad4.addColorStop(1, 'rgba(249, 115, 22, 0.0)');
       
       new Chart(ctx4,{
         type:"line",
         data:{
           labels:D.months,
           datasets:[{
             label: "مشتريات",
             data:D.mpurch,
             borderColor:"#f97316",
             borderWidth: 3,
             backgroundColor: grad4,
             fill:true,
             tension:0.4,
             pointRadius: 0,
             pointHoverRadius: 6,
             pointBackgroundColor: "#fff",
             pointBorderColor: "#f97316",
             pointBorderWidth: 2
           }]
         },
         options: {
           ...commonOptions,
           interaction: { mode: 'index', intersect: false }
         }
       });
     });
     </script>
     {% endif %}
     {% else %}
     <div class="pills">
       {% for r in tab.reports %}{% if (cur_tab ~ '/' ~ r.id) not in hidden_reports %}
         <a class="pill {{ 'on' if r.id==rpt.id else '' }}" href="/?tab={{cur_tab}}&report={{r.id}}">{{ r.title }}</a>
       {% endif %}{% endfor %}
     </div>
     <div class="rhead">
  <h1>{{ rpt.title }}</h1>
  <div class="exps">
    <a class="exp xl" href="/export?{{qs}}&format=xlsx">Excel</a>
    {% if rpt.id == 'collection_adopted' %}
      <select id="pdfModel" style="padding:4px 8px; border:1px solid #cbd5e1; border-radius:4px; margin-left:4px; font-family:inherit; font-size:13px;">
        <option value="1">PDF (النموذج الافتراضي)</option>
        <option value="2">PDF (نموذج 2)</option>
      </select>
      <button class="exp pf" style="border:none; cursor:pointer;" onclick="window.open('/print?{{qs|safe}}&model=' + document.getElementById('pdfModel').value, '_blank')">طباعة</button>
    {% else %}
      <a class="exp pf" href="/print?{{qs}}" target="_blank">PDF</a>
    {% endif %}
  </div>
</div>
     {% if rpt.params %}
     <form class="filters" method="get" action="/">
       <input type="hidden" name="tab" value="{{cur_tab}}"><input type="hidden" name="report" value="{{rpt.id}}">
       {% for p in rpt.params %}
         <div><label>{{p.label}}</label>
         {% if p.type=='select' %}
           <select name="{{p.name}}">{% for o in p.options %}<option value="{{o[0]}}" {{'selected' if binds.get(p.name)==o[0] else ''}}>{{o[1]}}</option>{% endfor %}</select>
         {% elif p.get('_list') %}
           <input type="text" name="{{p.name}}" list="dl_{{p.name}}" autocomplete="off" placeholder="ابحث بالكود أو الاسم" value="{{ binds.get(p.name) if binds.get(p.name) is not none else '' }}">
           <datalist id="dl_{{p.name}}">{% for o in p.get('_list') %}<option value="{{o}}"></option>{% endfor %}</datalist>
         {% else %}
           <input type="{{p.type}}" name="{{p.name}}" value="{{ binds.get(p.name) if binds.get(p.name) is not none else '' }}">
         {% endif %}
         </div>
       {% endfor %}
       <div><button class="btn" type="submit">عرض التقرير</button></div>
     </form>
     {% endif %}
     {% if error %}<div class="err">خطأ: {{error}}</div>
     {% else %}
       <div class="cnt">عدد الصفوف: {{rows|length}}</div>
       <div class="tw"><table><thead><tr>{% for c in cols %}<th>{{c}}</th>{% endfor %}</tr></thead>
       <tbody>{% for row in rows %}<tr>{% for cell in row %}<td>{{ '' if cell is none else cell }}</td>{% endfor %}</tr>{% endfor %}</tbody></table></div>
     {% endif %}
     {% endif %}
   </div>
 </div>
</div>
    <script>
      document.addEventListener("DOMContentLoaded", function() {
        const typeSelect = document.querySelector('select[name="p_type"]');
        const valSelect = document.querySelector('select[name="p_val"]');
        if(typeSelect && valSelect) {
          const valWrapper = valSelect.parentElement; // Usually a div grouping label + select
          
          function updateOptions() {
            const val = typeSelect.value;
            valSelect.innerHTML = ''; // clear options
            valWrapper.style.display = 'block';
            
            let options = [];
            if(val === 'month') {
              for(let i=1; i<=12; i++) options.push([i, "شهر " + i]);
            } else if(val === 'quarter') {
              options = [[1, 'الربع الأول'], [2, 'الربع الثاني'], [3, 'الربع الثالث'], [4, 'الربع الرابع']];
            } else if(val === 'half') {
              options = [[1, 'النصف الأول'], [2, 'النصف الثاني']];
            } else if(val === 'year') {
              valWrapper.style.display = 'none';
            }
            
            options.forEach(opt => {
              const el = document.createElement('option');
              el.value = opt[0];
              el.textContent = opt[1];
              valSelect.appendChild(el);
            });
            
            // try to re-select previous value if any from URL
            const urlParams = new URLSearchParams(window.location.search);
            const prevVal = urlParams.get('p_val');
            if(prevVal) {
                valSelect.value = prevVal;
                if(!valSelect.value) valSelect.value = options.length > 0 ? options[0][0] : '';
            }
          }
          
          typeSelect.addEventListener('change', updateOptions);
          updateOptions(); // call on load
        }
      });
    </script>
</body></html>"""

@app.route("/")
def index():
    hidden_tabs, hidden_reports = load_hidden()
    _vis = [t for t in TABS if t["id"] not in hidden_tabs] or TABS
    cur_tab = request.args.get("tab", _vis[0]["id"])
    rid = request.args.get("report", "")
    tab, rpt = find_report(cur_tab, rid)
    cur_tab = tab["id"]
    for _p in rpt["params"]:
        if _p.get("dynamic") == "jv": _p["options"] = jv_options()
        if _p["name"] in ("rep_code","c_code","v_code","i_code","a_code"): _p["_list"] = lookups(_p["name"])
    display = {p["name"]: request.args.get(p["name"], p.get("default","")) for p in rpt["params"]}
    qsp = {"tab": cur_tab, "report": rpt["id"]}
    for p in rpt["params"]: qsp[p["name"]] = request.args.get(p["name"], p.get("default",""))
    qs = urlencode(qsp)
    error = None; cols=[]; rows=[]; dash=None
    if tab.get("dash"):
        try:
            dash = compute_dash(request.args.get("date_from","2026-01-01"), request.args.get("date_to","2026-12-31"))
        except Exception as e:
            error = str(e)
    else:
        try:
            cols, rows = run_report(rpt, request.args)
        except Exception as e:
            error = str(e)
    return render_template_string(PAGE, tabs=TABS, tab=tab, cur_tab=cur_tab, rpt=rpt,
                                  binds=display, cols=cols, rows=rows, error=error, qs=qs, dash=dash, hidden_tabs=hidden_tabs, hidden_reports=hidden_reports, hide_profit=load_hide_profit())

PRINT_PAGE = """<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<style>
@page { size: A4 landscape; margin: 10mm; }
*{-webkit-print-color-adjust:exact;print-color-adjust:exact}
body{font-family:'Cairo','Inter',Tahoma,sans-serif;direction:rtl;color:#1e293b;margin:0;font-size:11px;}
.hd{display:flex;align-items:center;justify-content:space-between;border-bottom:2px solid #4f46e5;padding-bottom:10px;margin-bottom:15px}
.hd h1{font-size:18px;margin:0;color:#0f172a;font-weight:800}
.hd .dt{font-size:11px;color:#64748b;margin-top:6px;font-weight:600}
.logo{height:35px}
.filt{font-size:11px;color:#475569;margin-bottom:10px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:8px 10px;font-weight:600}
.filt b{color:#4f46e5}
table{border-collapse:collapse;width:100%;table-layout:fixed;word-wrap:break-word;}
thead th{background:#4f46e5;color:#fff;padding:4px 4px;font-size:11px;text-align:right;border:1px solid #4338ca;font-weight:700}
tbody td{padding:3px 4px;font-size:11px;border:1px solid #e2e8f0;text-align:right;font-weight:500;color:#1e293b}
tbody tr:nth-child(even) td{background:#f8fafc}
.ft{margin-top:20px;font-size:11px;color:#94a3b8;text-align:center;border-top:1px solid #e2e8f0;padding-top:10px;font-weight:600}
</style></head>
<body onload="setTimeout(function(){window.print()},250)">
<div class="hd">
  <h1 style="color:#4f46e5;font-weight:900;margin:0;font-size:26px">تقارير الأونكس الحديثة</h1>
  <div><h1>{{title}}</h1><div class="dt">تاريخ الطباعة: {{now}}</div></div>
</div>
{% if filt %}<div class="filt">الفلاتر — {% for f in filt %}<b>{{f[0]}}</b>: {{f[1]}}{% if not loop.last %} &nbsp;|&nbsp; {% endif %}{% endfor %}</div>{% endif %}
<table><thead><tr>{% for c in cols %}<th>{{c}}</th>{% endfor %}</tr></thead>
<tbody>{% for row in rows %}<tr>{% for cell in row %}<td>{{ '' if cell is none else cell }}</td>{% endfor %}</tr>{% endfor %}</tbody></table>
<div class="ft">لوحة تقارير SREEN — عدد الصفوف: {{rows|length}}</div>
</body></html>"""

@app.route("/export")
def export():
    tab, rpt = find_report(request.args.get("tab", TABS[0]["id"]), request.args.get("report",""))
    try:
        cols, rows = run_report(rpt, request.args)
    except Exception as e:
        return "خطأ: " + str(e), 500
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws = wb.active; ws.title = "تقرير"
        ws.sheet_view.rightToLeft = True
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
            cell.alignment = Alignment(horizontal="right")
        for r in rows:
            ws.append([str(x) if x is not None else '' for x in r])
        for i in range(1, len(cols)+1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=%s.xlsx" % rpt["id"]})
    except ImportError:
        import csv
        buf = io.StringIO()
        buf.write('\ufeff')
        writer = csv.writer(buf)
        writer.writerow(cols)
        for r in rows:
            writer.writerow([str(x) if x is not None else '' for x in r])
        return Response(buf.getvalue().encode('utf-8'),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=%s.csv" % rpt["id"]})

@app.route("/print")
def printview():
    tab, rpt = find_report(request.args.get("tab", TABS[0]["id"]), request.args.get("report",""))
    try:
        cols, rows = run_report(rpt, request.args)
        model = request.args.get("model", "1")
        if model == "2" and rpt["id"] == "collection_adopted":
            new_cols = ["الرمز", "الاسم / الوصف", "إجمالي السندات", "قيود الشبكة المنفصلة", "صافي المبيعات النقدية", "الإجمالي النهائي"]
            new_rows = []
            for r in rows:
                def parse_num(v):
                    if not v: return 0.0
                    if isinstance(v, str):
                        try: return float(v.replace(',',''))
                        except: return 0.0
                    return float(v)
                
                tot_rcpt = parse_num(r[2]) + parse_num(r[3]) + parse_num(r[4])
                net_jrn = parse_num(r[6])
                net_cash = parse_num(r[7]) - parse_num(r[10])
                final_tot = tot_rcpt + net_jrn + net_cash
                
                fmt = lambda x: f"{x:,.2f}" if x != 0 else "0.00"
                new_rows.append((r[0], r[1], fmt(tot_rcpt), fmt(net_jrn), fmt(net_cash), fmt(final_tot)))
            
            cols = new_cols
            rows = new_rows

    except Exception as e:
        return "خطأ: " + str(e), 500
    filt = []
    for p in rpt["params"]:
        v = request.args.get(p["name"], p.get("default",""))
        if v not in ("", None): filt.append((p["label"], v))
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    title = rpt["title"] + (" (نموذج 2)" if request.args.get("model") == "2" else "")
    return render_template_string(PRINT_PAGE, title=title, cols=cols, rows=rows, filt=filt, now=now)

SETTINGS_PAGE = """<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>الإعدادات</title>""" + STYLE + """</head><body>
<div class="app"><div class="main">
 <div class="top">""" + LOGO + """<div class="ttl">الإعدادات</div></div>
 <div class="wrap">
   <a class="back" href="/" style="color:#4f46e5;font-weight:700;display:inline-block;margin-bottom:10px">&#8594; رجوع للتقارير</a>
   {% if saved %}<div style="background:#e8f4ec;color:#1e7b34;padding:10px 14px;border-radius:8px;margin:6px 0 12px">تم حفظ الإعدادات</div>{% endif %}
   <h1>إظهار / إخفاء التبويبات والتقارير</h1>
   <p style="color:#6b7280;font-size:13px;margin-bottom:12px">ضع علامة على ما تريد إخفاءه من الواجهة، ثم احفظ.</p>
   <form method="post" action="/settings">
     <input type="hidden" name="action" value="save">
     <div class="card" style="margin-bottom:16px;border:2px solid #f59e0b;background:#fffbeb">
       <label style="font-weight:800;font-size:15px;color:#b45309"><input type="checkbox" name="hide_profit" {{ 'checked' if hide_profit else '' }}> 🔒 إخفاء كل ما يخص الربح من النظام</label>
       <div style="margin-top:6px;color:#92400e;font-size:12.5px">عند التفعيل يُخفى: تبويب «الربحية» بالكامل، بطاقتا «مجمل الربح» و«صافي الربح» في لوحة القيادة، وتقريرا «قائمة الدخل» و«مراكز التكلفة» في التبويب المالي.</div>
     </div>
     {% for t in tabs %}
       <div class="card" style="margin-bottom:12px">
         <label style="font-weight:700;font-size:15px"><input type="checkbox" name="tab_{{t.id}}" {{ 'checked' if t.id in hidden_tabs else '' }}> إخفاء التبويب كاملاً: {{t.title}}</label>
         <div style="margin-top:10px;padding-right:18px;display:flex;flex-wrap:wrap;gap:14px">
           {% for r in t.reports %}
             <label style="font-size:13px;color:#374151"><input type="checkbox" name="rep_{{t.id}}/{{r.id}}" {{ 'checked' if (t.id ~ '/' ~ r.id) in hidden_reports else '' }}> {{r.title}}</label>
           {% endfor %}
         </div>
       </div>
     {% endfor %}
     <button type="submit" style="background:#4f46e5;color:#fff;border:0;padding:12px 24px;border-radius:9px;font-weight:700;font-size:15px;cursor:pointer">حفظ الإعدادات</button>
   </form>
 </div>
</div></div></body></html>"""

PIN_PAGE = """<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>رمز الدخول</title>""" + STYLE + """</head><body>
<div class="app"><div class="main">
 <div class="top">""" + LOGO + """<div class="ttl">الإعدادات</div></div>
 <div class="wrap">
   <a class="back" href="/" style="color:#4f46e5;font-weight:700;display:inline-block;margin-bottom:16px">&#8594; رجوع للتقارير</a>
   <div class="card" style="max-width:380px;margin:40px auto;text-align:center">
     <div style="font-size:40px;margin-bottom:6px">🔒</div>
     <h1 style="font-size:18px;margin:0 0 4px">تبويب الإعدادات محمي</h1>
     <p style="color:#6b7280;font-size:13px;margin:0 0 16px">أدخل رمز الدخول للمتابعة</p>
     {% if error %}<div class="err" style="margin-bottom:12px">رمز الدخول غير صحيح</div>{% endif %}
     <form method="post" action="/settings">
       <input type="password" name="pin" autofocus inputmode="numeric" placeholder="• • • • •"
              style="width:100%;text-align:center;letter-spacing:8px;font-size:22px;padding:12px;border:1.5px solid #cbd5e1;border-radius:10px;margin-bottom:14px">
       <button type="submit" style="width:100%;background:#4f46e5;color:#fff;border:0;padding:12px;border-radius:10px;font-weight:700;font-size:15px;cursor:pointer">دخول</button>
     </form>
   </div>
 </div>
</div></div></body></html>"""

@app.route("/settings/logout")
def settings_logout():
    session.pop("set_ok", None)
    return render_template_string(PIN_PAGE, error=False)

@app.route("/settings", methods=["GET","POST"])
def settings():
    # بوابة رمز الدخول
    if not session.get("set_ok"):
        if request.method == "POST" and request.form.get("pin") is not None:
            if request.form.get("pin") == SETTINGS_PIN:
                session["set_ok"] = True
            else:
                return render_template_string(PIN_PAGE, error=True)
        else:
            return render_template_string(PIN_PAGE, error=False)
    saved = False
    if request.method == "POST" and request.form.get("action") == "save":
        htabs = [t["id"] for t in TABS if request.form.get("tab_"+t["id"])]
        hreps = []
        for t in TABS:
            for r in t["reports"]:
                key = t["id"]+"/"+r["id"]
                if request.form.get("rep_"+key):
                    hreps.append(key)
        save_hidden(htabs, hreps, bool(request.form.get("hide_profit")))
        saved = True
    ht, hr = load_hidden_raw()
    return render_template_string(SETTINGS_PAGE, tabs=TABS, hidden_tabs=ht, hidden_reports=hr,
                                  saved=saved, hide_profit=load_hide_profit())



DASHBOARD_PAGE = '''<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>لوحة القيادة SREEN</title>
     <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
     <script>
     var D={{ dash|tojson }};
     window.addEventListener("load",function(){ 
       if(!window.Chart) return; 
       Chart.defaults.font.family = "'Cairo', 'Inter', sans-serif";
       Chart.defaults.color = "#64748b";
       
       const commonOptions = {
         responsive: true,
         maintainAspectRatio: false,
         plugins: {
           legend: { display: false },
           tooltip: { backgroundColor: '#1e293b', padding: 14, titleFont: { size: 14, family: "'Cairo', sans-serif", weight: 'bold' }, bodyFont: { size: 14, family: "'Cairo', sans-serif" }, cornerRadius: 10, displayColors: true, boxPadding: 6 }
         }
       };
       
       // C1: Bar Chart (Sales & Collection)
       new Chart(document.getElementById("c1"),{
         type:"bar",
         data:{
           labels:D.months,
           datasets:[
             {label:"مبيعات", data:D.msales, backgroundColor:"#4f46e5", borderRadius:8, maxBarThickness: 32},
             {label:"تحصيل", data:D.mcollect, backgroundColor:"#38bdf8", borderRadius:8, maxBarThickness: 32}
           ]
         },
         options: {
           ...commonOptions,
           plugins: { ...commonOptions.plugins, legend: { display: true, position: 'top', align: 'end', labels: { usePointStyle: true, boxWidth: 10, font: { family: "'Cairo'", size: 13, weight: 'bold' } } } },
           scales: {
             x: { grid: { display: false }, border: { display: false } },
             y: { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false } }
           }
         }
       });

       // C2: Doughnut (Salesmen)
       new Chart(document.getElementById("c2"),{
         type:"doughnut",
         data:{
           labels:D.rep_labels.slice(0,5),
           datasets:[{data:D.rep_vals.slice(0,5), backgroundColor:["#4f46e5", "#38bdf8", "#10b981", "#f59e0b", "#8b5cf6"], borderWidth: 0, hoverOffset: 4}]
         },
         options: {
           responsive: true, maintainAspectRatio: false, cutout: '75%',
           plugins: { legend: { display: true, position: 'bottom', labels: { usePointStyle: true, boxWidth: 8, font: { size: 11 } } } }
         }
       });

       // C3: Doughnut (Items)
       new Chart(document.getElementById("c3"),{
         type:"doughnut",
         data:{
           labels:D.itm_labels.slice(0,5),
           datasets:[{data:D.itm_vals.slice(0,5), backgroundColor:["#f43f5e", "#d946ef", "#0ea5e9", "#14b8a6", "#eab308"], borderWidth: 0, hoverOffset: 4}]
         },
         options: {
           responsive: true, maintainAspectRatio: false, cutout: '75%',
           plugins: { legend: { display: true, position: 'bottom', labels: { usePointStyle: true, boxWidth: 8, font: { size: 11 } } } }
         }
       });

       // C4: Line Chart (Purchases)
       const ctx4 = document.getElementById("c4").getContext('2d');
       const grad4 = ctx4.createLinearGradient(0, 0, 0, 300);
       grad4.addColorStop(0, 'rgba(16, 185, 129, 0.4)');
       grad4.addColorStop(1, 'rgba(16, 185, 129, 0.0)');
       
       new Chart(ctx4,{
         type:"line",
         data:{
           labels:D.months,
           datasets:[{
             label: "مشتريات", data:D.mpurch, borderColor:"#10b981", borderWidth: 3, backgroundColor: grad4, fill:true, tension:0.4, pointRadius: 0, pointHoverRadius: 6, pointBackgroundColor: "#fff", pointBorderColor: "#10b981", pointBorderWidth: 2
           }]
         },
         options: {
           ...commonOptions,
           interaction: { mode: 'index', intersect: false },
           scales: {
             x: { grid: { display: false }, border: { display: false } },
             y: { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false } }
           }
         }
       });
     });
     </script>
<style>
 *{box-sizing:border-box;font-family:Tahoma,Arial}
 body{margin:0;background:#f1f5f9;color:#0f172a}
 .hd{background:linear-gradient(90deg,#0f766e,#134e4a);color:#fff;padding:14px 24px;display:flex;align-items:center;gap:14px;flex-wrap:wrap}
 .hd h1{margin:0;font-size:19px} .hd .sp{flex:1}
 .hd a{color:#fff;text-decoration:none;background:rgba(255,255,255,.15);padding:8px 14px;border-radius:8px;font-size:14px}
 .hd form{display:flex;gap:8px;align-items:center;font-size:13px}
 .hd input{padding:7px;border:0;border-radius:6px} .hd button{padding:8px 14px;border:0;border-radius:6px;background:#fbbf24;font-weight:700;cursor:pointer}
 .wrap{padding:22px;max-width:1300px;margin:auto}
 .kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:22px}
 @media(max-width:900px){.kpis{grid-template-columns:repeat(2,1fr)}}
 .kpi{background:#fff;border-radius:14px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.08);border-right:4px solid #0f766e}
 .kpi .l{color:#64748b;font-size:13px;margin-bottom:6px} .kpi .v{font-size:21px;font-weight:800}
 .kpi.g{border-color:#16a34a}.kpi.b{border-color:#2563eb}.kpi.o{border-color:#ea580c}.kpi.r{border-color:#dc2626}.kpi.p{border-color:#7c3aed}
 .charts{display:grid;grid-template-columns:1fr 1fr;gap:14px}
 @media(max-width:900px){.charts{grid-template-columns:1fr}}
 .ch{background:#fff;border-radius:14px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
 .ch h3{margin:0 0 12px;font-size:15px}
</style></head><body>
 <div class="hd"><h1>📊 لوحة القيادة — SREEN</h1>
   <form method="get" action="/dashboard"><span>من</span><input type="date" name="date_from" value="{{f}}"><span>إلى</span><input type="date" name="date_to" value="{{t}}"><button type="submit">تحديث</button></form>
   <div class="sp"></div><a href="/">← التقارير</a></div>
 <div class="wrap">
   <div class="kpis">
     <div class="kpi b"><div class="l">إجمالي المبيعات</div><div class="v">{{ "{:,.0f}".format(data.sales) }}</div></div>
     <div class="kpi g"><div class="l">إجمالي التحصيل</div><div class="v">{{ "{:,.0f}".format(data.collect) }}</div></div>
     <div class="kpi o"><div class="l">إجمالي المشتريات</div><div class="v">{{ "{:,.0f}".format(data.purch) }}</div></div>
     {% if not hide_profit|default(false) %}<div class="kpi p"><div class="l">مجمل الربح</div><div class="v">{{ "{:,.0f}".format(data.gross) }}</div></div>
     <div class="kpi g"><div class="l">صافي الربح</div><div class="v">{{ "{:,.0f}".format(data.netprofit) }}</div></div>{% endif %}
     <div class="kpi r"><div class="l">الذمم المدينة</div><div class="v">{{ "{:,.0f}".format(data.recv) }}</div></div>
     <div class="kpi b"><div class="l">قيمة المخزون</div><div class="v">{{ "{:,.0f}".format(data.invval) }}</div></div>
     <div class="kpi o"><div class="l">صافي الضريبة</div><div class="v">{{ "{:,.0f}".format(data.vat) }}</div></div>
   </div>
   <div class="charts">
     <div class="ch"><h3>المبيعات والتحصيل شهرياً</h3><canvas id="c1" height="140"></canvas></div>
     <div class="ch"><h3>أفضل المناديب (مبيعات)</h3><canvas id="c2" height="140"></canvas></div>
     <div class="ch"><h3>أفضل الأصناف (مبيعات)</h3><canvas id="c3" height="140"></canvas></div>
     <div class="ch"><h3>المشتريات شهرياً</h3><canvas id="c4" height="140"></canvas></div>
   </div></div>
<script>
const D={{ data|tojson }};
Chart.defaults.font.family="Tahoma";
new Chart(c1,{type:"bar",data:{labels:D.months,datasets:[{label:"مبيعات",data:D.msales,backgroundColor:"#2563eb"},{label:"تحصيل",data:D.mcollect,backgroundColor:"#16a34a"}]}});
new Chart(c2,{type:"bar",data:{labels:D.rep_labels,datasets:[{label:"مبيعات",data:D.rep_vals,backgroundColor:"#0f766e"}]},options:{indexAxis:"y",plugins:{legend:{display:false}}}});
new Chart(c3,{type:"bar",data:{labels:D.itm_labels,datasets:[{label:"مبيعات",data:D.itm_vals,backgroundColor:"#ea580c"}]},options:{indexAxis:"y",plugins:{legend:{display:false}}}});
new Chart(c4,{type:"line",data:{labels:D.months,datasets:[{label:"مشتريات",data:D.mpurch,borderColor:"#ea580c",backgroundColor:"rgba(234,88,12,.12)",fill:true,tension:.3}]},options:{plugins:{legend:{display:false}}}});
</script></body></html>'''

def compute_dash(f, t):
    b = {"f": f, "t": t}
    P="TO_DATE(:f,\'YYYY-MM-DD\')"; Q="TO_DATE(:t,\'YYYY-MM-DD\')+1"
    d = {"sales":0,"collect":0,"purch":0,"gross":0,"netprofit":0,"recv":0,"invval":0,"vat":0,
         "months":[],"msales":[],"mcollect":[],"mpurch":[],"rep_labels":[],"rep_vals":[],"itm_labels":[],"itm_vals":[]}
    try:
        with get_conn() as con:
            cur = con.cursor()
            def sc(sql):
                try:
                    cur.execute(sql,{k:v for k,v in b.items() if (":"+k) in sql}); r=cur.fetchone()
                    return round(float(r[0]),2) if r and r[0] is not None else 0.0
                except Exception: return 0.0
            def rw(sql):
                try:
                    cur.execute(sql,{k:v for k,v in b.items() if (":"+k) in sql}); return cur.fetchall()
                except Exception: return []
            d["sales"]=sc("SELECT NVL(SUM(NVL(BILL_AMT,0)-NVL(DISC_AMT,0)+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)),0) FROM IAS20261.IAS_BILL_MST WHERE BILL_DOC_TYPE IN (1,4) AND BILL_DATE>="+P+" AND BILL_DATE<"+Q)
            d["collect"]=sc("SELECT NVL(SUM(NVL(CR_AMT,0)),0) FROM IAS20261.IAS_POST_DTL WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=2 AND C_CODE IS NOT NULL AND DOC_DATE>="+P+" AND DOC_DATE<"+Q)
            d["purch"]=sc("SELECT NVL(SUM(NVL(BILL_AMT,0)-NVL(DISC_AMT,0)+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)),0) FROM IAS20261.IAS_PI_BILL_MST WHERE BILL_DATE>="+P+" AND BILL_DATE<"+Q)
            d["gross"]=sc("SELECT NVL(SUM(NVL(x.I_QTY,0)*(NVL(x.I_PRICE,0)-NVL(x.DIS_AMT,0)+NVL(x.OTHR_AMT,0))-NVL(x.I_QTY,0)*NVL(x.STK_COST,0)),0) FROM IAS20261.IAS_BILL_DTL x JOIN IAS20261.IAS_BILL_MST m ON m.BILL_SER=x.BILL_SER WHERE m.BILL_DOC_TYPE IN (1,4) AND m.BILL_DATE>="+P+" AND m.BILL_DATE<"+Q)
            d["netprofit"]=sc("SELECT NVL(SUM(NVL(p.CR_AMT,0)-NVL(p.DR_AMT,0)),0) FROM IAS20261.IAS_POST_DTL p JOIN IAS20261.ACCOUNT a ON a.A_CODE=p.A_CODE WHERE NVL(p.DOC_POST,0)=1 AND a.A_REPORT=2 AND p.DOC_DATE>="+P+" AND p.DOC_DATE<"+Q)
            d["recv"]=sc("SELECT NVL(SUM(bal),0) FROM (SELECT SUM(NVL(DR_AMT,0)-NVL(CR_AMT,0)) bal FROM IAS20261.IAS_POST_DTL WHERE NVL(DOC_POST,0)=1 AND C_CODE IS NOT NULL AND DOC_DATE<"+Q+" GROUP BY C_CODE HAVING SUM(NVL(DR_AMT,0)-NVL(CR_AMT,0))>0)")
            d["invval"]=sc("SELECT NVL(SUM(NVL(I_QTY,0)*NVL(IN_OUT,0)*NVL(STK_COST,0)),0) FROM IAS20261.ITEM_MOVEMENT WHERE I_DATE<"+Q)
            ov=sc("SELECT NVL(SUM(NVL(VAT_AMT,0)),0) FROM IAS20261.IAS_BILL_MST WHERE BILL_DOC_TYPE IN (1,4) AND BILL_DATE>="+P+" AND BILL_DATE<"+Q)
            iv=sc("SELECT NVL(SUM(NVL(VAT_AMT,0)),0) FROM IAS20261.IAS_PI_BILL_MST WHERE BILL_DATE>="+P+" AND BILL_DATE<"+Q)
            d["vat"]=round(ov-iv,2)
            def mm(sql):
                m={}
                for r in rw(sql):
                    m[str(r[0])]=round(float(r[1] or 0),2)
                return m
            ms=mm("SELECT TO_CHAR(BILL_DATE,\'YYYY-MM\'), SUM(NVL(BILL_AMT,0)-NVL(DISC_AMT,0)+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)) FROM IAS20261.IAS_BILL_MST WHERE BILL_DOC_TYPE IN (1,4) AND BILL_DATE>="+P+" AND BILL_DATE<"+Q+" GROUP BY TO_CHAR(BILL_DATE,\'YYYY-MM\')")
            mc=mm("SELECT TO_CHAR(DOC_DATE,\'YYYY-MM\'), SUM(NVL(CR_AMT,0)) FROM IAS20261.IAS_POST_DTL WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=2 AND C_CODE IS NOT NULL AND DOC_DATE>="+P+" AND DOC_DATE<"+Q+" GROUP BY TO_CHAR(DOC_DATE,\'YYYY-MM\')")
            mp=mm("SELECT TO_CHAR(BILL_DATE,\'YYYY-MM\'), SUM(NVL(BILL_AMT,0)-NVL(DISC_AMT,0)+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)) FROM IAS20261.IAS_PI_BILL_MST WHERE BILL_DATE>="+P+" AND BILL_DATE<"+Q+" GROUP BY TO_CHAR(BILL_DATE,\'YYYY-MM\')")
            months=sorted(set(list(ms)+list(mc)+list(mp)))
            d["months"]=months
            d["msales"]=[ms.get(x,0) for x in months]
            d["mcollect"]=[mc.get(x,0) for x in months]
            d["mpurch"]=[mp.get(x,0) for x in months]
            for r in rw("SELECT NVL(sm.REPRS_A_NAME, m.REP_CODE) nm, SUM(NVL(m.BILL_AMT,0)-NVL(m.DISC_AMT,0)+NVL(m.VAT_AMT,0)+NVL(m.OTHR_AMT,0)) v FROM IAS20261.IAS_BILL_MST m LEFT JOIN IAS20261.SALES_MAN sm ON sm.REPRS_CODE=m.REP_CODE WHERE m.BILL_DOC_TYPE IN (1,4) AND m.BILL_DATE>="+P+" AND m.BILL_DATE<"+Q+" GROUP BY NVL(sm.REPRS_A_NAME,m.REP_CODE) ORDER BY v DESC FETCH FIRST 7 ROWS ONLY"):
                d["rep_labels"].append(str(r[0])); d["rep_vals"].append(round(float(r[1] or 0),2))
            for r in rw("SELECT NVL(i.I_NAME, x.I_CODE) nm, SUM(NVL(x.I_QTY,0)*(NVL(x.I_PRICE,0)-NVL(x.DIS_AMT,0)+NVL(x.OTHR_AMT,0))) v FROM IAS20261.IAS_BILL_DTL x JOIN IAS20261.IAS_BILL_MST m ON m.BILL_SER=x.BILL_SER LEFT JOIN IAS20261.IAS_ITM_MST i ON i.I_CODE=x.I_CODE WHERE m.BILL_DOC_TYPE IN (1,4) AND m.BILL_DATE>="+P+" AND m.BILL_DATE<"+Q+" GROUP BY NVL(i.I_NAME,x.I_CODE) ORDER BY v DESC FETCH FIRST 7 ROWS ONLY"):
                d["itm_labels"].append(str(r[0])[:22]); d["itm_vals"].append(round(float(r[1] or 0),2))
    except Exception as e:
        d["err"]=str(e)
    return d


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
