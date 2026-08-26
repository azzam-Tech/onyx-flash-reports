import openpyxl

try:
    new_file = r"C:\Users\amarn\Downloads\الثلاجات_تصنيف_اونكس_برو_مرتب (1).xlsx"
    wb = openpyxl.load_workbook(new_file, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    print(f"Headers in NEW file: {headers}\n")
    
    print("Searching for ADSD50MWQ in NEW file:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_str = [str(x) if x is not None else "" for x in row]
        if 'ADSD50MWQ' in row_str:
            print(f"Row {row_idx}: {row}")
            
    print("\nSearching for ARRRE-69S in NEW file:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_str = [str(x) if x is not None else "" for x in row]
        if 'ARRRE-69S' in row_str:
            print(f"Row {row_idx}: {row}")
            
except Exception as e:
    print(f"Error: {e}")
