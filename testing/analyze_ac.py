import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_conn

def get_latest_excel():
    downloads_dir = r"C:\Users\amarn\Downloads"
    files = []
    for f in os.listdir(downloads_dir):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            files.append(os.path.join(downloads_dir, f))
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0] if files else None

def analyze_ac_file():
    file_path = get_latest_excel()
    if not file_path:
        print("No Excel files found in Downloads.")
        return

    print(f"Using file: {os.path.basename(file_path)}")
    out_path = os.path.join(os.path.dirname(__file__), '..', 'migration_plan', 'update_ac_group.sql')
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # Let's see the header first to confirm columns
        print("\n--- HEADER ROW ---")
        for idx, col in enumerate(list(sheet.iter_rows(values_only=True))[0]):
            print(f"Col {idx}: {col}")
        
        excel_data = {}
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i > 0:
                # AC file: Col 0 is I_CODE, Col 5 is SG_CODE, Col 6 is SSG_CODE
                icode = str(row[0]).strip() if row[0] else None
                sg_code = str(row[5]).strip().zfill(3) if row[5] else None
                ssg_code = str(row[6]).strip().zfill(3) if row[6] else None
                
                # Verify Group Code? Usually Col 5 starts with group code
                if icode and sg_code and ssg_code:
                    excel_data[icode] = (sg_code, ssg_code)
                    
        print(f"\nTotal valid items in AC Excel: {len(excel_data)}")
        
        # Now check DB for G_CODE '003' (Assuming AC is 003, let's verify if AC is 003 or 004)
        # Actually I need to know the G_CODE for AC. Let's query it from one of the items.
        first_item = list(excel_data.keys())[0] if excel_data else None
        
        with get_conn() as con:
            with con.cursor() as cur:
                if first_item:
                    cur.execute("SELECT G_CODE FROM IAS_ITM_MST WHERE I_CODE = :1", [first_item])
                    res = cur.fetchone()
                    ac_g_code = res[0] if res else '003'
                else:
                    ac_g_code = '003'
                    
                print(f"Detected G_CODE for ACs: {ac_g_code}")
                
                cur.execute("SELECT I_CODE, NVL(INACTIVE, 0) FROM IAS_ITM_MST WHERE G_CODE = :1", [ac_g_code])
                db_items = cur.fetchall()
                
                db_icodes = {str(r[0]).strip(): r[1] for r in db_items}
                
                matched = sum(1 for icode in excel_data if icode in db_icodes)
                missing = sum(1 for icode in db_icodes if icode not in excel_data)
                
                print(f"Total AC items in DB (Group {ac_g_code}): {len(db_icodes)}")
                print(f"Matched items (will be updated): {matched}")
                print(f"Missing items (will be deactivated): {missing}")
                
                # Generate SQL
                with open(out_path, 'w', encoding='utf-8') as f:
                    f.write("-- ==================================================\n")
                    f.write(f"-- تحديث المجموعات الفرعية وتحت الفرعية للمكيفات (مجموعة {ac_g_code})\n")
                    f.write("-- ==================================================\n")
                    
                    for icode, (sg, ssg) in excel_data.items():
                        if icode in db_icodes:
                            f.write(f"UPDATE IAS_ITM_MST SET MNG_CODE = '{sg}', SUBG_CODE = '{ssg}', INACTIVE = 0 WHERE I_CODE = '{icode}';\n")
                        
                    f.write("\n-- ==================================================\n")
                    f.write("-- تجميد المكيفات المتبقية (التي لم ترد في الملف)\n")
                    f.write("-- ==================================================\n")
                    
                    for icode in db_icodes:
                        if icode not in excel_data:
                            f.write(f"UPDATE IAS_ITM_MST SET INACTIVE = 1 WHERE I_CODE = '{icode}';\n")
                            
                    f.write("\nCOMMIT;\n")
                    
                print(f"\nSQL script generated at: {out_path}")
                
    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    analyze_ac_file()
