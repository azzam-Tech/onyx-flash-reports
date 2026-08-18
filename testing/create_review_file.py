import sys
import os
import openpyxl
from openpyxl.styles import PatternFill

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_conn

def create_consolidated_review_file():
    manager_file = r"C:\Users\amarn\Downloads\الثلاجات_تصنيف_اونكس_برو_مرتب.xlsx"
    out_file = r"C:\Users\amarn\Downloads\الثلاجات_للمراجعة_النهائية.xlsx"
    
    try:
        # Read Manager's Excel
        wb_mgr = openpyxl.load_workbook(manager_file)
        sheet_mgr = wb_mgr.active
        manager_icodes = set()
        for i, row in enumerate(sheet_mgr.iter_rows(values_only=True)):
            if i == 0: continue
            icode = str(row[8]).strip() if row[8] else None
            if icode:
                manager_icodes.add(icode)
                
        with get_conn() as con:
            with con.cursor() as cur:
                # Active Items
                cur.execute("""
                    SELECT DISTINCT I_CODE 
                    FROM ITEM_MOVEMENT 
                    WHERE I_CODE IN (SELECT I_CODE FROM IAS_ITM_MST WHERE G_CODE = '003')
                """)
                active_icodes = {r[0] for r in cur.fetchall()}
                
                # Stagnant Items
                cur.execute("""
                    SELECT I_CODE 
                    FROM IAS_ITM_MST 
                    WHERE G_CODE = '003' 
                    AND I_CODE NOT IN (
                        SELECT DISTINCT I_CODE FROM ITEM_MOVEMENT WHERE I_CODE IS NOT NULL
                    )
                """)
                stagnant_icodes = {r[0] for r in cur.fetchall()}
                
                missed_active = active_icodes - manager_icodes
                manager_stagnant = manager_icodes.intersection(stagnant_icodes)
                manager_active = manager_icodes.intersection(active_icodes)
                
                # All icodes to process
                all_icodes = manager_icodes.union(missed_active)
                
                # Fetch details
                format_strings = ','.join([f"'{c}'" for c in all_icodes])
                cur.execute(f"SELECT I_CODE, I_NAME FROM IAS_ITM_MST WHERE I_CODE IN ({format_strings}) ORDER BY I_CODE")
                rows = cur.fetchall()
                
                # Create Output Excel
                wb_out = openpyxl.Workbook()
                ws_out = wb_out.active
                ws_out.title = "المراجعة النهائية"
                
                # Headers
                ws_out.append(["رقم الصنف", "اسم الصنف", "ملاحظة"])
                
                # Colors
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                for r in rows:
                    icode = r[0]
                    iname = r[1]
                    note = ""
                    
                    ws_out.append([icode, iname, note])
                    current_row = ws_out.max_row
                    
                    if icode in missed_active:
                        ws_out.cell(row=current_row, column=1).fill = green_fill
                        ws_out.cell(row=current_row, column=2).fill = green_fill
                        ws_out.cell(row=current_row, column=3).value = "منسي (نشط - مضاف حديثاً)"
                        ws_out.cell(row=current_row, column=3).fill = green_fill
                    elif icode in manager_stagnant:
                        ws_out.cell(row=current_row, column=1).fill = red_fill
                        ws_out.cell(row=current_row, column=2).fill = red_fill
                        ws_out.cell(row=current_row, column=3).value = "مدرج بالخطأ (راكد - للحذف)"
                        ws_out.cell(row=current_row, column=3).fill = red_fill
                    else:
                        ws_out.cell(row=current_row, column=3).value = "صحيح ومعتمد"

                wb_out.save(out_file)
                print(f"Consolidated review file created: {out_file} with {len(rows)} items.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    create_consolidated_review_file()
