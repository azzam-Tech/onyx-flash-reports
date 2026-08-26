import openpyxl

def safe_str(val):
    if val is None: return ""
    return str(val).strip()

try:
    file_path = r"C:\Users\amarn\Downloads\الثلاجات_تصنيف_اونكس_برو.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print("Looking for ADSD50MWQ and ARRRE-69S in Excel file...")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        i_code = safe_str(row[5])
        if i_code in ('ADSD50MWQ', 'ARRRE-69S'):
            print(f"Row {row_idx}: {row}")
            
    # Also print the total number of rows openpyxl sees
    print(f"\nTotal rows in Excel sheet according to openpyxl: {ws.max_row}")
    
except Exception as e:
    print(f"Error: {e}")
