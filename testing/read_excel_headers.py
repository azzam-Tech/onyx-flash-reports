import openpyxl
import sys

try:
    file_path = r"C:\Users\amarn\Downloads\الثلاجات_تصنيف_اونكس_برو.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Read headers
    headers = [cell.value for cell in ws[1]]
    print("Columns found in Excel:")
    for col in headers:
        print(f"  - {col}")
        
    print("\nFirst 3 rows of data:")
    for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
        print(row)
        
except Exception as e:
    print(f"Error reading Excel: {e}")
