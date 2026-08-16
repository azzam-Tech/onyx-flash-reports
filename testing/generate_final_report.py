import oracledb
import os
import openpyxl

os.environ["PATH"] = r"C:\oracle\instantclient\instantclient_23_0;" + os.environ.get("PATH", "")
oracledb.init_oracle_client(lib_dir=r"C:\oracle\instantclient\instantclient_23_0")

def normalize(code):
    cleaned = code.replace('-', '').replace('.', '').replace('_', '').replace(' ', '').upper()
    return ''.join(sorted(cleaned))

def generate_final_report():
    try:
        manual_overrides = {
            'TWS-8SRWM': 'SRWM-8TW.',
            'TWS-12SRWM': 'SRWM-12TW.',
            'HIKT-50S4KW3': 'HIKT-50S3'
        }
        
        conn = oracledb.connect(user='RPT_USER', password='ULT2016', dsn='100.100.1.100:1521/ORCL')
        cur = conn.cursor()
        
        # 1. Get all items and mapping
        cur.execute("SELECT I_CODE, I_NAME FROM IAS20261.IAS_ITM_MST")
        all_items = cur.fetchall()
        
        db_map = {}
        for row in all_items:
            i_code = row[0]
            if i_code:
                norm_code = normalize(i_code)
                if norm_code not in db_map:
                    db_map[norm_code] = []
                db_map[norm_code].append((i_code, row[1]))
                
        # 2. Fetch current retail prices
        cur.execute("SELECT I_CODE, I_PRICE FROM IAS20261.IAS_ITEM_PRICE WHERE LEV_NO = 2")
        db_prices = {r[0]: r[1] for r in cur.fetchall() if r[0]}

        # 3. Read Excel
        file_path = r"C:\Users\amarn\OneDrive\Desktop\dbOnyxOnAntigravity\pricing\تحديث الاسعار.xlsx"
        wb_in = openpyxl.load_workbook(file_path, data_only=True)
        sheet_in = wb_in.active
        
        final_list = []
        
        for row in sheet_in.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[2]:
                continue
            
            excel_code = str(row[0]).strip()
            
            correct_code = "غير موجود"
            correct_name = "غير موجود"
            current_db_price = "غير متوفر"
            
            if excel_code in manual_overrides:
                correct_code = manual_overrides[excel_code]
            else:
                temp_code = excel_code
                if temp_code.startswith('.'):
                    temp_code = temp_code[1:] + '.'
                    
                if temp_code in db_prices:
                    correct_code = temp_code
                else:
                    norm = normalize(excel_code)
                    if norm in db_map:
                        correct_code = db_map[norm][0][0]
                        correct_name = db_map[norm][0][1]
                        
            # Get name and price if code was found
            if correct_code != "غير موجود":
                if not correct_name or correct_name == "غير موجود":
                    norm = normalize(correct_code)
                    if norm in db_map:
                        correct_name = db_map[norm][0][1]
                        
                if correct_code in db_prices:
                    current_db_price = db_prices[correct_code]

            final_list.append((excel_code, correct_code, correct_name, current_db_price))
                    
        cur.close()
        conn.close()
        
        # 4. Write to Excel
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "التقرير النهائي للأصناف"
        
        ws_out.append(["كود الصنف في الملف (الملخبط)", "الكود الصحيح في النظام", "اسم الصنف", "سعر التجزئة الحالي في النظام (مستوى 2)"])
        
        for item in final_list:
            ws_out.append([item[0], item[1], item[2], item[3]])
            
        out_path = r"C:\Users\amarn\OneDrive\Desktop\dbOnyxOnAntigravity\pricing\التقرير_النهائي_لجميع_الأصناف.xlsx"
        wb_out.save(out_path)
        print(f"Successfully generated Excel with {len(final_list)} items at: {out_path}")
        
    except Exception as e:
        print(f"Failed: {e}")

if __name__ == '__main__':
    generate_final_report()
