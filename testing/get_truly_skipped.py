import oracledb
import os
import openpyxl

os.environ["PATH"] = r"C:\oracle\instantclient\instantclient_23_0;" + os.environ.get("PATH", "")
oracledb.init_oracle_client(lib_dir=r"C:\oracle\instantclient\instantclient_23_0")

def normalize(code):
    cleaned = code.replace('-', '').replace('.', '').replace('_', '').replace(' ', '').upper()
    return ''.join(sorted(cleaned))

def get_truly_skipped():
    try:
        manual_overrides = {
            'TWS-8SRWM': 'SRWM-8TW.',
            'TWS-12SRWM': 'SRWM-12TW.',
            'HIKT-50S4KW3': 'HIKT-50S3'
        }
        
        conn = oracledb.connect(user='RPT_USER', password='ULT2016', dsn='100.100.1.100:1521/ORCL')
        cur = conn.cursor()
        
        # 1. Get ALL items updated today by our script
        cur.execute("""
            SELECT DISTINCT I_CODE 
            FROM IAS_ITEM_PRICE_HISTORY 
            WHERE AUD_U_ID = 999 AND AUD_DATE >= TRUNC(SYSDATE)
        """)
        updated_codes = {r[0] for r in cur.fetchall() if r[0]}
        
        # 2. Get all DB items
        cur.execute("SELECT I_CODE, I_NAME FROM IAS_ITM_MST")
        all_items = cur.fetchall()
        
        db_map = {}
        for row in all_items:
            i_code = row[0]
            if i_code:
                norm_code = normalize(i_code)
                if norm_code not in db_map:
                    db_map[norm_code] = []
                db_map[norm_code].append((i_code, row[1]))
                
        # 3. Get current retail prices
        cur.execute("SELECT I_CODE, I_PRICE FROM IAS_ITEM_PRICE WHERE LEV_NO = 2")
        db_prices = {r[0]: r[1] for r in cur.fetchall() if r[0]}

        # 4. Read Excel
        file_path = r"C:\Users\amarn\OneDrive\Desktop\dbOnyxOnAntigravity\pricing\تحديث الاسعار.xlsx"
        wb_in = openpyxl.load_workbook(file_path, data_only=True)
        sheet_in = wb_in.active
        
        skipped_items = []
        
        for row in sheet_in.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[2]:
                continue
            
            excel_code = str(row[0]).strip()
            try:
                excel_price = float(row[2])
            except ValueError:
                continue
                
            correct_code = None
            correct_name = ""
            
            if excel_code in manual_overrides:
                correct_code = manual_overrides[excel_code]
            else:
                temp_code = excel_code
                if temp_code.startswith('.'):
                    temp_code = temp_code[1:] + '.'
                    
                if temp_code in db_prices:
                    correct_code = temp_code
                else:
                    norm = normalize(excel_code)
                    if norm in db_map:
                        correct_code = db_map[norm][0][0]
                        correct_name = db_map[norm][0][1]
                        
            if not correct_code:
                continue
                
            if not correct_name:
                norm = normalize(correct_code)
                if norm in db_map:
                    correct_name = db_map[norm][0][1]

            # The core logic: if it is IN updated_codes, it was modified today (not skipped)
            if correct_code in updated_codes:
                continue

            # If it's NOT in updated_codes, and its price matches Excel, it was perfectly identical from the start
            if correct_code in db_prices:
                db_price = db_prices[correct_code]
                if db_price == excel_price:
                    skipped_items.append((excel_code, correct_code, correct_name, db_price))
                    
        cur.close()
        conn.close()
        
        # 5. Write to Excel
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "الأصناف المتطابقة من الأساس"
        
        ws_out.append(["الكود الملخبط في الإكسيل", "الكود الصحيح في النظام", "اسم الصنف", "سعر التجزئة (كان صحيحاً من البداية)"])
        
        for item in skipped_items:
            ws_out.append([item[0], item[1], item[2], item[3]])
            
        out_path = r"C:\Users\amarn\OneDrive\Desktop\dbOnyxOnAntigravity\pricing\الأصناف_المتطابقة_من_الأساس.xlsx"
        wb_out.save(out_path)
        print(f"Successfully generated Excel with {len(skipped_items)} truly skipped items at: {out_path}")
        
    except Exception as e:
        print(f"Failed: {e}")

if __name__ == '__main__':
    get_truly_skipped()
