import sys
import os
import openpyxl

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_pooled_conn

try:
    with get_pooled_conn() as conn:
        with conn.cursor() as cur:
            # Query the full hierarchy
            query = """
                SELECT 
                    cntry.CNTRY_NO, cntry.CNTRY_A_NAME,
                    prov.PROV_NO, prov.PROV_A_NAME,
                    city.CITY_NO, city.CITY_A_NAME,
                    rgn.R_CODE, rgn.R_A_NAME
                FROM CNTRY cntry
                LEFT JOIN IAS_PROVINCES prov ON prov.CNTRY_NO = cntry.CNTRY_NO
                LEFT JOIN CITIES city ON city.PROV_NO = prov.PROV_NO AND city.CNTRY_NO = cntry.CNTRY_NO
                LEFT JOIN REGIONS rgn ON rgn.CITY_NO = city.CITY_NO AND rgn.PROV_NO = prov.PROV_NO AND rgn.CNTRY_NO = cntry.CNTRY_NO
                ORDER BY cntry.CNTRY_NO, prov.PROV_NO, city.CITY_NO, rgn.R_CODE
            """
            cur.execute(query)
            rows = cur.fetchall()
            
            # Create Excel workbook and sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "الهيكل الجغرافي كامل"
            ws.sheet_view.rightToLeft = True # Enable RTL for Arabic
            
            # Write Headers
            headers = [
                'رقم الدولة', 'اسم الدولة',
                'رقم المحافظة', 'اسم المحافظة',
                'رقم المدينة', 'اسم المدينة',
                'رقم المنطقة البيعية', 'اسم المنطقة البيعية'
            ]
            ws.append(headers)
            
            # Add some basic styling
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                
            # Write Data
            for row in rows:
                ws.append(row)
                
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
                
            os.makedirs('Results', exist_ok=True)
            output_path = 'Results/Full_Geographical_Tree.xlsx'
            wb.save(output_path)
            
            print(f"Successfully generated Excel file at {output_path}")

except Exception as e:
    print(f"Error: {e}")
