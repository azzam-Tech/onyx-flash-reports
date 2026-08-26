import openpyxl
import os

try:
    file_path = r"C:\Users\amarn\Downloads\الثلاجات_تصنيف_اونكس_برو.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    with open('testing/excel_structure.txt', 'w', encoding='utf-8') as f:
        f.write("Columns:\n")
        for i, col in enumerate(headers):
            f.write(f"{i}: {col}\n")
            
        f.write("\nFirst 3 rows:\n")
        for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
            f.write(str(row) + "\n")
            
except Exception as e:
    print(f"Error: {e}")
