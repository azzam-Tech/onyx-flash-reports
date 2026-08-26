import sys
import os
import openpyxl
from openpyxl.utils import get_column_letter

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_pooled_conn

try:
    with get_pooled_conn() as conn:
        with conn.cursor() as cur:
            # Get column names first
            cur.execute("""
                SELECT COLUMN_NAME 
                FROM ALL_TAB_COLUMNS 
                WHERE OWNER = 'IAS20261' AND TABLE_NAME = 'IAS_SUB_GRP_DTL'
                ORDER BY COLUMN_ID
            """)
            cols = [r[0] for r in cur.fetchall()]
            
            # Fetch the data
            cur.execute("""
                SELECT * 
                FROM IAS20261.IAS_SUB_GRP_DTL 
                ORDER BY SUBG_CODE
            """)
            rows = cur.fetchall()
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "المجموعات تحت الفرعية (الألوان)"
            
            # Write headers
            ws.append(cols)
            
            # Format datetimes in rows
            import datetime
            def format_cell(val):
                if isinstance(val, datetime.datetime):
                    return val.strftime("%Y-%m-%d %H:%M:%S")
                return val
                
            # Write rows
            for r in rows:
                ws.append([format_cell(val) for val in r])
                
            # Adjust column widths slightly
            for i, col in enumerate(cols, 1):
                ws.column_dimensions[get_column_letter(i)].width = 20
                
            # Save file
            out_path = os.path.join(os.path.dirname(__file__), '..', 'sub_sub_groups.xlsx')
            wb.save(out_path)
            print(f"File successfully saved to: {out_path}")
            print(f"Total Sub-Sub-Groups extracted: {len(rows)}")
            
except Exception as e:
    print(f"Error: {e}")
