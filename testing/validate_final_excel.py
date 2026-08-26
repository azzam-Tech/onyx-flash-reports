import openpyxl
from collections import Counter
import sys
import os

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_pooled_conn

def safe_str(val):
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val)).strip()
    return str(val).strip()

try:
    file_path = r"C:\Users\amarn\Downloads\الثلاجات_تصنيف_اونكس_برو_محدث3.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    
    # Let's guess the column indices based on typical structure:
    # 0: G_CODE, 2: MNG_CODE, 5: SUBG_CODE, 8: I_CODE, 10: DETAIL_NO?
    # We will just iterate and print the first row to be sure.
    first_data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    print(f"First data row: {first_data_row}")
    
    i_codes = []
    data_to_update = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # We will dynamically find columns if possible, or just assume:
        # Assuming typical 11 column structure: row[8] is I_CODE, row[2] is MNG_CODE, row[5] is SUBG_CODE
        # If it's a 7 column structure: row[5] is I_CODE, row[2] is MNG_CODE, row[3] is SUBG_CODE
        # Let's use heuristics based on length
        if len(row) > 8 and safe_str(row[8]):
            i_code = safe_str(row[8])
            mng_code = safe_str(row[2])
            subg_code = safe_str(row[5])
        elif len(row) > 5 and safe_str(row[5]):
            i_code = safe_str(row[5])
            mng_code = safe_str(row[2])
            subg_code = safe_str(row[3])
        else:
            continue
            
        if i_code:
            i_codes.append(i_code)
            data_to_update.append((i_code, mng_code, subg_code))
            
    # 1. Check duplicates
    counts = Counter(i_codes)
    duplicates = {k: v for k, v in counts.items() if v > 1}
    
    print(f"\n--- Validation Report ---")
    print(f"Total Rows with Items: {len(i_codes)}")
    print(f"Unique Items: {len(counts)}")
    if duplicates:
        print(f"WARNING: Found {len(duplicates)} duplicate item codes!")
        for k, v in list(duplicates.items())[:5]:
            print(f"  - {k} appears {v} times")
    else:
        print("SUCCESS: No duplicate items found in the Excel file.")
        
    # 2. Validate against Database
    with get_pooled_conn() as conn:
        with conn.cursor() as cur:
            i_codes_list = list(counts.keys())
            
            binds = [f":{i+1}" for i in range(len(i_codes_list))]
            query = f"""
                SELECT I_CODE 
                FROM IAS20261.IAS_ITM_MST 
                WHERE I_CODE IN ({','.join(binds)})
            """
            cur.execute(query, i_codes_list)
            
            db_codes = set(r[0] for r in cur.fetchall())
            missing_codes = set(i_codes_list) - db_codes
            
            if missing_codes:
                print(f"WARNING: {len(missing_codes)} items NOT found in the database!")
                for c in list(missing_codes)[:5]:
                    print(f"  - {c}")
            else:
                print("SUCCESS: All items in the Excel file exist in the database.")
                
except Exception as e:
    print(f"Error: {e}")
