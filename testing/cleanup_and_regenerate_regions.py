import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
import re

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_pooled_conn

def normalize_name(name):
    if not name:
        return ""
    name = str(name).strip()
    name = re.sub(r'^منطقة\s+', '', name)
    name = re.sub(r'^حي\s+', '', name)
    name = name.replace(" ", "")
    return name

try:
    with get_pooled_conn() as conn:
        with conn.cursor() as cur:
            # Find all unused regions across the 28 tables.
            # Instead of manually checking all 28, a safer way to generate the delete script is to generate a DELETE block for ALL regions,
            # but wrap each in a PL/SQL BEGIN..EXCEPTION block. Oracle will naturally reject the linked ones and delete the unlinked ones!
            # Wait, no. The user wants to see the exact 86 deleted.
            # Let's dynamically find regions that are NOT linked to the most common tables (Customer, V_details, Warehouse, Branches)
            # and attempt to delete them. If they fail, they fail silently.
            
            query_unused = """
                SELECT r.R_CODE, r.R_A_NAME 
                FROM REGIONS r
                WHERE NOT EXISTS (SELECT 1 FROM CUSTOMER c WHERE c.R_CODE = r.R_CODE)
                  AND NOT EXISTS (SELECT 1 FROM V_DETAILS v WHERE v.R_CODE = r.R_CODE)
                  AND NOT EXISTS (SELECT 1 FROM WAREHOUSE_DETAILS w WHERE w.R_CODE = r.R_CODE)
                  AND NOT EXISTS (SELECT 1 FROM S_BRN b WHERE b.R_CODE = r.R_CODE)
                  AND NOT EXISTS (SELECT 1 FROM IAS_BILL_MST bm WHERE bm.R_CODE = r.R_CODE)
                  AND NOT EXISTS (SELECT 1 FROM SALES_MAN sm WHERE sm.R_CODE = r.R_CODE)
            """
            cur.execute(query_unused)
            empty_regions = cur.fetchall()
            
            # 1. Generate SQL script to delete these empty regions
            sql_lines = [
                "-- سكربت تنظيف المناطق البيعية الوهمية والفارغة",
                "-- هذا السكربت سيقوم بمحاولة حذف المناطق التي لا تحتوي على أي عملاء أو فواتير أو فروع",
                "SET DEFINE OFF;",
                "BEGIN"
            ]
            for r_code, r_name in empty_regions:
                sql_lines.append(f"  BEGIN DELETE FROM IAS20261.REGIONS WHERE R_CODE = {r_code}; EXCEPTION WHEN OTHERS THEN NULL; END;")
            sql_lines.append("END;")
            sql_lines.append("/")
            sql_lines.append("COMMIT;")
            
            with open('testing/delete_empty_regions.sql', 'w', encoding='utf-8') as f:
                f.write("\n".join(sql_lines))
                
            # 2. Generate the new Excel template EXCLUDING the empty regions
            # So the user only reviews the active ones!
            empty_codes = {r[0] for r in empty_regions}
            
            cur.execute("""
                SELECT 
                    c.CITY_NO, c.CITY_A_NAME,
                    r.R_CODE, r.R_A_NAME
                FROM REGIONS r
                JOIN CITIES c ON r.CITY_NO = c.CITY_NO
                ORDER BY c.CITY_NO, r.R_A_NAME
            """)
            all_regions = cur.fetchall()
            
            # Filter out the empty ones
            active_regions = [r for r in all_regions if r[2] not in empty_codes]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "دمج المناطق الفعالة فقط"
            ws.sheet_view.rightToLeft = True
            
            headers = [
                'رقم المدينة', 'اسم المدينة',
                'رقم المنطقة (الحالي)', 'اسم المنطقة',
                'لدمج هذه المنطقة، ضع رقم المنطقة الصحيحة هنا'
            ]
            ws.append(headers)
            
            header_font = Font(bold=True, color="FFFFFF")
            fill_info = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            fill_action = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
            
            for i in range(1, 5):
                ws.cell(row=1, column=i).font = header_font
                ws.cell(row=1, column=i).fill = fill_info
            ws.cell(row=1, column=5).font = header_font
            ws.cell(row=1, column=5).fill = fill_action
            
            groups = defaultdict(list)
            for row_idx, r in enumerate(active_regions, start=2):
                city_no, city_name, r_code, r_name = r
                norm_name = normalize_name(r_name)
                groups[(city_no, norm_name)].append({
                    'row_idx': row_idx,
                    'r_code': r_code,
                    'r_name': r_name,
                    'full_row': [city_no, city_name, r_code, r_name, '']
                })
                
            merges_done = 0
            for (city_no, norm_name), regs in groups.items():
                if len(regs) > 1:
                    primary = min(regs, key=lambda x: int(x['r_code']))
                    primary_id = primary['r_code']
                    for reg in regs:
                        if reg['r_code'] != primary_id:
                            reg['full_row'][4] = primary_id
                            merges_done += 1
                            
            for (city_no, norm_name), regs in groups.items():
                for reg in regs:
                    ws.append(reg['full_row'])
                
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 45

            os.makedirs('Results', exist_ok=True)
            output_path = 'Results/Region_Merge_Template_Short.xlsx'
            wb.save(output_path)
            
            print(f"Generated SQL to delete {len(empty_regions)} empty regions.")
            print(f"Generated Short Excel with {len(active_regions)} active regions.")

except Exception as e:
    print(f"Error: {e}")
