import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'privet', 'onyx_reports'))
from database import get_pooled_conn

try:
    with get_pooled_conn() as conn:
        with conn.cursor() as cur:
            # 1. Fetch wrong provinces that are actually in use
            # (either in CUSTOMER, CITIES, or REGIONS)
            query_wrong = """
                SELECT DISTINCT p.PROV_NO, p.PROV_A_NAME 
                FROM IAS_PROVINCES p
                WHERE (p.PROV_NO < 101 OR p.PROV_NO > 113)
                  AND (
                      EXISTS (SELECT 1 FROM CUSTOMER c WHERE c.PROV_NO = p.PROV_NO) OR
                      EXISTS (SELECT 1 FROM CITIES city WHERE city.PROV_NO = p.PROV_NO) OR
                      EXISTS (SELECT 1 FROM REGIONS rgn WHERE rgn.PROV_NO = p.PROV_NO)
                  )
                ORDER BY p.PROV_NO
            """
            cur.execute(query_wrong)
            wrong_provs = cur.fetchall()

            # 2. Fetch correct provinces (101 to 113)
            query_correct = """
                SELECT PROV_NO, PROV_A_NAME 
                FROM IAS_PROVINCES 
                WHERE PROV_NO BETWEEN 101 AND 113
                ORDER BY PROV_NO
            """
            cur.execute(query_correct)
            correct_provs = cur.fetchall()

            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # --- Sheet 1: Mapping Template ---
            ws_mapping = wb.active
            ws_mapping.title = "خريطة دمج المحافظات"
            ws_mapping.sheet_view.rightToLeft = True
            
            headers_map = ['رقم المحافظة الخطأ (الحالي)', 'اسم المحافظة الخطأ', 'ضع رقم المحافظة الصحيح هنا (101-113)']
            ws_mapping.append(headers_map)
            
            # Styling headers
            header_font = Font(bold=True, color="FFFFFF")
            fill_wrong = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
            fill_action = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
            
            ws_mapping['A1'].font = header_font
            ws_mapping['A1'].fill = fill_wrong
            ws_mapping['B1'].font = header_font
            ws_mapping['B1'].fill = fill_wrong
            ws_mapping['C1'].font = header_font
            ws_mapping['C1'].fill = fill_action
            
            # Add data
            for row in wrong_provs:
                ws_mapping.append([row[0], row[1], '']) # Leave column C empty for the user
                
            # Formatting widths
            ws_mapping.column_dimensions['A'].width = 25
            ws_mapping.column_dimensions['B'].width = 35
            ws_mapping.column_dimensions['C'].width = 40

            # --- Sheet 2: Reference ---
            ws_ref = wb.create_sheet(title="قائمة المحافظات المعتمدة (مرجع)")
            ws_ref.sheet_view.rightToLeft = True
            
            headers_ref = ['رقم المحافظة الصحيح', 'اسم المحافظة الصحيح']
            ws_ref.append(headers_ref)
            
            fill_ref = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws_ref['A1'].font = header_font
            ws_ref['A1'].fill = fill_ref
            ws_ref['B1'].font = header_font
            ws_ref['B1'].fill = fill_ref
            
            for row in correct_provs:
                ws_ref.append([row[0], row[1]])
                
            ws_ref.column_dimensions['A'].width = 20
            ws_ref.column_dimensions['B'].width = 30

            # Save
            os.makedirs('Results', exist_ok=True)
            output_path = 'Results/Province_Mapping_Template.xlsx'
            wb.save(output_path)
            
            print(f"Generated Mapping Template with {len(wrong_provs)} wrong provinces and {len(correct_provs)} correct provinces.")
            print(f"File saved to {output_path}")

except Exception as e:
    print(f"Error: {e}")
