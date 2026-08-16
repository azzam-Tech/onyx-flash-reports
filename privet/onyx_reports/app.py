# -*- coding: utf-8 -*-
from flask import Flask, request, render_template_string, render_template, Response, session, redirect, jsonify
import os
import json
import io
from datetime import datetime
from urllib.parse import urlencode

from config import APP_PIN, SETTINGS_PIN, load_hidden, save_hidden, load_globals, save_globals, load_hide_profit, check_permission, load_hidden_raw
from database import get_conn
from reports_config import TABS, find_report
import report_handlers
from report_handlers import *

from flask_cors import CORS

app = Flask(__name__, static_folder='public', static_url_path='/')
CORS(app, supports_credentials=True)

@app.before_request
def set_target_year():
    from flask import request, g
    year_val = request.args.get('year_val')
    date_to = request.args.get('date_to')
    date_from = request.args.get('date_from')
    
    target_year = "2026"
    if year_val and len(year_val) == 4:
        target_year = year_val
    elif date_from and len(date_from) >= 4:
        target_year = date_from[:4]
    elif date_to and len(date_to) >= 4:
        target_year = date_to[:4]
        
    g.target_year = target_year
app.secret_key = os.environ.get("SREEN_SECRET", "sreen-reports-2026-secret-key")

from flask import redirect

@app.before_request
def require_login():
    print(f"DEBUG: method={request.method}, path={request.path}, endpoint={request.endpoint}")
    if request.method == 'OPTIONS':
        return None
    if request.path.startswith('/api/'):
        return None
    # Let React handle authentication flow; just serve static assets unconditionally
    return None

@app.route("/_old_login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        pin = request.form.get("pin", "")
        
        from config import load_users
        users = load_users()
        
        # Find user by PIN
        found_username = None
        found_user_data = None
        for uname, udata in users.items():
            if udata.get("password") == pin:
                found_username = uname
                found_user_data = udata
                break
                
        if found_username:
            session['logged_in'] = True
            session['username'] = found_username
            session['role'] = found_user_data.get('role', 'user')
            return redirect('/')
        else:
            error = "ط§ظ„ط±ظ…ط² ط؛ظٹط± طµط­ظٹط­طŒ ط­ط§ظˆظ„ ظ…ط±ط© ط£ط®ط±ظ‰."
    return render_template("login.html", error=error)

@app.route("/_old_logout")
def logout():
    session.pop('logged_in', None)
    return redirect('/login')

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.json or {}
    pin = data.get("pin", "")
    
    from config import load_users
    users = load_users()
    
    found_username = None
    found_user_data = None
    for uname, udata in users.items():
        if udata.get("password") == pin:
            found_username = uname
            found_user_data = udata
            break
            
    if found_username:
        session['logged_in'] = True
        session['username'] = found_username
        session['role'] = found_user_data.get('role', 'user')
        return jsonify({
            "success": True, 
            "user": {
                "username": found_username,
                "role": session['role'],
                "allowed_tabs": found_user_data.get("allowed_tabs", []),
                "allowed_reports": found_user_data.get("allowed_reports", [])
            }
        })
    else:
        return jsonify({"success": False, "error": "ط±ظ…ط² ط§ظ„ظ…ط±ظˆط± ط؛ظٹط± طµط­ظٹط­"}), 401

@app.route("/api/session", methods=["GET"])
def api_session():
    if not session.get('logged_in'):
        return jsonify({"authenticated": False}), 401
    
    from config import load_users
    users = load_users()
    username = session.get('username')
    user_data = users.get(username, {})
    
    return jsonify({
        "authenticated": True,
        "user": {
            "username": username,
            "role": session.get('role', 'user'),
            "allowed_tabs": user_data.get("allowed_tabs", []),
            "allowed_reports": user_data.get("allowed_reports", [])
        }
    })

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.pop('logged_in', None)
    session.pop('username', None)
    session.pop('role', None)
    return jsonify({"success": True})


@app.route("/_old_index")
def index():
    username = session.get('username')
    hidden_tabs, hidden_reports = load_hidden()
    
    # Filter by user permissions
    _vis = []
    for t in TABS:
        if t["id"] in hidden_tabs: continue
        if not check_permission(username, t["id"]): continue
        
        # Filter reports inside the tab
        allowed_reports = []
        for r in t["reports"]:
            if r["id"] in hidden_reports: continue
            if check_permission(username, t["id"], r["id"]):
                allowed_reports.append(r)
        
        if allowed_reports:
            t_copy = t.copy()
            t_copy["reports"] = allowed_reports
            _vis.append(t_copy)
            
    if not _vis:
        return "ظ„ط§ طھظ…ظ„ظƒ طµظ„ط§ط­ظٹط© ظ„ط¹ط±ط¶ ط£ظٹ طھظ‚ط§ط±ظٹط±."
        
    cur_tab = request.args.get("tab", _vis[0]["id"])
    rid = request.args.get("report", "")
    
    # Find the current tab in the filtered list
    tab = next((t for t in _vis if t["id"] == cur_tab), None)
    if not tab:
        return "ط؛ظٹط± ظ…طµط±ط­ ظ„ظƒ ط¨ط¹ط±ط¶ ظ‡ط°ط§ ط§ظ„طھط¨ظˆظٹط¨."
        
    # Find the requested report in the allowed reports of this tab
    rpt = next((r for r in tab["reports"] if r["id"] == rid), None)
    if not rpt:
        rpt = tab["reports"][0]
    
    cur_tab = tab["id"]
    for _p in rpt["params"]:
        if _p.get("dynamic") == "jv": _p["options"] = jv_options()
        if _p["name"] in ("rep_code","c_code","v_code","i_code","a_code","cc_code","grp_code"): _p["_list"] = lookups(_p["name"])
    display = {p["name"]: request.args.get(p["name"]) or (p["get_default"]() if "get_default" in p else p.get("default","")) for p in rpt["params"]}
    qsp = {"tab": cur_tab, "report": rpt["id"]}
    for p in rpt["params"]: qsp[p["name"]] = request.args.get(p["name"]) or (p["get_default"]() if "get_default" in p else p.get("default",""))
    qs = urlencode(qsp)
    error = None; cols=[]; rows=[]; dash=None
    if tab.get("dash"):
        try:
            dash = compute_dash(request.args.get("date_from","2026-01-01"), request.args.get("date_to","2026-12-31"))
        except Exception as e:
            error = str(e)
    else:
        try:
            cols, rows = run_report(rpt, request.args)
        except Exception as e:
            error = str(e)
    return render_template("index.html", tabs=_vis, tab=tab, cur_tab=cur_tab, rpt=rpt,
                                  binds=display, cols=cols, rows=rows, error=error, qs=qs, dash=dash, hidden_tabs=hidden_tabs, hidden_reports=hidden_reports, hide_profit=load_hide_profit())

PRINT_PAGE = """<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<style>
@page { size: A4 landscape; margin: 10mm; }
*{-webkit-print-color-adjust:exact;print-color-adjust:exact}
body{font-family:'Cairo','Inter',Tahoma,sans-serif;direction:rtl;color:#1e293b;margin:0;font-size:11px;}
.hd{display:flex;align-items:center;justify-content:space-between;border-bottom:2px solid #4f46e5;padding-bottom:10px;margin-bottom:15px}
.hd h1{font-size:18px;margin:0;color:#0f172a;font-weight:800}
.hd .dt{font-size:11px;color:#64748b;margin-top:6px;font-weight:600}
.logo{height:35px}
.filt{font-size:11px;color:#475569;margin-bottom:10px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:8px 10px;font-weight:600}
.filt b{color:#4f46e5}
table{border-collapse:collapse;width:100%;table-layout:fixed;word-wrap:break-word;}
thead th{background:#4f46e5;color:#fff;padding:4px 4px;font-size:11px;text-align:right;border:1px solid #4338ca;font-weight:700}
tbody td{padding:3px 4px;font-size:11px;border:1px solid #e2e8f0;text-align:right;font-weight:500;color:#1e293b}
tbody tr:nth-child(even) td{background:#f8fafc}
tbody tr:first-child td{background:#eef2ff;font-weight:800;color:#3730a3;border-bottom:2px solid #a5b4fc} /* طھظ…ظٹظٹط² طµظپ ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ */
.ft{margin-top:20px;font-size:11px;color:#94a3b8;text-align:center;border-top:1px solid #e2e8f0;padding-top:10px;font-weight:600}

</style></head>
<body onload="setTimeout(function(){window.print()},250)">
<div class="hd">
  <h1 style="color:#4f46e5;font-weight:900;margin:0;font-size:26px">طھظ‚ط§ط±ظٹط± ط§ظ„ط£ظˆظ†ظƒط³ ط§ظ„ط­ط¯ظٹط«ط©</h1>
  <div><h1>{{title}}</h1><div class="dt">طھط§ط±ظٹط® ط§ظ„ط·ط¨ط§ط¹ط©: {{now}}</div></div>
</div>
{% if filt %}<div class="filt">ط§ظ„ظپظ„ط§طھط± â€” {% for f in filt %}<b>{{f[0]}}</b>: {{f[1]}}{% if not loop.last %} &nbsp;|&nbsp; {% endif %}{% endfor %}</div>{% endif %}
{% set hidden_cols = ["ط§ظ„ط®طµظ… ظپظٹ ط§ظ„ظپط§طھظˆط±ط©", "ط¥ظٹط¯ط§ط¹ط§طھ ظˆطھط³ظˆظٹط§طھ (ط¨ط¯ظˆظ† ط¹ظ…ظٹظ„)"] %}
<table><thead><tr>{% for c in cols %}{% if c not in hidden_cols %}<th>{{c}}</th>{% endif %}{% endfor %}</tr></thead>
<tbody>{% for row in rows %}<tr>{% for cell in row %}{% if cols[loop.index0] not in hidden_cols %}<td>{{ '' if cell is none else cell }}</td>{% endif %}{% endfor %}</tr>{% endfor %}</tbody></table>
<div class="ft">ظ„ظˆط­ط© طھظ‚ط§ط±ظٹط± SREEN â€” ط¹ط¯ط¯ ط§ظ„طµظپظˆظپ: {{rows|length}}</div>
</body></html>"""

@app.route("/_old_export")
def export():
    username = session.get('username')
    tab_id = request.args.get("tab", TABS[0]["id"])
    report_id = request.args.get("report","")
    if not check_permission(username, tab_id, report_id):
        return "ط؛ظٹط± ظ…طµط±ط­ ظ„ظƒ ط¨طھطµط¯ظٹط± ظ‡ط°ط§ ط§ظ„طھظ‚ط±ظٹط±.", 403
    tab, rpt = find_report(tab_id, report_id)
    try:
        cols, rows = run_report(rpt, request.args)
    except Exception as e:
        return "ط®ط·ط£: " + str(e), 500
    try:
        hidden_cols = ["ط§ظ„ط®طµظ… ظپظٹ ط§ظ„ظپط§طھظˆط±ط©", "ط¥ظٹط¯ط§ط¹ط§طھ ظˆطھط³ظˆظٹط§طھ (ط¨ط¯ظˆظ† ط¹ظ…ظٹظ„)"]
        valid_indices = [i for i, col in enumerate(cols) if col not in hidden_cols]
        filtered_cols = [cols[i] for i in valid_indices]

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws = wb.active; ws.title = "طھظ‚ط±ظٹط±"
        ws.sheet_view.rightToLeft = True
        ws.append(filtered_cols)
        
        # طھظ†ط³ظٹظ‚ ط§ظ„ط±ط£ط³
        header_fill = PatternFill("solid", fgColor="4F46E5")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin', color='E2E8F0'), 
                        right=Side(style='thin', color='E2E8F0'), 
                        top=Side(style='thin', color='E2E8F0'), 
                        bottom=Side(style='thin', color='E2E8F0'))
                        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="right")
            cell.border = border
            
        # طھظ†ط³ظٹظ‚ طµظپ ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ (ط§ظ„طµظپ ط§ظ„ط£ظˆظ„ ظ…ظ† ط§ظ„ط¨ظٹط§ظ†ط§طھ ظˆظ‡ظˆ ط§ظ„طµظپ ط§ظ„ط«ط§ظ†ظٹ ظپظٹ ط§ظ„ط¥ظƒط³ظ„)
        total_fill = PatternFill("solid", fgColor="EEF2FF")
        total_font = Font(bold=True, color="3730A3")
        
        for row_idx, r in enumerate(rows, start=2):
            filtered_r = [str(r[i]) if r[i] is not None else '' for i in valid_indices]
            ws.append(filtered_r)
            for cell_idx, cell in enumerate(ws[row_idx], start=1):
                cell.border = border
                cell.alignment = Alignment(horizontal="right")
                if row_idx == 2:  # طµظپ ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ
                    cell.fill = total_fill
                    cell.font = total_font

        for i in range(1, len(filtered_cols)+1):
            ws.column_dimensions[get_column_letter(i)].width = 22
            
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=%s.xlsx" % rpt["id"]})
    except ImportError:
        import csv
        buf = io.StringIO()
        buf.write('\ufeff')
        writer = csv.writer(buf)
        writer.writerow(filtered_cols)
        for r in rows:
            filtered_r = [str(r[i]) if r[i] is not None else '' for i in valid_indices]
            writer.writerow(filtered_r)
        return Response(buf.getvalue().encode('utf-8'),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=%s.csv" % rpt["id"]})

@app.route("/_old_print")
def printview():
    username = session.get('username')
    tab_id = request.args.get("tab", TABS[0]["id"])
    report_id = request.args.get("report","")
    if not check_permission(username, tab_id, report_id):
        return "ط؛ظٹط± ظ…طµط±ط­ ظ„ظƒ ط¨ط·ط¨ط§ط¹ط© ظ‡ط°ط§ ط§ظ„طھظ‚ط±ظٹط±.", 403
    tab, rpt = find_report(tab_id, report_id)
    try:
        cols, rows = run_report(rpt, request.args)
        model = request.args.get("model", "1")
        if model == "2" and rpt["id"] == "collection_adopted":
            new_cols = ["ط§ظ„ط±ظ…ط²", "ط§ظ„ط§ط³ظ… / ط§ظ„ظˆطµظپ", "ط¥ط¬ظ…ط§ظ„ظٹ ط§ظ„ط³ظ†ط¯ط§طھ", "ظ‚ظٹظˆط¯ ط§ظ„ط´ط¨ظƒط© ط§ظ„ظ…ظ†ظپطµظ„ط©", "طµط§ظپظٹ ط§ظ„ظ…ط¨ظٹط¹ط§طھ ط§ظ„ظ†ظ‚ط¯ظٹط©", "ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ ط§ظ„ظ†ظ‡ط§ط¦ظٹ"]
            new_rows = []
            for r in rows:
                def parse_num(v):
                    if not v: return 0.0
                    if isinstance(v, str):
                        try: return float(v.replace(',',''))
                        except: return 0.0
                    return float(v)
                
                tot_rcpt = parse_num(r[2]) + parse_num(r[3]) + parse_num(r[4])
                net_jrn = parse_num(r[6])
                net_cash = parse_num(r[7]) - parse_num(r[10])
                final_tot = tot_rcpt + net_jrn + net_cash
                
                fmt = lambda x: f"{x:,.2f}" if x != 0 else "0.00"
                new_rows.append((r[0], r[1], fmt(tot_rcpt), fmt(net_jrn), fmt(net_cash), fmt(final_tot)))
            
            cols = new_cols
            rows = new_rows

    except Exception as e:
        return "ط®ط·ط£: " + str(e), 500
    filt = []
    for p in rpt["params"]:
        v = request.args.get(p["name"], p.get("default",""))
        if v not in ("", None): filt.append((p["label"], v))
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    title = rpt["title"] + (" (ظ†ظ…ظˆط°ط¬ 2)" if request.args.get("model") == "2" else "")
    return render_template("print.html", title=title, cols=cols, rows=rows, filt=filt, now=now)

SETTINGS_PAGE = """<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>ط§ظ„ط¥ط¹ط¯ط§ط¯ط§طھ</title>""" + STYLE + """</head><body>
<div class="app"><div class="main">
 <div class="wrap">
   <a class="back" href="/" style="color:#4f46e5;font-weight:700;display:inline-block;margin-bottom:10px">&#8594; ط±ط¬ظˆط¹ ظ„ظ„طھظ‚ط§ط±ظٹط±</a>
   {% if saved %}<div style="background:#e8f4ec;color:#1e7b34;padding:10px 14px;border-radius:8px;margin:6px 0 12px">طھظ… ط­ظپط¸ ط§ظ„ط¥ط¹ط¯ط§ط¯ط§طھ</div>{% endif %}
   <h1>ط¥ط¸ظ‡ط§ط± / ط¥ط®ظپط§ط، ط§ظ„طھط¨ظˆظٹط¨ط§طھ ظˆط§ظ„طھظ‚ط§ط±ظٹط±</h1>
   <p style="color:#6b7280;font-size:13px;margin-bottom:12px">ط¶ط¹ ط¹ظ„ط§ظ…ط© ط¹ظ„ظ‰ ظ…ط§ طھط±ظٹط¯ ط¥ط®ظپط§ط،ظ‡ ظ…ظ† ط§ظ„ظˆط§ط¬ظ‡ط©طŒ ط«ظ… ط§ط­ظپط¸.</p>
   <form method="post">
     <input type="hidden" name="action" value="save">
     <div class="card" style="margin-bottom:16px;border:2px solid #f59e0b;background:#fffbeb">
       <label style="font-weight:800;font-size:15px;color:#b45309"><input type="checkbox" name="hide_profit" {{ 'checked' if hide_profit else '' }}> ًں”’ ط¥ط®ظپط§ط، ظƒظ„ ظ…ط§ ظٹط®طµ ط§ظ„ط±ط¨ط­ ظ…ظ† ط§ظ„ظ†ط¸ط§ظ…</label>
       <div style="margin-top:6px;color:#92400e;font-size:12.5px">ط¹ظ†ط¯ ط§ظ„طھظپط¹ظٹظ„ ظٹظڈط®ظپظ‰: طھط¨ظˆظٹط¨ آ«ط§ظ„ط±ط¨ط­ظٹط©آ» ط¨ط§ظ„ظƒط§ظ…ظ„طŒ ط¨ط·ط§ظ‚طھط§ آ«ظ…ط¬ظ…ظ„ ط§ظ„ط±ط¨ط­آ» ظˆآ«طµط§ظپظٹ ط§ظ„ط±ط¨ط­آ» ظپظٹ ظ„ظˆط­ط© ط§ظ„ظ‚ظٹط§ط¯ط©طŒ ظˆطھظ‚ط±ظٹط±ط§ آ«ظ‚ط§ط¦ظ…ط© ط§ظ„ط¯ط®ظ„آ» ظˆآ«ظ…ط±ط§ظƒط² ط§ظ„طھظƒظ„ظپط©آ» ظپظٹ ط§ظ„طھط¨ظˆظٹط¨ ط§ظ„ظ…ط§ظ„ظٹ.</div>
     </div>
     {% for t in tabs %}
       <div class="card" style="margin-bottom:12px">
         <label style="font-weight:700;font-size:15px"><input type="checkbox" name="tab_{{t.id}}" {{ 'checked' if t.id in hidden_tabs else '' }}> ط¥ط®ظپط§ط، ط§ظ„طھط¨ظˆظٹط¨ ظƒط§ظ…ظ„ط§ظ‹: {{t.title}}</label>
         <div style="margin-top:10px;padding-right:18px;display:flex;flex-wrap:wrap;gap:14px">
           {% for r in t.reports %}
             <label style="font-size:13px;color:#374151"><input type="checkbox" name="rep_{{t.id}}/{{r.id}}" {{ 'checked' if (t.id ~ '/' ~ r.id) in hidden_reports else '' }}> {{r.title}}</label>
           {% endfor %}
         </div>
       </div>
     {% endfor %}
     <button type="submit" style="background:#4f46e5;color:#fff;border:0;padding:12px 24px;border-radius:9px;font-weight:700;font-size:15px;cursor:pointer">ط­ظپط¸ ط§ظ„ط¥ط¹ط¯ط§ط¯ط§طھ</button>
   </form>
 </div>
</div></div></body></html>"""

GLOBALS_PAGE = """<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>ط§ظ„ظ…طھط؛ظٹط±ط§طھ ط§ظ„ط¹ط§ظ…ط©</title>""" + STYLE + """
<style>
.tgt-table { width:100%; border-collapse:collapse; font-size:13px; }
.tgt-table th, .tgt-table td { border:1px solid #e2e8f0; padding:6px; text-align:center; }
.tgt-table th { background:#f8fafc; font-weight:700; color:#475569; position:sticky; top:0; }
.tgt-input { width:100%; min-width:70px; padding:4px; border:1px solid #cbd5e1; border-radius:4px; text-align:center; }
</style>
</head><body>
<div class="app"><div class="main">
 <div class="wrap">
   <a class="back" href="/" style="color:#4f46e5;font-weight:700;display:inline-block;margin-bottom:16px">&#8594; ط§ظ„ط¹ظˆط¯ط© ظ„ظ„ط±ط¦ظٹط³ظٹط©</a>
   <div class="rhead">
     <h1>ط§ظ„ظ…طھط؛ظٹط±ط§طھ ط§ظ„ط¹ط§ظ…ط© (ط§ظ„طھط§ط±ط¬طھ)</h1>
     {% if saved %}<div style="color:#10b981;font-weight:bold;margin-top:10px">طھظ… ط§ظ„ط­ظپط¸ ط¨ظ†ط¬ط§ط­!</div>{% endif %}
   </div>
   <form method="post" action="/globals">
     <div style="overflow-x:auto; max-height: 70vh; margin-bottom: 20px; border-radius: 8px; border: 1px solid #e2e8f0">
       <table class="tgt-table">
         <thead>
           <tr>
             <th style="min-width:150px">ط§ط³ظ… ط§ظ„ظ…ظ†ط¯ظˆط¨</th>
             {% for m in range(1, 13) %}
             <th>ط´ظ‡ط± {{m}}</th>
             {% endfor %}
           </tr>
         </thead>
         <tbody>
           {% for rep in reps %}
           <tr>
             <td style="text-align:right">{{ rep.name }}</td>
             {% for m in range(1, 13) %}
             <td>
               <input type="number" class="tgt-input" 
                      name="target_{{rep.code}}_{{m}}" 
                      value="{{ targets.get(rep.code|string, {}).get(m|string, 1000000) }}">
             </td>
             {% endfor %}
           </tr>
           {% endfor %}
         </tbody>
       </table>
     </div>
     <button type="submit" style="background:#4f46e5;color:#fff;border:0;padding:12px 24px;border-radius:9px;font-weight:700;font-size:15px;cursor:pointer">ط­ظپط¸ ط§ظ„ظ…طھط؛ظٹط±ط§طھ</button>
   </form>
 </div>
</div></div></body></html>"""


PIN_PAGE = """<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>ط±ظ…ط² ط§ظ„ط¯ط®ظˆظ„</title>""" + STYLE + """</head><body>
<div class="app"><div class="main">
 <div class="wrap">
   <a class="back" href="/" style="color:#4f46e5;font-weight:700;display:inline-block;margin-bottom:16px">&#8594; ط±ط¬ظˆط¹ ظ„ظ„طھظ‚ط§ط±ظٹط±</a>
   <div class="card" style="max-width:380px;margin:40px auto;text-align:center">
     <div style="font-size:40px;margin-bottom:6px">ًں”’</div>
     <h1 style="font-size:18px;margin:0 0 4px">طھط¨ظˆظٹط¨ ط§ظ„ط¥ط¹ط¯ط§ط¯ط§طھ ظ…ط­ظ…ظٹ</h1>
     <p style="color:#6b7280;font-size:13px;margin:0 0 16px">ط£ط¯ط®ظ„ ط±ظ…ط² ط§ظ„ط¯ط®ظˆظ„ ظ„ظ„ظ…طھط§ط¨ط¹ط©</p>
     {% if error %}<div class="err" style="margin-bottom:12px">ط±ظ…ط² ط§ظ„ط¯ط®ظˆظ„ ط؛ظٹط± طµط­ظٹط­</div>{% endif %}
     <form method="post">
       <input type="password" name="pin" autofocus inputmode="numeric" placeholder="â€¢ â€¢ â€¢ â€¢ â€¢"
              style="width:100%;text-align:center;letter-spacing:8px;font-size:22px;padding:12px;border:1.5px solid #cbd5e1;border-radius:10px;margin-bottom:14px">
       <button type="submit" style="width:100%;background:#4f46e5;color:#fff;border:0;padding:12px;border-radius:10px;font-weight:700;font-size:15px;cursor:pointer">ط¯ط®ظˆظ„</button>
     </form>
   </div>
 </div>
</div></div></body></html>"""

@app.route("/_old_settings_logout")
def settings_logout():
    session.pop("set_ok", None)
    return render_template("pin.html", error=False)

@app.route("/_old_settings", methods=["GET","POST"])
def settings():
    if session.get("role") != "admin": return "ظ„ط§ طھظ…ظ„ظƒ طµظ„ط§ط­ظٹط©", 403
    # ط¨ظˆط§ط¨ط© ط±ظ…ط² ط§ظ„ط¯ط®ظˆظ„

    saved = False
    if request.method == "POST" and request.form.get("action") == "save":
        htabs = [t["id"] for t in TABS if request.form.get("tab_"+t["id"])]
        hreps = []
        for t in TABS:
            for r in t["reports"]:
                key = t["id"]+"/"+r["id"]
                if request.form.get("rep_"+key):
                    hreps.append(key)
        save_hidden(htabs, hreps, bool(request.form.get("hide_profit")))
        saved = True
    ht, hr = load_hidden_raw()
    return render_template("settings.html", tabs=_vis, hidden_tabs=ht, hidden_reports=hr,
                                  saved=saved, hide_profit=load_hide_profit())




@app.route("/_old_globals", methods=["GET","POST"])
def globals_page():
    if session.get("role") != "admin": return "ظ„ط§ طھظ…ظ„ظƒ طµظ„ط§ط­ظٹط©", 403

            
    saved = False
    targets_data = load_globals()
    if not targets_data: targets_data = {}
    
    if request.method == "POST":
        hide_profit = "hide_profit" in request.form
        save_hidden([], [], hide_profit=hide_profit)
        for key in request.form:
            if key.startswith("target_2026_"):
                parts = key.split("_")
                if len(parts) == 4:
                    year = parts[1]
                    rep_code = parts[2]
                    month = parts[3]
                    try:
                        val = float(request.form[key])
                        if year not in targets_data: targets_data[year] = {}
                        if rep_code not in targets_data[year]: targets_data[year][rep_code] = {}
                        targets_data[year][rep_code][month] = val
                    except ValueError:
                        pass
        save_globals(targets_data)
        saved = True
        
    reps = []
    try:
        with get_conn() as con:
            with con.cursor() as cur:
                cur.execute("SELECT REPRS_CODE, REPRS_A_NAME FROM IAS20261.SALES_MAN ORDER BY REPRS_CODE")
                for c, n in cur.fetchall():
                    reps.append({"code": str(c), "name": n or str(c)})
    except Exception as e:
        print("Error loading reps:", e)
        
    return render_template("globals.html", reps=reps, targets=targets_data.get("2026", {}), saved=saved, hide_profit=load_hide_profit())


DASHBOARD_PAGE = '''<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>ظ„ظˆط­ط© ط§ظ„ظ‚ظٹط§ط¯ط© SREEN</title>
     <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
     <script>
     var D={{ dash|tojson }};
     window.addEventListener("load",function(){ 
       if(!window.Chart) return; 
       Chart.defaults.font.family = "'Cairo', 'Inter', sans-serif";
       Chart.defaults.color = "#64748b";
       
       const commonOptions = {
         responsive: true,
         maintainAspectRatio: false,
         plugins: {
           legend: { display: false },
           tooltip: { backgroundColor: '#1e293b', padding: 14, titleFont: { size: 14, family: "'Cairo', sans-serif", weight: 'bold' }, bodyFont: { size: 14, family: "'Cairo', sans-serif" }, cornerRadius: 10, displayColors: true, boxPadding: 6 }
         }
       };
       
       // C1: Bar Chart (Sales & Collection)
       new Chart(document.getElementById("c1"),{
         type:"bar",
         data:{
           labels:D.months,
           datasets:[
             {label:"ظ…ط¨ظٹط¹ط§طھ", data:D.msales, backgroundColor:"#4f46e5", borderRadius:8, maxBarThickness: 32},
             {label:"طھط­طµظٹظ„", data:D.mcollect, backgroundColor:"#38bdf8", borderRadius:8, maxBarThickness: 32}
           ]
         },
         options: {
           ...commonOptions,
           plugins: { ...commonOptions.plugins, legend: { display: true, position: 'top', align: 'end', labels: { usePointStyle: true, boxWidth: 10, font: { family: "'Cairo'", size: 13, weight: 'bold' } } } },
           scales: {
             x: { grid: { display: false }, border: { display: false } },
             y: { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false } }
           }
         }
       });

       // C2: Doughnut (Salesmen)
       new Chart(document.getElementById("c2"),{
         type:"doughnut",
         data:{
           labels:D.rep_labels.slice(0,5),
           datasets:[{data:D.rep_vals.slice(0,5), backgroundColor:["#4f46e5", "#38bdf8", "#10b981", "#f59e0b", "#8b5cf6"], borderWidth: 0, hoverOffset: 4}]
         },
         options: {
           responsive: true, maintainAspectRatio: false, cutout: '75%',
           plugins: { legend: { display: true, position: 'bottom', labels: { usePointStyle: true, boxWidth: 8, font: { size: 11 } } } }
         }
       });

       // C3: Doughnut (Items)
       new Chart(document.getElementById("c3"),{
         type:"doughnut",
         data:{
           labels:D.itm_labels.slice(0,5),
           datasets:[{data:D.itm_vals.slice(0,5), backgroundColor:["#f43f5e", "#d946ef", "#0ea5e9", "#14b8a6", "#eab308"], borderWidth: 0, hoverOffset: 4}]
         },
         options: {
           responsive: true, maintainAspectRatio: false, cutout: '75%',
           plugins: { legend: { display: true, position: 'bottom', labels: { usePointStyle: true, boxWidth: 8, font: { size: 11 } } } }
         }
       });

       // C4: Line Chart (Purchases)
       const ctx4 = document.getElementById("c4").getContext('2d');
       const grad4 = ctx4.createLinearGradient(0, 0, 0, 300);
       grad4.addColorStop(0, 'rgba(16, 185, 129, 0.4)');
       grad4.addColorStop(1, 'rgba(16, 185, 129, 0.0)');
       
       new Chart(ctx4,{
         type:"line",
         data:{
           labels:D.months,
           datasets:[{
             label: "ظ…ط´طھط±ظٹط§طھ", data:D.mpurch, borderColor:"#10b981", borderWidth: 3, backgroundColor: grad4, fill:true, tension:0.4, pointRadius: 0, pointHoverRadius: 6, pointBackgroundColor: "#fff", pointBorderColor: "#10b981", pointBorderWidth: 2
           }]
         },
         options: {
           ...commonOptions,
           interaction: { mode: 'index', intersect: false },
           scales: {
             x: { grid: { display: false }, border: { display: false } },
             y: { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false } }
           }
         }
       });
     });
     </script>
<style>
 *{box-sizing:border-box;font-family:Tahoma,Arial}
 body{margin:0;background:#f1f5f9;color:#0f172a}
 .hd{background:linear-gradient(90deg,#0f766e,#134e4a);color:#fff;padding:14px 24px;display:flex;align-items:center;gap:14px;flex-wrap:wrap}
 .hd h1{margin:0;font-size:19px} .hd .sp{flex:1}
 .hd a{color:#fff;text-decoration:none;background:rgba(255,255,255,.15);padding:8px 14px;border-radius:8px;font-size:14px}
 .hd form{display:flex;gap:8px;align-items:center;font-size:13px}
 .hd input{padding:7px;border:0;border-radius:6px} .hd button{padding:8px 14px;border:0;border-radius:6px;background:#fbbf24;font-weight:700;cursor:pointer}
 .wrap{padding:22px;max-width:1300px;margin:auto}
 .kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:22px}
 @media(max-width:900px){.kpis{grid-template-columns:repeat(2,1fr)}}
 .kpi{background:#fff;border-radius:14px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.08);border-right:4px solid #0f766e}
 .kpi .l{color:#64748b;font-size:13px;margin-bottom:6px} .kpi .v{font-size:21px;font-weight:800}
 .kpi.g{border-color:#16a34a}.kpi.b{border-color:#2563eb}.kpi.o{border-color:#ea580c}.kpi.r{border-color:#dc2626}.kpi.p{border-color:#7c3aed}
 .charts{display:grid;grid-template-columns:1fr 1fr;gap:14px}
 @media(max-width:900px){.charts{grid-template-columns:1fr}}
 .ch{background:#fff;border-radius:14px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
 .ch h3{margin:0 0 12px;font-size:15px}

</style></head><body>
 <div class="hd"><h1>ًں“ٹ ظ„ظˆط­ط© ط§ظ„ظ‚ظٹط§ط¯ط© â€” SREEN</h1>
   <form method="get" action="/dashboard"><span>ظ…ظ†</span><input type="date" name="date_from" value="{{f}}"><span>ط¥ظ„ظ‰</span><input type="date" name="date_to" value="{{t}}"><button type="submit">طھط­ط¯ظٹط«</button></form>
   <div class="sp"></div><a href="/">â†گ ط§ظ„طھظ‚ط§ط±ظٹط±</a></div>
 <div class="wrap">
   <div class="kpis">
     <div class="kpi b"><div class="l">ط¥ط¬ظ…ط§ظ„ظٹ ط§ظ„ظ…ط¨ظٹط¹ط§طھ</div><div class="v">{{ "{:,.0f}".format(data.sales) }}</div></div>
     <div class="kpi g"><div class="l">ط¥ط¬ظ…ط§ظ„ظٹ ط§ظ„طھط­طµظٹظ„</div><div class="v">{{ "{:,.0f}".format(data.collect) }}</div></div>
     <div class="kpi o"><div class="l">ط¥ط¬ظ…ط§ظ„ظٹ ط§ظ„ظ…ط´طھط±ظٹط§طھ</div><div class="v">{{ "{:,.0f}".format(data.purch) }}</div></div>
     {% if not hide_profit|default(false) %}<div class="kpi p"><div class="l">ظ…ط¬ظ…ظ„ ط§ظ„ط±ط¨ط­</div><div class="v">{{ "{:,.0f}".format(data.gross) }}</div></div>
     <div class="kpi g"><div class="l">طµط§ظپظٹ ط§ظ„ط±ط¨ط­</div><div class="v">{{ "{:,.0f}".format(data.netprofit) }}</div></div>{% endif %}
     <div class="kpi r"><div class="l">ط§ظ„ط°ظ…ظ… ط§ظ„ظ…ط¯ظٹظ†ط©</div><div class="v">{{ "{:,.0f}".format(data.recv) }}</div></div>
     <div class="kpi b"><div class="l">ظ‚ظٹظ…ط© ط§ظ„ظ…ط®ط²ظˆظ†</div><div class="v">{{ "{:,.0f}".format(data.invval) }}</div></div>
     <div class="kpi o"><div class="l">طµط§ظپظٹ ط§ظ„ط¶ط±ظٹط¨ط©</div><div class="v">{{ "{:,.0f}".format(data.vat) }}</div></div>
   </div>
   <div class="charts">
     <div class="ch"><h3>ط§ظ„ظ…ط¨ظٹط¹ط§طھ ظˆط§ظ„طھط­طµظٹظ„ ط´ظ‡ط±ظٹط§ظ‹</h3><canvas id="c1" height="140"></canvas></div>
     <div class="ch"><h3>ط£ظپط¶ظ„ ط§ظ„ظ…ظ†ط§ط¯ظٹط¨ (ظ…ط¨ظٹط¹ط§طھ)</h3><canvas id="c2" height="140"></canvas></div>
     <div class="ch"><h3>ط£ظپط¶ظ„ ط§ظ„ط£طµظ†ط§ظپ (ظ…ط¨ظٹط¹ط§طھ)</h3><canvas id="c3" height="140"></canvas></div>
     <div class="ch"><h3>ط§ظ„ظ…ط´طھط±ظٹط§طھ ط´ظ‡ط±ظٹط§ظ‹</h3><canvas id="c4" height="140"></canvas></div>
   </div></div>
<script>
const D={{ data|tojson }};
Chart.defaults.font.family="Tahoma";
new Chart(c1,{type:"bar",data:{labels:D.months,datasets:[{label:"ظ…ط¨ظٹط¹ط§طھ",data:D.msales,backgroundColor:"#2563eb"},{label:"طھط­طµظٹظ„",data:D.mcollect,backgroundColor:"#16a34a"}]}});
new Chart(c2,{type:"bar",data:{labels:D.rep_labels,datasets:[{label:"ظ…ط¨ظٹط¹ط§طھ",data:D.rep_vals,backgroundColor:"#0f766e"}]},options:{indexAxis:"y",plugins:{legend:{display:false}}}});
new Chart(c3,{type:"bar",data:{labels:D.itm_labels,datasets:[{label:"ظ…ط¨ظٹط¹ط§طھ",data:D.itm_vals,backgroundColor:"#ea580c"}]},options:{indexAxis:"y",plugins:{legend:{display:false}}}});
new Chart(c4,{type:"line",data:{labels:D.months,datasets:[{label:"ظ…ط´طھط±ظٹط§طھ",data:D.mpurch,borderColor:"#ea580c",backgroundColor:"rgba(234,88,12,.12)",fill:true,tension:.3}]},options:{plugins:{legend:{display:false}}}});
</script></body></html>'''

def compute_dash(f, t):
    import concurrent.futures
    from database import get_pooled_conn
    b = {"f": f, "t": t}
    P="TO_DATE(:f,'YYYY-MM-DD')"; Q="TO_DATE(:t,'YYYY-MM-DD')+1"
    d = {"sales":0,"collect":0,"purch":0,"gross":0,"netprofit":0,"recv":0,"invval":0,"vat":0,
         "months":[],"msales":[],"mcollect":[],"mpurch":[],"rep_labels":[],"rep_vals":[],"itm_labels":[],"itm_vals":[]}
    try:
        sqls = {
            "sales": "SELECT NVL(SUM((NVL(BILL_AMT,0)-(NVL(DISC_AMT,0)-NVL(ADD_DISC_AMT_MST,0))+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)) * DECODE(BILL_DOC_TYPE, 3, -1, 1)),0) FROM IAS20261.IAS_BILL_MST WHERE BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND BILL_DATE>="+P+" AND BILL_DATE<"+Q,
            "sales_ret": "SELECT NVL(SUM((NVL(BILL_AMT,0)-(NVL(DISC_AMT,0)-NVL(ADD_DISC_AMT_MST,0))+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)) * DECODE(RT_BILL_DOC_TYPE, 3, -1, 1)),0) FROM IAS20261.IAS_RT_BILL_MST WHERE RT_BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND RT_BILL_DATE>="+P+" AND RT_BILL_DATE<"+Q,
            "collect": "SELECT NVL(SUM(NVL(CR_AMT,0)),0) FROM IAS20261.IAS_POST_DTL WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=2 AND C_CODE IS NOT NULL AND DOC_DATE>="+P+" AND DOC_DATE<"+Q,
            "purch": "SELECT NVL(SUM((NVL(BILL_AMT,0)-(NVL(DISC_AMT,0)-NVL(ADD_DISC_AMT_MST,0))+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)) * DECODE(BILL_DOC_TYPE, 3, -1, 1)),0) FROM IAS20261.IAS_PI_BILL_MST WHERE BILL_DATE>="+P+" AND BILL_DATE<"+Q,
            "gross": "SELECT NVL(SUM((NVL(x.I_QTY,0)*(NVL(x.I_PRICE,0)-NVL(x.DIS_AMT,0)+NVL(x.OTHR_AMT,0))-NVL(x.I_QTY,0)*NVL(x.STK_COST,0)) * DECODE(m.BILL_DOC_TYPE, 3, -1, 1)),0) FROM IAS20261.IAS_BILL_DTL x JOIN IAS20261.IAS_BILL_MST m ON m.BILL_DOC_TYPE=x.BILL_DOC_TYPE AND m.BILL_NO=x.BILL_NO AND m.BILL_SER=x.BILL_SER WHERE m.BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND m.BILL_DATE>="+P+" AND m.BILL_DATE<"+Q,
            "gross_ret": "SELECT NVL(SUM((NVL(x.I_QTY,0)*(NVL(x.I_PRICE,0)-NVL(x.DIS_AMT,0)+NVL(x.OTHR_AMT,0))-NVL(x.I_QTY,0)*NVL(x.STK_COST,0)) * DECODE(m.RT_BILL_DOC_TYPE, 3, -1, 1)),0) FROM IAS20261.IAS_RT_BILL_DTL x JOIN IAS20261.IAS_RT_BILL_MST m ON m.RT_BILL_DOC_TYPE=x.RT_BILL_DOC_TYPE AND m.RT_BILL_NO=x.RT_BILL_NO AND m.RT_BILL_SER=x.RT_BILL_SER WHERE m.RT_BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND m.RT_BILL_DATE>="+P+" AND m.RT_BILL_DATE<"+Q,
            "netprofit": "SELECT NVL(SUM(NVL(p.CR_AMT,0)-NVL(p.DR_AMT,0)),0) FROM IAS20261.IAS_POST_DTL p JOIN IAS20261.ACCOUNT a ON a.A_CODE=p.A_CODE WHERE NVL(p.DOC_POST,0)=1 AND a.A_REPORT=2 AND p.DOC_DATE>="+P+" AND p.DOC_DATE<"+Q,
            "recv": "SELECT NVL(SUM(bal),0) FROM (SELECT SUM(NVL(DR_AMT,0)-NVL(CR_AMT,0)) bal FROM IAS20261.IAS_POST_DTL WHERE NVL(DOC_POST,0)=1 AND C_CODE IS NOT NULL AND DOC_DATE<"+Q+" GROUP BY C_CODE HAVING SUM(NVL(DR_AMT,0)-NVL(CR_AMT,0))>0)",
            "invval": "SELECT NVL(SUM(NVL(I_QTY,0)*NVL(IN_OUT,0)*NVL(STK_COST,0)),0) FROM IAS20261.ITEM_MOVEMENT WHERE I_DATE<"+Q,
            "ov": "SELECT NVL(SUM(NVL(VAT_AMT,0) * DECODE(BILL_DOC_TYPE, 3, -1, 1)),0) FROM IAS20261.IAS_BILL_MST WHERE BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND BILL_DATE>="+P+" AND BILL_DATE<"+Q,
            "ov_ret": "SELECT NVL(SUM(NVL(VAT_AMT,0) * DECODE(RT_BILL_DOC_TYPE, 3, -1, 1)),0) FROM IAS20261.IAS_RT_BILL_MST WHERE RT_BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND RT_BILL_DATE>="+P+" AND RT_BILL_DATE<"+Q,
            "iv": "SELECT NVL(SUM(NVL(VAT_AMT,0) * DECODE(BILL_DOC_TYPE, 3, -1, 1)),0) FROM IAS20261.IAS_PI_BILL_MST WHERE BILL_DATE>="+P+" AND BILL_DATE<"+Q,
            "ms": "SELECT TO_CHAR(BILL_DATE,'YYYY-MM'), SUM((NVL(BILL_AMT,0)-(NVL(DISC_AMT,0)-NVL(ADD_DISC_AMT_MST,0))+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)) * DECODE(BILL_DOC_TYPE, 3, -1, 1)) FROM IAS20261.IAS_BILL_MST WHERE BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND BILL_DATE>="+P+" AND BILL_DATE<"+Q+" GROUP BY TO_CHAR(BILL_DATE,'YYYY-MM')",
            "ms_ret": "SELECT TO_CHAR(RT_BILL_DATE,'YYYY-MM'), SUM((NVL(BILL_AMT,0)-(NVL(DISC_AMT,0)-NVL(ADD_DISC_AMT_MST,0))+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)) * DECODE(RT_BILL_DOC_TYPE, 3, -1, 1)) FROM IAS20261.IAS_RT_BILL_MST WHERE RT_BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND RT_BILL_DATE>="+P+" AND RT_BILL_DATE<"+Q+" GROUP BY TO_CHAR(RT_BILL_DATE,'YYYY-MM')",
            "mc": "SELECT TO_CHAR(DOC_DATE,'YYYY-MM'), SUM(CR_AMT) FROM (SELECT DOC_DATE, CR_AMT FROM IAS20261.IAS_POST_DTL WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=2 AND C_CODE IS NOT NULL AND DOC_DATE>="+P+" AND DOC_DATE<"+Q+" UNION ALL SELECT DOC_DATE, CR_AMT FROM IAS20261.IAS_POST_DTL WHERE NVL(DOC_POST,0)=0 AND DOC_TYPE=2 AND DOC_DATE>="+P+" AND DOC_DATE<"+Q+" UNION ALL SELECT DOC_DATE, CR_AMT FROM IAS20261.IAS_POST_DTL WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=1 AND JV_TYPE=2 AND C_CODE IS NOT NULL AND DOC_DATE>="+P+" AND DOC_DATE<"+Q+" UNION ALL SELECT b.BILL_DATE AS DOC_DATE, NVL(p.DR_AMT,0) AS CR_AMT FROM IAS20261.IAS_BILL_MST b JOIN IAS20261.IAS_POST_DTL p ON p.DOC_NO = b.BILL_NO AND p.DOC_SER = b.BILL_SER AND p.DOC_TYPE = 4 AND TO_CHAR(p.A_CODE) LIKE '111%' WHERE b.BILL_DOC_TYPE=1 AND NVL(p.DOC_POST,0)=1 AND p.DR_AMT > 0 AND b.BILL_DATE>="+P+" AND b.BILL_DATE<"+Q+" UNION ALL SELECT DOC_DATE, -CR_AMT FROM IAS20261.IAS_POST_DTL WHERE NVL(DOC_POST,0)=1 AND DOC_TYPE=5 AND TO_CHAR(A_CODE) LIKE '111%' AND DOC_DATE>="+P+" AND DOC_DATE<"+Q+") GROUP BY TO_CHAR(DOC_DATE,'YYYY-MM')",
            "mp": "SELECT TO_CHAR(BILL_DATE,'YYYY-MM'), SUM((NVL(BILL_AMT,0)-(NVL(DISC_AMT,0)-NVL(ADD_DISC_AMT_MST,0))+NVL(VAT_AMT,0)+NVL(OTHR_AMT,0)) * DECODE(BILL_DOC_TYPE, 3, -1, 1)) FROM IAS20261.IAS_PI_BILL_MST WHERE BILL_DATE>="+P+" AND BILL_DATE<"+Q+" GROUP BY TO_CHAR(BILL_DATE,'YYYY-MM')",
            "rs": "SELECT NVL(sm.REPRS_A_NAME, m.REP_CODE), SUM((NVL(m.BILL_AMT,0)-(NVL(m.DISC_AMT,0)-NVL(m.ADD_DISC_AMT_MST,0))+NVL(m.VAT_AMT,0)+NVL(m.OTHR_AMT,0)) * DECODE(m.BILL_DOC_TYPE, 3, -1, 1)) FROM IAS20261.IAS_BILL_MST m LEFT JOIN IAS20261.SALES_MAN sm ON sm.REPRS_CODE=m.REP_CODE WHERE m.BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND m.BILL_DATE>="+P+" AND m.BILL_DATE<"+Q+" GROUP BY NVL(sm.REPRS_A_NAME,m.REP_CODE)",
            "rs_ret": "SELECT NVL(sm.REPRS_A_NAME, m.REP_CODE_BILL), SUM((NVL(m.BILL_AMT,0)-(NVL(m.DISC_AMT,0)-NVL(m.ADD_DISC_AMT_MST,0))+NVL(m.VAT_AMT,0)+NVL(m.OTHR_AMT,0)) * DECODE(m.RT_BILL_DOC_TYPE, 3, -1, 1)) FROM IAS20261.IAS_RT_BILL_MST m LEFT JOIN IAS20261.SALES_MAN sm ON sm.REPRS_CODE=m.REP_CODE_BILL WHERE m.RT_BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND m.RT_BILL_DATE>="+P+" AND m.RT_BILL_DATE<"+Q+" GROUP BY NVL(sm.REPRS_A_NAME,m.REP_CODE_BILL)",
            "its": "SELECT NVL(i.I_NAME, x.I_CODE), SUM((NVL(x.I_QTY,0)*(NVL(x.I_PRICE,0)-NVL(x.DIS_AMT,0)+NVL(x.OTHR_AMT,0))) * DECODE(m.BILL_DOC_TYPE, 3, -1, 1)) FROM IAS20261.IAS_BILL_DTL x JOIN IAS20261.IAS_BILL_MST m ON m.BILL_DOC_TYPE=x.BILL_DOC_TYPE AND m.BILL_NO=x.BILL_NO AND m.BILL_SER=x.BILL_SER LEFT JOIN IAS20261.IAS_ITM_MST i ON i.I_CODE=x.I_CODE WHERE m.BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND m.BILL_DATE>="+P+" AND m.BILL_DATE<"+Q+" GROUP BY NVL(i.I_NAME,x.I_CODE)",
            "its_ret": "SELECT NVL(i.I_NAME, x.I_CODE), SUM((NVL(x.I_QTY,0)*(NVL(x.I_PRICE,0)-NVL(x.DIS_AMT,0)+NVL(x.OTHR_AMT,0))) * DECODE(m.RT_BILL_DOC_TYPE, 3, -1, 1)) FROM IAS20261.IAS_RT_BILL_DTL x JOIN IAS20261.IAS_RT_BILL_MST m ON m.RT_BILL_DOC_TYPE=x.RT_BILL_DOC_TYPE AND m.RT_BILL_NO=x.RT_BILL_NO AND m.RT_BILL_SER=x.RT_BILL_SER LEFT JOIN IAS20261.IAS_ITM_MST i ON i.I_CODE=x.I_CODE WHERE m.RT_BILL_DOC_TYPE IN (1,2,3,4,5,6,7,8) AND m.RT_BILL_DATE>="+P+" AND m.RT_BILL_DATE<"+Q+" GROUP BY NVL(i.I_NAME,x.I_CODE)"
        }

        results = {}
        with get_pooled_conn() as con:
            with con.cursor() as cur:
                for key, sql in sqls.items():
                    try:
                        cur.execute(sql, {k:v for k,v in b.items() if (":"+k) in sql})
                        if key in ["sales", "sales_ret", "collect", "purch", "gross", "gross_ret", "netprofit", "recv", "invval", "ov", "ov_ret", "iv"]:
                            r = cur.fetchone()
                            results[key] = round(float(r[0]),2) if r and r[0] is not None else 0.0
                        else:
                            m = {}
                            for r in cur.fetchall():
                                m[str(r[0])] = round(float(r[1] or 0),2)
                            results[key] = m
                    except Exception as e:
                        print(f"Error in {key}: {e}")
                        results[key] = (0.0 if key not in ["ms", "ms_ret", "mc", "mp", "rs", "rs_ret", "its", "its_ret"] else {})

        sales = results["sales"]; sales_ret = results["sales_ret"]
        d["sales"] = round(sales - sales_ret, 2)
        d["collect"] = results["collect"]
        d["purch"] = results["purch"]
        d["gross"] = round(results["gross"] - results["gross_ret"], 2)
        d["netprofit"] = results["netprofit"]
        d["recv"] = results["recv"]
        d["invval"] = results["invval"]
        d["vat"] = round((results["ov"] - results["ov_ret"]) - results["iv"], 2)
        
        ms = results["ms"]; ms_ret = results["ms_ret"]
        mc = results["mc"]; mp = results["mp"]
        
        months=sorted(set(list(ms)+list(ms_ret)+list(mc)+list(mp)))
        d["months"]=months
        d["msales"]=[round(ms.get(x,0) - ms_ret.get(x,0), 2) for x in months]
        d["mcollect"]=[mc.get(x,0) for x in months]
        d["mpurch"]=[mp.get(x,0) for x in months]
        
        rs = results["rs"]; rs_ret = results["rs_ret"]
        rs_net = {k: round(rs.get(k,0) - rs_ret.get(k,0), 2) for k in set(list(rs)+list(rs_ret))}
        for k, v in sorted(rs_net.items(), key=lambda item: item[1], reverse=True):
            if v != 0:
                d["rep_labels"].append(str(k))
                d["rep_vals"].append(v)
        
        its = results["its"]; its_ret = results["its_ret"]
        its_net = {k: round(its.get(k,0) - its_ret.get(k,0), 2) for k in set(list(its)+list(its_ret))}
        for k, v in sorted(its_net.items(), key=lambda item: item[1], reverse=True)[:50]:
            if v != 0:
                d["itm_labels"].append(str(k)[:22])
                d["itm_vals"].append(v)
    except Exception as e:
        d["err"]=str(e)
    return d




@app.route("/_old_users", methods=["GET", "POST"])
def users_manage():
    if session.get("role") != "admin":
        return "ط؛ظٹط± ظ…طµط±ط­ ظ„ظƒ ط¨ط¯ط®ظˆظ„ ظ‡ط°ظ‡ ط§ظ„طµظپط­ط©.", 403
        
    from config import load_users, save_users
    users = load_users()
    
    if request.method == "POST":
        action = request.form.get("action")
        username = request.form.get("username", "").strip()
        
        if action == "add_or_update" and username:
            password = request.form.get("password")
            role = request.form.get("role", "user")
            
            allowed_tabs = request.form.getlist("allowed_tabs")
            allowed_reports = request.form.getlist("allowed_reports")
            
            if role == "admin":
                allowed_tabs = ["*"]
                allowed_reports = ["*"]
                
            users[username] = {
                "password": password if password else users.get(username, {}).get("password", ""),
                "role": role,
                "allowed_tabs": allowed_tabs,
                "allowed_reports": allowed_reports
            }
            save_users(users)
            
        elif action == "delete" and username:
            if username in users and username != "admin":
                del users[username]
                save_users(users)
                
        return redirect("/users")
        
    return render_template("users_manage.html", users=users, TABS=TABS)

@app.route("/api/users_manage", methods=["GET", "POST"])
def api_users_manage():
    from config import load_users, save_users
    
    if request.method == "GET":
        # Return all users and TABS structure for UI
        users = load_users()
        safe_users = {}
        for uname, udata in users.items():
            safe_users[uname] = {
                "role": udata.get("role", "user"),
                "allowed_tabs": udata.get("allowed_tabs", []),
                "allowed_reports": udata.get("allowed_reports", [])
                # Not sending passwords for security
            }
        
        # Prepare tabs data
        tabs_data = []
        for t in TABS:
            reports_data = [{"id": r["id"], "title": r["title"]} for r in t["reports"]]
            tabs_data.append({"id": t["id"], "title": t["title"], "reports": reports_data})
            
        return jsonify({"users": safe_users, "tabs": tabs_data})

    elif request.method == "POST":
        if session.get("role") != "admin":
            return jsonify({"error": "Unauthorized"}), 403
            
        data = request.json
        if not data:
            return jsonify({"error": "No data provided"}), 400
            
        action = data.get("action")
        username = data.get("username", "").strip()
        users = load_users()
        
        if action == "add_or_update" and username:
            password = data.get("password")
            role = data.get("role", "user")
            
            allowed_tabs = data.get("allowed_tabs", [])
            allowed_reports = data.get("allowed_reports", [])
            
            if role == "admin":
                allowed_tabs = ["*"]
                allowed_reports = ["*"]
                
            users[username] = {
                "password": password if password else users.get(username, {}).get("password", ""),
                "role": role,
                "allowed_tabs": allowed_tabs,
                "allowed_reports": allowed_reports
            }
            save_users(users)
            return jsonify({"success": True})
            
        elif action == "delete" and username:
            if username in users and username != "admin":
                del users[username]
                save_users(users)
                return jsonify({"success": True})
            return jsonify({"error": "Cannot delete this user"}), 400
            
        return jsonify({"error": "Invalid action"}), 400

@app.route("/api/settings_manage", methods=["GET", "POST"])
def api_settings_manage():
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
        
    if request.method == "GET":
        from database import get_pooled_conn as get_db_connection
        reps = []
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT REPRS_CODE, REPRS_A_NAME FROM IAS20261.SALES_MAN")
                    for c, n in cur.fetchall():
                        reps.append({"code": str(c), "name": n or str(c)})
        except Exception as e:
            print("Error loading reps:", e)
            
        targets_data = load_globals()
        ht, hr = load_hidden_raw()
        
        return jsonify({
            "reps": reps,
            "targets": targets_data.get("2026", {}),
            "hide_profit": load_hide_profit(),
            "hidden_tabs": list(ht),
            "hidden_reports": list(hr)
        })

    elif request.method == "POST":
        data = request.json
        if not data:
            return jsonify({"error": "No data provided"}), 400
            
        # Extract fields
        hide_profit = bool(data.get("hide_profit", False))
        hidden_tabs = data.get("hidden_tabs", [])
        hidden_reports = data.get("hidden_reports", [])
        targets = data.get("targets", {}) # expecting dict of {rep_code: {month: val}}
        
        # Save hidden configs
        save_hidden(hidden_tabs, hidden_reports, hide_profit=hide_profit)
        
        # Save targets for 2026
        targets_data = load_globals()
        if not targets_data:
            targets_data = {}
        targets_data["2026"] = targets
        save_globals(targets_data)
        
        return jsonify({"success": True})

@app.route("/api/tabs")
def api_tabs():
    username = session.get('username')
    from config import check_permission
    hidden_tabs, hidden_reports = load_hidden()
    
    _vis = []
    for t in TABS:
        if t["id"] in hidden_tabs: continue
        if not check_permission(username, t["id"]): continue
        
        allowed_reports = []
        for r in t["reports"]:
            if r["id"] in hidden_reports: continue
            if not r.get("hide_from_menu", False):
                if check_permission(username, t["id"], r["id"]):
                    allowed_reports.append({"id": r["id"], "title": r["title"]})
        
        if allowed_reports:
            _vis.append({"id": t["id"], "title": t["title"], "reports": allowed_reports})
            
    return jsonify({"tabs": _vis})

_DASH_CACHE = {}

@app.route("/api/dashboard")
def api_dashboard():
    try:
        from datetime import datetime
        import time
        d_from = request.args.get("date_from", datetime.now().strftime("%Y-01-01"))
        d_to = request.args.get("date_to", datetime.now().strftime("%Y-12-31"))
        force_refresh = request.args.get("force_refresh", "0")
        
        cache_key = f"{d_from}_{d_to}"
        current_time = time.time()
        
        if force_refresh != "1" and cache_key in _DASH_CACHE:
            if current_time - _DASH_CACHE[cache_key]['time'] < 300: # 5 minutes TTL
                dash_data = _DASH_CACHE[cache_key]['data']
            else:
                dash_data = compute_dash(d_from, d_to)
                _DASH_CACHE[cache_key] = {'time': current_time, 'data': dash_data}
        else:
            dash_data = compute_dash(d_from, d_to)
            _DASH_CACHE[cache_key] = {'time': current_time, 'data': dash_data}
            
        # Determine if profit should be hidden
        hide_profit = load_hide_profit()
        
        return jsonify({
            "data": dash_data,
            "hide_profit": hide_profit,
            "date_from": d_from,
            "date_to": d_to
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reports/<tab_id>/<report_id>")
def api_report_data(tab_id, report_id):
    username = session.get('username')
    # if not check_permission(username, tab_id, report_id):
    #     return jsonify({"error": "ط؛ظٹط± ظ…طµط±ط­ ظ„ظƒ ط¨ط¹ط±ط¶ ظ‡ط°ط§ ط§ظ„طھظ‚ط±ظٹط±."}), 403
        
    tab, rpt = find_report(tab_id, report_id)
    if not rpt:
        return jsonify({"error": "ط§ظ„طھظ‚ط±ظٹط± ط؛ظٹط± ظ…ظˆط¬ظˆط¯"}), 404
        
    try:
        # Prepare parameters for the frontend UI
        resolved_params = []
        for p in rpt.get('params', []):
            param_def = dict(p) # Copy
            if callable(param_def.get('get_default')):
                param_def['default'] = param_def['get_default']()
                del param_def['get_default']
            
            # Convert lookup texts to SearchableSelect options
            if param_def["name"] in ("rep_code", "c_code", "v_code", "i_code", "a_code", "cc_code", "grp_code"):
                print(f"DEBUG: Processing param {param_def['name']} with type {param_def.get('type')}")
                if "options" not in param_def or not param_def["options"]:
                    try:
                        print(f"DEBUG: Fetching lookups for {param_def['name']}")
                        l_items = lookups(param_def["name"])
                        print(f"DEBUG: Fetched {len(l_items)} items")
                        param_def["type"] = "select"
                        param_def["options"] = [["", "ط§ظ„ظƒظ„ / ط¨ط¯ظˆظ† طھط­ط¯ظٹط¯"]] + [[x, x] for x in l_items]
                    except Exception as e:
                        print("Lookup error:", e)

            resolved_params.append(param_def)
            
        binds = dict(request.args)
        
        cols, rows = run_report(rpt, request.args)
        return jsonify({"cols": cols, "rows": rows, "params": resolved_params, "binds": binds})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=False)

import os
from flask import send_from_directory

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react(path):
    if path.startswith('api/'):
        return 'API route not found', 404
    public_dir = os.path.join(os.path.dirname(__file__), 'public')
    if path and os.path.exists(os.path.join(public_dir, path)):
        return send_from_directory(public_dir, path)
    return send_from_directory(public_dir, 'index.html')

