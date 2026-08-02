# -*- coding: utf-8 -*-
import re

with open('app.py', 'r', encoding='utf-8') as f:
    text = f.read()

# 1. Update PRINT_PAGE logo
old_logo = '<svg class="logo" viewBox="0 0 230 60" xmlns="http://www.w3.org/2000/svg"><text x="6" y="46" font-family="Arial" font-weight="900" font-style="italic" font-size="48" fill="#4f46e5">SREEN</text><polygon points="60,4 43,33 56,33 47,56 78,22 63,22" fill="#4f46e5"/></svg>'
# Wait, it might have been partially patched in previous step, so let's use regex
text = re.sub(r'<svg class="logo".*?</svg>', '<h1 style="color:#4f46e5;font-weight:900;margin:0;font-size:26px">تقارير الأونكس الحديثة</h1>', text)
text = text.replace('LOGO = \'<h1 style="color:#4f46e5;font-weight:900;margin:0;font-size:26px">تقارير الأونكس الحديثة</h1>\'', 'LOGO = \'<div style="color:#4f46e5;font-weight:900;font-size:26px;letter-spacing:-1px">تقارير الأونكس الحديثة</div>\'')

# 2. Update Excel Export (fallback to CSV if openpyxl missing, and fix color if openpyxl exists)
old_export = '''def export():
    tab, rpt = find_report(request.args.get("tab", TABS[0]["id"]), request.args.get("report",""))
    try:
        cols, rows = run_report(rpt, request.args)
    except Exception as e:
        return "خطأ: " + str(e), 500
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "تقرير"
    ws.sheet_view.rightToLeft = True
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="12333C")
        cell.alignment = Alignment(horizontal="right")
    for r in rows:
        ws.append(list(r))
    for i in range(1, len(cols)+1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=%s.xlsx" % rpt["id"]})'''

new_export = '''def export():
    tab, rpt = find_report(request.args.get("tab", TABS[0]["id"]), request.args.get("report",""))
    try:
        cols, rows = run_report(rpt, request.args)
    except Exception as e:
        return "خطأ: " + str(e), 500
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws = wb.active; ws.title = "تقرير"
        ws.sheet_view.rightToLeft = True
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
            cell.alignment = Alignment(horizontal="right")
        for r in rows:
            ws.append([str(x) if x is not None else '' for x in r])
        for i in range(1, len(cols)+1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=%s.xlsx" % rpt["id"]})
    except ImportError:
        import csv
        buf = io.StringIO()
        buf.write('\\ufeff')
        writer = csv.writer(buf)
        writer.writerow(cols)
        for r in rows:
            writer.writerow([str(x) if x is not None else '' for x in r])
        return Response(buf.getvalue().encode('utf-8'),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=%s.csv" % rpt["id"]})'''
if "def export():" in text:
    text = re.sub(r'def export\(\):.*?return Response\(.*?filename=%s\.xlsx.*?\}\)', new_export.replace('\\', '\\\\'), text, flags=re.DOTALL)

# 3. Change Charts Concept
old_charts_html = '''<div class="gcharts">
          <div class="gc"><h3>المبيعات والتحصيل شهرياً</h3><canvas id="c1" height="150"></canvas></div>
          <div class="gc"><h3>أفضل المناديب (مبيعات)</h3><canvas id="c2" height="150"></canvas></div>
          <div class="gc"><h3>أفضل الأصناف (مبيعات)</h3><canvas id="c3" height="150"></canvas></div>
          <div class="gc"><h3>المشتريات شهرياً</h3><canvas id="c4" height="150"></canvas></div>
        </div>'''

new_charts_html = '''<div class="gcharts" style="grid-template-columns: 1fr 1fr 1fr;">
          <div class="gc" style="grid-column: span 2;"><h3>المبيعات والتحصيل شهرياً</h3><canvas id="c1" height="250"></canvas></div>
          <div class="gc" style="display:flex; flex-direction:column; align-items:center;"><h3>أفضل 5 مناديب</h3><div style="width:100%; max-width:250px; flex:1;"><canvas id="c2"></canvas></div></div>
          <div class="gc" style="display:flex; flex-direction:column; align-items:center;"><h3>أفضل 5 أصناف</h3><div style="width:100%; max-width:250px; flex:1;"><canvas id="c3"></canvas></div></div>
          <div class="gc" style="grid-column: span 2;"><h3>المشتريات شهرياً</h3><canvas id="c4" height="250"></canvas></div>
        </div>'''
text = text.replace(old_charts_html, new_charts_html)
# Also change it for the initial case where style="grid-template-columns: 1fr 1fr;"
text = re.sub(r'<div class="gcharts">.*?<div class="gc">.*?</div>\s*</div>', new_charts_html, text, flags=re.DOTALL)

new_script = '''     <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
     <script>
     var D={{ dash|tojson }};
     window.addEventListener("load",function(){ 
       if(!window.Chart) return; 
       Chart.defaults.font.family = "'Cairo', 'Inter', sans-serif";
       Chart.defaults.color = "#64748b";
       
       const commonOptions = {
         responsive: true,
         maintainAspectRatio: false,
         plugins: {
           legend: { display: false },
           tooltip: { backgroundColor: '#1e293b', padding: 14, titleFont: { size: 14, family: "'Cairo', sans-serif", weight: 'bold' }, bodyFont: { size: 14, family: "'Cairo', sans-serif" }, cornerRadius: 10, displayColors: true, boxPadding: 6 }
         }
       };
       
       // C1: Bar Chart (Sales & Collection)
       new Chart(document.getElementById("c1"),{
         type:"bar",
         data:{
           labels:D.months,
           datasets:[
             {label:"مبيعات", data:D.msales, backgroundColor:"#4f46e5", borderRadius:8, maxBarThickness: 32},
             {label:"تحصيل", data:D.mcollect, backgroundColor:"#38bdf8", borderRadius:8, maxBarThickness: 32}
           ]
         },
         options: {
           ...commonOptions,
           plugins: { ...commonOptions.plugins, legend: { display: true, position: 'top', align: 'end', labels: { usePointStyle: true, boxWidth: 10, font: { family: "'Cairo'", size: 13, weight: 'bold' } } } },
           scales: {
             x: { grid: { display: false }, border: { display: false } },
             y: { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false } }
           }
         }
       });

       // C2: Doughnut (Salesmen)
       new Chart(document.getElementById("c2"),{
         type:"doughnut",
         data:{
           labels:D.rep_labels.slice(0,5),
           datasets:[{data:D.rep_vals.slice(0,5), backgroundColor:["#4f46e5", "#38bdf8", "#10b981", "#f59e0b", "#8b5cf6"], borderWidth: 0, hoverOffset: 4}]
         },
         options: {
           responsive: true, maintainAspectRatio: false, cutout: '75%',
           plugins: { legend: { display: true, position: 'bottom', labels: { usePointStyle: true, boxWidth: 8, font: { size: 11 } } } }
         }
       });

       // C3: Doughnut (Items)
       new Chart(document.getElementById("c3"),{
         type:"doughnut",
         data:{
           labels:D.itm_labels.slice(0,5),
           datasets:[{data:D.itm_vals.slice(0,5), backgroundColor:["#f43f5e", "#d946ef", "#0ea5e9", "#14b8a6", "#eab308"], borderWidth: 0, hoverOffset: 4}]
         },
         options: {
           responsive: true, maintainAspectRatio: false, cutout: '75%',
           plugins: { legend: { display: true, position: 'bottom', labels: { usePointStyle: true, boxWidth: 8, font: { size: 11 } } } }
         }
       });

       // C4: Line Chart (Purchases)
       const ctx4 = document.getElementById("c4").getContext('2d');
       const grad4 = ctx4.createLinearGradient(0, 0, 0, 300);
       grad4.addColorStop(0, 'rgba(16, 185, 129, 0.4)');
       grad4.addColorStop(1, 'rgba(16, 185, 129, 0.0)');
       
       new Chart(ctx4,{
         type:"line",
         data:{
           labels:D.months,
           datasets:[{
             label: "مشتريات", data:D.mpurch, borderColor:"#10b981", borderWidth: 3, backgroundColor: grad4, fill:true, tension:0.4, pointRadius: 0, pointHoverRadius: 6, pointBackgroundColor: "#fff", pointBorderColor: "#10b981", pointBorderWidth: 2
           }]
         },
         options: {
           ...commonOptions,
           interaction: { mode: 'index', intersect: false },
           scales: {
             x: { grid: { display: false }, border: { display: false } },
             y: { grid: { color: '#f1f5f9', borderDash: [6, 4] }, border: { display: false } }
           }
         }
       });
     });
     </script>'''

text = re.sub(r'<script src="https://cdnjs\.cloudflare\.com.*?</script>', new_script, text, flags=re.DOTALL)

with open('app.py', 'w', encoding='utf-8') as f:
    f.write(text)

print("Final adjustments applied!")
